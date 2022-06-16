import os
import sys
import re
import datetime
import time
from io import StringIO
from datetime import date
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import gc
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from pytimekr import pytimekr
import pyodbc
import pandas as pd
import numpy as np
import openpyxl
from threading import Thread

class AddForm(QGroupBox):
    """Dialog 창에서 공통 조건 입력 UI를 만드는 클래스"""

    def __init__(self):
        SegmentLabel = QLabel('Segment :            ')
        SegmentLabel.setStyleSheet("color: white; font-weight : bold")
        self.SegmentBox1 = QLineEdit()
        self.SegmentBox2 = QLineEdit()
        self.SegmentBox3 = QLineEdit()
        self.SegmentBox4 = QLineEdit()
        self.SegmentBox5 = QLineEdit()
        self.SegmentBox1.setStyleSheet("background-color: white;")
        self.SegmentBox2.setStyleSheet("background-color: white;")
        self.SegmentBox3.setStyleSheet("background-color: white;")
        self.SegmentBox4.setStyleSheet("background-color: white;")
        self.SegmentBox5.setStyleSheet("background-color: white;")
        self.SegmentBox1.setPlaceholderText('※ Segment01')
        self.SegmentBox2.setPlaceholderText('※ Segment02')
        self.SegmentBox3.setPlaceholderText('※ Segment03')
        self.SegmentBox4.setPlaceholderText('※ Segment04')
        self.SegmentBox5.setPlaceholderText('※ Segment05')

        UserDefineLabel = QLabel('UserDefined :          ')
        UserDefineLabel.setStyleSheet("color: white; font-weight : bold")
        self.UserDefine1 = QLineEdit()
        self.UserDefine2 = QLineEdit()
        self.UserDefine3 = QLineEdit()
        self.UserDefine1.setStyleSheet("background-color: white;")
        self.UserDefine2.setStyleSheet("background-color: white;")
        self.UserDefine3.setStyleSheet("background-color: white;")
        self.UserDefine1.setPlaceholderText('※ UserDefined1')
        self.UserDefine2.setPlaceholderText('※ UserDefined2')
        self.UserDefine3.setPlaceholderText('※ UserDefined3')

        self.UserLabel = QLabel('전표입력자 :        ')
        self.UserLabel.setStyleSheet("color: white; font-weight : bold")
        self.User = QLineEdit()
        self.User.setStyleSheet("background-color: white;")
        self.User.setPlaceholderText('※ 전표입력자를 입력하세요')
        self.Acount = QTextEdit()
        self.Acount.setStyleSheet("background-color: white;")
        self.Acount.setPlaceholderText('※ 특정 계정코드를 입력하거나 위 트리에서 선택하세요')

        self.sourceLabel = QLabel('Source :        ')
        self.sourceLabel.setStyleSheet("color: white; font-weight : bold")
        self.source = QLineEdit()
        self.source.setStyleSheet("background-color: white;")
        self.source.setPlaceholderText('※ Source를 입력하세요')

        self.btnMid = QPushButton('Account Update')
        self.btnMid.setStyleSheet('color:white;  background-image : url(./bar.png)')
        fontm = self.btnMid.font()
        fontm.setBold(True)
        self.btnMid.setFont(fontm)

        self.sublayout1 = QGridLayout()
        self.sublayout1.addWidget(SegmentLabel, 0, 0)
        self.sublayout1.addWidget(self.SegmentBox1, 0, 1)
        self.sublayout1.addWidget(self.SegmentBox2, 0, 2)
        self.sublayout1.addWidget(self.SegmentBox3, 0, 3)
        self.sublayout1.addWidget(self.SegmentBox4, 0, 4)
        self.sublayout1.addWidget(self.SegmentBox5, 0, 5)

        self.sublayout1.addWidget(UserDefineLabel, 1, 0)
        self.sublayout1.addWidget(self.UserDefine1, 1, 1)
        self.sublayout1.addWidget(self.UserDefine2, 1, 2)
        self.sublayout1.addWidget(self.UserDefine3, 1, 3)


class Communicate(QObject):
    """ExtButtonClicked 함수에서 발생하는 시그널을 doneAction 함수로 보내는 함수"""

    def resource_path(self, relative_path):
        """PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    closeApp = pyqtSignal()
    closeApp2 = pyqtSignal(str)


class Form(QGroupBox):
    """CoA를 이용하여 계정 트리를 구축하는 클래스"""

    def resource_path(self, relative_path):
        """PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self, parent):
        super(Form, self).__init__(parent)

        grid = QGridLayout()
        qh = QHBoxLayout()

        self.setLayout(grid)

        self.btnSelect = QPushButton("Select All")
        self.btnSelect.resize(65, 22)
        self.btnSelect.clicked.connect(self.select_all)
        self.btnSelect.clicked.connect(self.get_selected_leaves)
        self.btnSelect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnSelect.font()
        font11.setBold(True)
        self.btnSelect.setFont(font11)

        self.btnUnselect = QPushButton("Unselect All")
        self.btnUnselect.resize(65, 22)
        self.btnUnselect.clicked.connect(self.unselect_all)
        self.btnUnselect.clicked.connect(self.get_selected_leaves)
        self.btnUnselect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnUnselect.font()
        font11.setBold(True)
        self.btnUnselect.setFont(font11)

        self.setStyleSheet('QGroupBox  {color: white; background-color: white}')

        self.tree = QTreeWidget(self)
        self.tree.setStyleSheet("border-style: outset; border-color : white; background-color:white;")

        headerItem = QTreeWidgetItem()
        item = QTreeWidgetItem()

        qh.addWidget(self.btnSelect)
        qh.addWidget(self.btnUnselect)

        grid.addLayout(qh, 0, 0)
        grid.addWidget(self.tree, 1, 0)

        self.tree.setHeaderHidden(True)
        self.tree.itemClicked.connect(self.get_selected_leaves)

    ### 계정 트리 Unselect All 클릭 시 적용
    def unselect_all(self):
        def recurse_unselect(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_unselect(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            grandchild.setCheckState(0, Qt.Unchecked)

        recurse_unselect(self.tree.invisibleRootItem())

    ### 계정 트리 Select All 클릭 시 적용
    def select_all(self):
        def recurse_select(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_select(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Unchecked:
                            grandchild.setCheckState(0, Qt.Checked)

        recurse_select(self.tree.invisibleRootItem())

    ### 계정 트리 선택된 값 가지고 오기
    def get_selected_leaves(self):
        checked_items = []

        def recurse(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            checked_items.append(grandchild.text(0).split(' ')[0])

        recurse(self.tree.invisibleRootItem())

        global checked_name
        checked_name = ''
        for i in checked_items:
            checked_name = checked_name + ',' + '\'' + i + '\''

        checked_name = checked_name[1:]


class Form1(QGroupBox):
    """CoA를 이용하여 계정 트리를 구축하는 클래스"""

    def resource_path(self, relative_path):
        """PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self, parent):
        super(Form1, self).__init__(parent)

        grid = QGridLayout()
        qh = QHBoxLayout()

        self.setLayout(grid)

        self.btnSelect = QPushButton("Select All")
        self.btnSelect.resize(65, 22)
        self.btnSelect.clicked.connect(self.select_all)
        self.btnSelect.clicked.connect(self.get_selected_leaves_1)
        self.btnSelect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnSelect.font()
        font11.setBold(True)
        self.btnSelect.setFont(font11)

        self.btnUnselect = QPushButton("Unselect All")
        self.btnUnselect.resize(65, 22)
        self.btnUnselect.clicked.connect(self.unselect_all)
        self.btnUnselect.clicked.connect(self.get_selected_leaves_1)
        self.btnUnselect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnUnselect.font()
        font11.setBold(True)
        self.btnUnselect.setFont(font11)

        self.setStyleSheet('QGroupBox  {color: white; background-color: white}')

        self.tree = QTreeWidget(self)
        self.tree.setStyleSheet("border-style: outset; border-color : white; background-color:white;")

        headerItem = QTreeWidgetItem()
        item = QTreeWidgetItem()

        qh.addWidget(self.btnSelect)
        qh.addWidget(self.btnUnselect)

        grid.addLayout(qh, 0, 0)
        grid.addWidget(self.tree, 1, 0)

        self.tree.setHeaderHidden(True)
        self.tree.itemClicked.connect(self.get_selected_leaves_1)

    ### 계정 트리 Unselect All 클릭 시 적용
    def unselect_all(self):
        def recurse_unselect(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_unselect(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            grandchild.setCheckState(0, Qt.Unchecked)

        recurse_unselect(self.tree.invisibleRootItem())

    ### 계정 트리 Select All 클릭 시 적용
    def select_all(self):
        def recurse_select(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_select(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Unchecked:
                            grandchild.setCheckState(0, Qt.Checked)

        recurse_select(self.tree.invisibleRootItem())

    ### 계정 트리 선택된 값 가지고 오기
    def get_selected_leaves_1(self):
        checked_items = []

        def recurse(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            checked_items.append(grandchild.text(0).split(' ')[0])

        recurse(self.tree.invisibleRootItem())

        global checked_name2
        checked_name2 = ''
        for i in checked_items:
            checked_name2 = checked_name2 + ',' + '\'' + i + '\''

        checked_name2 = checked_name2[1:]


class DataFrameModel(QAbstractTableModel):
    """추출된 데이터를 Main UI에 출력하는 클래스"""
    DtypeRole = Qt.UserRole + 1000
    ValueRole = Qt.UserRole + 1001

    def resource_path(self, relative_path):
        """PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self, df=pd.DataFrame(), parent=None):
        super(DataFrameModel, self).__init__(parent)
        self._dataframe = df

    def setDataFrame(self, dataframe):
        self.beginResetModel()
        self._dataframe = dataframe.copy()
        self.endResetModel()

    def dataFrame(self):
        return self._dataframe

    dataFrame = pyqtProperty(pd.DataFrame, fget=dataFrame, fset=setDataFrame)

    @pyqtSlot(int, Qt.Orientation, result=str)
    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._dataframe.columns[section]
            else:
                return str(self._dataframe.index[section])
        return QVariant()

    def rowCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        return len(self._dataframe.index)

    def columnCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        return self._dataframe.columns.size

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or not (0 <= index.row() < self.rowCount() and 0 <= index.column() < self.columnCount()):
            return QVariant()
        row = self._dataframe.index[index.row()]
        col = self._dataframe.columns[index.column()]
        dt = self._dataframe[col].dtype

        val = self._dataframe.iloc[row][col]
        if role == Qt.DisplayRole:
            return str(val)
        elif role == DataFrameModel.ValueRole:
            return val
        if role == DataFrameModel.DtypeRole:
            return dt
        return QVariant()

    def roleNames(self):
        roles = {
            Qt.DisplayRole: b'display',
            DataFrameModel.DtypeRole: b'dtype',
            DataFrameModel.ValueRole: b'value'
        }
        return roles


class MyApp(QWidget):
    # Resource
    def resource_path(self, relative_path):
        """PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수"""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self):
        """class 변수 초기화 및 Dialog별 시그널 생성 함수"""
        super().__init__()
        self.init_UI()

        ##Initialize Variables
        self.selected_project_id = None
        self.selected_server_name = "--서버 목록--"
        self.dataframe = None
        self.dataframe_refer = None
        self.cnxn = None
        self.my_query = None
        self.selected_scenario_subclass_index = 0
        self.scenario_dic = {}
        self.new_tree = None
        self.new_prep = None
        self.dateList = []
        self.string_date_list = []
        self.finalDate = []
        self.dialoglist = set()
        self.timerVar = QTimer()
        self.timerVar.setInterval(1000)
        self.timerVar.timeout.connect(self.printTime)

        ##다이얼로그별 시그널 생성
        self.communicate4 = Communicate()
        self.communicate4.closeApp.connect(self.doneAction4)
        self.communicate5 = Communicate()
        self.communicate5.closeApp.connect(self.doneAction5)
        self.communicate6 = Communicate()
        self.communicate6.closeApp.connect(self.doneAction6)
        self.communicate7 = Communicate()
        self.communicate7.closeApp.connect(self.doneAction7)
        self.communicate8 = Communicate()
        self.communicate8.closeApp.connect(self.doneAction8)
        self.communicate9 = Communicate()
        self.communicate9.closeApp.connect(self.doneAction9)
        self.communicate10 = Communicate()
        self.communicate10.closeApp.connect(self.doneAction10)
        self.communicate12 = Communicate()
        self.communicate12.closeApp.connect(self.doneAction12)
        self.communicate13 = Communicate()
        self.communicate13.closeApp.connect(self.doneAction13)
        self.communicate14 = Communicate()
        self.communicate14.closeApp.connect(self.doneAction14)
        self.communicate15 = Communicate()
        self.communicate15.closeApp.connect(self.doneAction15)
        self.communicate16 = Communicate()
        self.communicate16.closeApp.connect(self.doneAction16)
        self.communicate17 = Communicate()
        self.communicate17.closeApp.connect(self.doneAction17)
        self.communicateC = Communicate()
        self.communicateC.closeApp2.connect(self.doneActionC)

    def MessageBox_Open(self, text):
        """입력된 text를 표시하는 경고창을 만드는 함수"""
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Information)
        self.msg.setWindowTitle("Message")
        self.msg.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.msg.setText(text)
        self.msg.exec_()

    def MessageBox_Open2(self, text):
        """프로젝트 연결이 성공했음을 알리는 알림창을 만드는 함수"""
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Information)
        self.msg.setWindowTitle("Project Connected")
        self.msg.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.msg.setText(text)
        self.msg.exec_()

    def alertbox_open(self):
        """각 시나리오별 필수 입력값이 누락되었음을 알리는 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('필수 입력값 누락')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('필수 입력값이 누락되었습니다.')
        self.alt.exec_()

    def alertbox_open2(self, state):
        """필수 입력값의 데이터타입이 숫자가 아닌 경우 발생하는 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        txt = state
        self.alt.setWindowTitle('필수 입력값 타입 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText(txt + ' 값을 숫자로만 입력해주시기 바랍니다.')
        self.alt.exec_()

    def alertbox_open3(self):
        """최대 추출 라인수 50만 건을 초과한 데이터가 추출되었음을 알리는 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('최대 라인 수 초과 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('50만건 이상 추출되어 상위 1000건만 선출하였습니다.')
        self.alt.exec_()

    def alertbox_open4(self, state):
        """조건 입력값의 데이터타입 이상을 알리는 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        txt = state
        self.alt.setWindowTitle('입력값 타입 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText(txt)
        self.alt.exec_()

    def alertbox_open5(self):
        """중복된 시나리오 번호(시트명)임을 알리는 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('시트명 중복')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('이미 해당 시트명이 존재합니다.')
        self.alt.exec_()

    def alertbox_open6(self):
        """제외 키워드를 activate한 상태에서 제외 키워드를 입력하지 않을 경우의 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('제외 키워드 입력 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('제외할 키워드를 입력하세요.')
        self.alt.exec_()

    def alertbox_open13(self):
        """전기일과 입력일 간 차이가 70만 일 이상 차이나는 경우 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('필수 입력값 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))
        self.alt.setText('N일은 0이상 70만 미만의 정수로만 입력 바랍니다.')
        self.alt.exec_()

    def alertbox_open19(self):
        """날짜 입력 형식에서 이상이 있을 경우 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('날짜 입력 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('날짜를 yyyyMMdd 형식으로 입력해주세요.')
        self.alt.exec_()

    def alertbox_open20(self):
        """기능영역이 존재하지 않을 경우 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('기능영역 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('해당 프로젝트는 기능영역이 존재하지 않습니다.')
        self.alt.exec_()

    def alertbox_open21(self):
        """입력일과 전기일이 모두 선택되거나, 모두 선택되지 않을 경우 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('날짜 형식 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('입력일과 전기일 중 하나를 선택하시길 바랍니다.')
        self.alt.exec_()

    def alertbox_open22(self):
        """특정 계정코드 입력 조건에 이상이 있을 시 경고창 생성 함수"""
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('계정 입력 오류')
        self.alt.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))
        self.alt.setText('계정 코드 쿼리문을 확인하시길 바랍니다.')
        self.alt.exec_()

    def check_account(self, acc):
        """특정 계정코드 조건값에 이상이 있는지 확인하는 함수"""
        ## 예외 처리 - 콤마(,)가 정상적으로 입력되지 않은 경우
        if acc.strip() != '' and (acc.count(',') + 1) * 2 != acc.count("'"):
            self.alertbox_open22()
            return False

        sql = '''
                                SET NOCOUNT ON;
                                SELECT TOP 1 JournalEntries.GLAccountNumber
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                                WHERE 1=1 {Account}
         '''.format(field=self.selected_project_id, Account=acc)

        ## 예외처리 - 따옴표(')가 정상적으로 입력되지 않은 경우
        try:
            self.dataframe_check = pd.read_sql(sql, self.cnxn)
        except:
            self.alertbox_open22()
            return False

    def check_account2(self, acc1, acc2):
        """상대계정 시나리오에서 주계정(A)과 상대게정(B)의 특정 게정코드 입력값에 이상이 있는지 확인하는 함수"""

        ## 예외 처리 - 주계정과 상대계정에서 콤마(,)가 정상적으로 입력되지 않은 경우
        if acc1.strip() != '' and (acc1.count(',') + 1) * 2 != acc1.count("'"):
            self.alertbox_open22()
            return False

        elif acc2.strip() != '' and (acc2.count(',') + 1) * 2 != acc2.count("'"):
            self.alertbox_open22()
            return False

        sql1 = '''
                               SET NOCOUNT ON;
                               SELECT TOP 1 *
                               FROM (SELECT GLAccountNumber AS GL_Account_Number From [{field}_Import_CY_01].[dbo].[pbcJournalEntries]) AS LVL4
                               WHERE 1=1 {Account}
        '''.format(field=self.selected_project_id, Account=acc1)

        sql2 = '''
                               SET NOCOUNT ON;
                               SELECT TOP 1 *
                               FROM (SELECT GLAccountNumber AS Analysis_GL_Account_Number From [{field}_Import_CY_01].[dbo].[pbcJournalEntries]) AS LVL4
                               WHERE 1=1 {Account}
                '''.format(field=self.selected_project_id, Account=acc2)

        ## 예외 처리 - 주계정과 상대계정에서 따옴표표(')가 정상적으로 입력되지않은 경우
        try:
            pd.read_sql(sql1, self.cnxn)
            pd.read_sql(sql2, self.cnxn)
        except:
            self.alertbox_open22()
            return False

    def NewQueryConcat(self, Segment1, Segment2, Segment3, Segment4, Segment5, UserDefine1, UserDefine2, UserDefine3,
                       UserList1, SourceList1, Manual, Auto):
        """Segment01 ~ 05, UserDefined1 ~ 3, 전표입력자, 전표유형, 수자동 전표 등의 조건 입력값을 쿼리 조건문으로 변환하는 함수"""

        ## Segment01
        SplitedSegment1 = Segment1.text().split(',')
        SplitedSegment1List = []
        for a in SplitedSegment1:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.Segment01 LIKE N'" + a + "'"
                SplitedSegment1List.append(b)
        Segment1Clean = str(' OR '.join(SplitedSegment1List))

        ## Segment02
        SplitedSegment2 = Segment2.text().split(',')
        SplitedSegment2List = []
        for a in SplitedSegment2:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.Segment02 LIKE N'" + a + "'"
                SplitedSegment2List.append(b)
        Segment2Clean = str(' OR '.join(SplitedSegment2List))

        ## Segment03
        SplitedSegment3 = Segment3.text().split(',')
        SplitedSegment3List = []
        for a in SplitedSegment3:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.Segment03 LIKE N'" + a + "'"
                SplitedSegment3List.append(b)
        Segment3Clean = str(' OR '.join(SplitedSegment3List))

        ## Segment04
        SplitedSegment4 = Segment4.text().split(',')
        SplitedSegment4List = []
        for a in SplitedSegment4:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.Segment04 LIKE N'" + a + "'"
                SplitedSegment4List.append(b)
        Segment4Clean = str(' OR '.join(SplitedSegment4List))

        ## Segment05
        SplitedSegment5 = Segment5.text().split(',')
        SplitedSegment5List = []
        for a in SplitedSegment5:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.Segment05 LIKE N'" + a + "'"
                SplitedSegment5List.append(b)
        Segment5Clean = str(' OR '.join(SplitedSegment5List))

        ## UserDefined1
        SplitedUserDefine1 = UserDefine1.text().split(',')
        SplitedUserDefine1List = []
        for a in SplitedUserDefine1:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.UserDefined1 LIKE N'" + a + "'"
                SplitedUserDefine1List.append(b)
        UserDefine1Clean = str(' OR '.join(SplitedUserDefine1List))

        ## UserDefined2
        SplitedUserDefine2 = UserDefine2.text().split(',')
        SplitedUserDefine2List = []
        for a in SplitedUserDefine2:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.UserDefined2 LIKE N'" + a + "'"
                SplitedUserDefine2List.append(b)
        UserDefine2Clean = str(' OR '.join(SplitedUserDefine2List))

        ## UserDefined3
        SplitedUserDefine3 = UserDefine3.text().split(',')
        SplitedUserDefine3List = []
        for a in SplitedUserDefine3:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.UserDefined3 LIKE N'" + a + "'"
                SplitedUserDefine3List.append(b)
        UserDefine3Clean = str(' OR '.join(SplitedUserDefine3List))

        ## PerparerID(전표입력자)
        SplitedUserList1 = UserList1.text().split(',')
        SplitedUserList1List = []
        for a in SplitedUserList1:
            a = a.strip()
            if a == '':
                b = ""
            elif a.upper() == '[NULL]':
                b = "JournalEntries.PreparerID LIKE N'' OR JournalEntries.PreparerID LIKE N' ' OR JournalEntries.PreparerID IS NULL"
                SplitedUserList1List.append(b)
            else:
                b = "JournalEntries.PreparerID LIKE N'" + a + "'"
                SplitedUserList1List.append(b)
        UserList1Clean = str(' OR '.join(SplitedUserList1List))

        ## Source(전표유형)
        SplitedSourceList1 = SourceList1.text().split(',')
        SplitedSourceList1List = []
        for a in SplitedSourceList1:
            a = a.strip()
            if a == '':
                b = ""
            else:
                b = "JournalEntries.Source LIKE N'" + a + "'"
                SplitedSourceList1List.append(b)
        SourceList1Clean = str(' OR '.join(SplitedSourceList1List))

        ## 조건 병합
        ConcatSQLlist = [Segment1Clean, Segment2Clean, Segment3Clean, Segment4Clean, Segment5Clean, UserDefine1Clean,
                         UserDefine2Clean, UserDefine3Clean]
        ConcatSQLlistClean = []
        for i in ConcatSQLlist:
            if len(i) > 0:
                ConcatSQLlistClean.append("(" + i + ")")
        ConcatSQL = str(' AND '.join(ConcatSQLlistClean))

        ## 조건이 존재하는 칼럼을 Select 절에 추가
        ConcatSQL3List = []
        if Segment1.text() != '':
            ConcatSQL3List.append('JournalEntries.Segment01')
        if Segment2.text() != '':
            ConcatSQL3List.append('JournalEntries.Segment02')
        if Segment3.text() != '':
            ConcatSQL3List.append('JournalEntries.Segment03')
        if Segment4.text() != '':
            ConcatSQL3List.append('JournalEntries.Segment04')
        if Segment5.text() != '':
            ConcatSQL3List.append('JournalEntries.Segment05')
        if UserDefine1.text() != '':
            ConcatSQL3List.append('CONVERT(CHAR(10), CONVERT(DATE, JournalEntries.UserDefined1), 23) AS UserDefined1')
        if UserDefine2.text() != '':
            ConcatSQL3List.append('JournalEntries.UserDefined2')
        if UserDefine3.text() != '':
            ConcatSQL3List.append('JournalEntries.UserDefined3')

        if len(ConcatSQL3List) > 0:
            ConcatSQL3Clean = "," + str(','.join(ConcatSQL3List))
        else:
            ConcatSQL3Clean = ""

        if not ConcatSQL:
            ConcatSQL2 = ""
        else:
            ConcatSQL2 = "AND " + ConcatSQL

        if len(UserList1Clean) > 0:
            ConcatSQL2 = ConcatSQL2 + "AND (" + UserList1Clean + ")"
        else:
            ConcatSQL2 = ConcatSQL2

        if len(SourceList1Clean) > 0:
            ConcatSQL2 = ConcatSQL2 + "AND (" + SourceList1Clean + ")"
        else:
            ConcatSQL2 = ConcatSQL2

        ## 수/자동 전표
        AutoManual = ''

        if Manual.isChecked() and Auto.isChecked():
            AutoManual = ''
        elif Manual.isChecked():
            AutoManual = "AND Details.SystemManualIndicator = 'Manual' "
        elif Auto.isChecked():
            AutoManual = "AND Details.SystemManualIndicator = 'System' "

        return ConcatSQL2, ConcatSQL3Clean, AutoManual

    ### 사용자가 선택한 계정에 대하여 checked_account로 계정 조건문을 업데이트
    def AccountUpdate(self, AccountText):
        AccountText.setPlainText(checked_name)

    ### 상대계정 시나리오에서 사용자가 선택한 주계정 A에 대하여 checked_account_A로 계정 조건문을 업데이트
    def AccountUpdate_A(self, AccountText):
        AccountText.setPlainText(checked_name)

    ### 상대계정 시나리오에서 사용자가 선택한 상대계정 B에 대하여 checked_account_B로 계정 조건문을 업데이트
    def AccountUpdate_B(self, AccountText):
        AccountText.setPlainText(checked_name2)

    ### 메인 UI 디자인 설정
    def init_UI(self):
        image = QImage(self.resource_path('./dark_gray.png'))
        scaledImg = image.scaled(QSize(1000, 900))
        palette = QPalette()
        palette.setBrush(10, QBrush(scaledImg))
        self.setPalette(palette)

        ### 타이틀 바 설정
        pixmap = QPixmap(self.resource_path('./title.png'))
        lbl_img = QLabel()
        lbl_img.setPixmap(pixmap)
        lbl_img.setStyleSheet("background-color:#000000")

        ### 상하 방향으로 layout을 쌓을 수 있게끔 설정
        widget_layout = QBoxLayout(QBoxLayout.TopToBottom)
        self.splitter_layout = QSplitter(Qt.Vertical)

        self.splitter_layout.addWidget(lbl_img)
        self.splitter_layout.addWidget(self.Connect_ServerInfo_Group())
        self.splitter_layout.addWidget(self.Show_DataFrame_Group())
        self.splitter_layout.addWidget(self.Save_Buttons_Group())
        self.splitter_layout.setHandleWidth(0)
        self.splitter_layout.setStretchFactor(0, 3)
        self.splitter_layout.setStretchFactor(1, 2)
        self.splitter_layout.setStretchFactor(2, 4)
        self.splitter_layout.setStretchFactor(3, 1)

        widget_layout.addWidget(self.splitter_layout)
        self.setLayout(widget_layout)

        ### 좌측 상단 아이콘 설정
        self.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))
        self.setWindowTitle('Scenario')

        ### 창 크기 설정
        self.setGeometry(300, 100, 1000, 900)
        self.show()

    def connectButtonClicked(self):
        """Project connect 버튼 클릭시 SQL 서버와 프로그램을 연결하는 함수"""

        password = ''
        ecode = self.line_ecode.text().strip()
        ecode = "'" + ecode + "'"
        user = 'guest'
        server = self.selected_server_name
        db = 'master'

        # 예외처리 - 서버 선택
        if server == "--서버 목록--":
            self.MessageBox_Open("서버가 선택되어 있지 않습니다.")
            return

        server_path = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};uid={user};pwd={password};DATABASE={db};trusted_connection=yes"

        # 예외처리 - 접속 정보 오류
        try:
            self.cnxn = pyodbc.connect(server_path)
        except:
            QMessageBox.about(self, "Warning", "접속에 실패하였습니다.")
            return

        cursor = self.cnxn.cursor()

        sql_query = f"""
                           SELECT ProjectName
                           From [DataAnalyticsRepository].[dbo].[Projects]
                           WHERE EngagementCode IN ({ecode})
                           AND DeletedBy IS NULL
                     """
        ## 예외 처리 - project name을 찾는 SQL query에서 오류가 발생하는 경우
        try:
            selected_project_names = pd.read_sql(sql_query, self.cnxn)
        except:
            self.MessageBox_Open("Engagement Code를 입력하세요.")
            self.ProjectCombobox.clear()
            self.ProjectCombobox.addItem("프로젝트가 없습니다")
            return

        # 예외처리 - 입력된 ecode에 해당하는 프로젝트가 존재하지 않는 경우
        if len(selected_project_names) == 0:
            self.MessageBox_Open("해당 Engagement Code 내 프로젝트가 존재하지 않습니다.")
            self.ProjectCombobox.clear()
            self.ProjectCombobox.addItem("프로젝트가 없습니다.")
            return
        else:
            self.MessageBox_Open2("프로젝트가 연결되었습니다.")

        ## 서버 연결 시 - 기존 저장 정보를 초기화(메모리 관리)
        del self.selected_project_id, self.dataframe, self.scenario_dic, self.my_query

        ## 추출 시나리오 query 저장 변수
        self.my_query = pd.DataFrame(columns=["Sheet name", "Scenario number", "Query"])

        ## 해당 ecode의 프로젝트들을 combo box에 표시
        self.ProjectCombobox.clear()
        self.ProjectCombobox.addItem("--프로젝트 목록--")
        for i in range(len(selected_project_names)):
            self.ProjectCombobox.addItem(selected_project_names.iloc[i, 0])

        self.combo_sheet.clear()

        ## 서버 연결시 - 기존 저장 정보 초기화(메모리 관리)
        self.selected_project_id = None
        self.dataframe = None
        self.dataframe_refer = None
        self.viewtable.setModel(self.dataframe)
        self.scenario_dic = {}
        self.string_date_list = []
        self.finalDate = []
        self.clickCount = 0
        gc.collect()

    def Server_ComboBox_Selected(self, text):
        """콤보박스에서 선택된 SQL 서버를 class 변수에 담는 함수"""
        self.selected_server_name = text

    def Project_ComboBox_Selected(self, text):
        """콤보박스에서 선택된 프로젝트 정보를 class 변수에 담는 함수"""
        ## 예외처리 - 서버가 연결되지 않은 상태로 Projectname Combobox를 건드리는 경우
        if self.cnxn is None:
            return

        ecode = self.line_ecode.text().strip()  # leading/trailing space 제거
        ecode = "'" + ecode + "'"

        pname = text
        self.pname_year = "20" + str(pname)[2:4]
        cursor = self.cnxn.cursor()

        sql_query = f"""
                                SELECT Project_ID
                                FROM [DataAnalyticsRepository].[dbo].[Projects]
                                WHERE ProjectName IN (\'{pname}\')
                                AND EngagementCode IN ({ecode})
                                AND DeletedBy is Null
                             """

        ## 예외처리 - 에러 표시인 "프로젝트가 없습니다"가 콤보박스에서 선택되어 있는 경우
        try:
            self.selected_project_id = pd.read_sql(sql_query, self.cnxn).iloc[0, 0]

        except:
            self.selected_project_id = None

    def Connect_ServerInfo_Group(self):
        """SQL 서버 상의 프로젝트와 연결하기 위한 정보를 입력하는 UI"""

        groupbox = QGroupBox('접속 정보')
        self.setStyleSheet('QGroupBox  {color: white;}')
        font_groupbox = groupbox.font()
        font_groupbox.setBold(True)
        groupbox.setFont(font_groupbox)

        ##labels 생성 및 스타일 지정
        label1 = QLabel('Server : ', self)
        label2 = QLabel('Engagement Code : ', self)
        label3 = QLabel('Project Name : ', self)
        label4 = QLabel('Scenario : ', self)

        font1 = label1.font()
        font1.setBold(True)
        label1.setFont(font1)
        font2 = label2.font()
        font2.setBold(True)
        label2.setFont(font2)
        font3 = label3.font()
        font3.setBold(True)
        label3.setFont(font3)
        font4 = label4.font()
        font4.setBold(True)
        label4.setFont(font4)

        label1.setStyleSheet("color: white;")
        label2.setStyleSheet("color: white;")
        label3.setStyleSheet("color: white;")
        label4.setStyleSheet("color: white;")

        ##서버 선택 콤보박스
        self.cb_server = QComboBox(self)
        self.cb_server.addItem('--서버 목록--')
        for i in [1, 2, 3, 4, 6, 7, 8]:
            self.cb_server.addItem(f'KRSEOVMPPACSQ0{i}\INST1')

        ### Scenario 유형 콤보박스 - 소분류
        self.comboScenario = QComboBox(self)

        self.comboScenario.addItem('--시나리오 목록--')
        self.comboScenario.addItem('01 : 계정 사용빈도 N번 이하인 계정이 포함된 전표리스트')
        self.comboScenario.addItem('02 : 당기 생성된 계정리스트 추출')
        self.comboScenario.addItem('03 : 결산일 전후 T일 입력 전표')
        self.comboScenario.addItem('04 : 비영업일 전기/입력 전표')
        self.comboScenario.addItem('05 : 효력, 입력 일자 간 차이가 N일 이상인 전표')
        self.comboScenario.addItem('06 : 전표 작성 빈도수가 N회 이하인 작성자에 의한 생성된 전표')
        self.comboScenario.addItem('07 : 특정 전표 입력자(W)에 의해 생성된 전표')
        self.comboScenario.addItem('08 : 특정 계정(A) 상대계정 리스트 검토')
        self.comboScenario.addItem('09 : 연속된 숫자로 끝나는 금액 검토')
        self.comboScenario.addItem('10 : 전표 description에 공란 또는 특정단어(key word)가 입력되어 있는 전표 리스트')
        self.comboScenario.addItem('11 : 증빙일과 전기일의 회계기간이 다른 전표 추출 및 검토')
        self.comboScenario.addItem('12 : 차/대변 합계가 중요성금액 이상인 전표')
        self.comboScenario.addItem('13 : 전표 입력자와 승인자가 동일한 전표')

        self.ProjectCombobox = QComboBox(self)

        ##Engagement code 입력 line
        self.line_ecode = QLineEdit(self)
        self.line_ecode.setText("")

        ##Project Connect 버튼 생성 및 스타일 지정
        btn_connect = QPushButton('   Project Connect', self)
        font_btn_connect = btn_connect.font()
        font_btn_connect.setBold(True)
        btn_connect.setFont(font_btn_connect)
        btn_connect.setStyleSheet('color:white;  background-image : url(./bar.png)')

        ##Input Conditions 버튼 생성 및 스타일 지정
        btn_condition = QPushButton('   Input Conditions', self)
        font_btn_condition = btn_condition.font()
        font_btn_condition.setBold(True)
        btn_condition.setFont(font_btn_condition)
        btn_condition.setStyleSheet('color:white;  background-image : url(./bar.png)')

        ### Signal 함수들
        self.comboScenario.activated[str].connect(self.ComboSmall_Selected)
        self.cb_server.activated[str].connect(self.Server_ComboBox_Selected)
        btn_connect.clicked.connect(self.connectButtonClicked)
        btn_connect.setShortcut("Ctrl+P")  # remove sheet 업데이트 부분
        self.ProjectCombobox.activated[str].connect(self.Project_ComboBox_Selected)
        btn_condition.clicked.connect(self.connectDialog)

        ##layout 쌓기
        grid = QGridLayout()
        grid.addWidget(label1, 0, 0)
        grid.addWidget(label2, 1, 0)
        grid.addWidget(label3, 2, 0)
        grid.addWidget(label4, 3, 0)
        grid.addWidget(self.cb_server, 0, 1)
        grid.addWidget(btn_connect, 1, 2)
        grid.addWidget(self.comboScenario, 3, 1)
        grid.addWidget(btn_condition, 3, 2)
        grid.addWidget(self.line_ecode, 1, 1)
        grid.addWidget(self.ProjectCombobox, 2, 1)

        groupbox.setLayout(grid)
        return groupbox

    def ComboSmall_Selected(self, text):
        """선택된 시나리오 유형 정보를 class 변수에 저장하는 함수"""
        self.selected_scenario_subclass_index = self.comboScenario.currentIndex()

    def connectDialog(self):
        """입력 정보를 바탕으로 해당 시나리오의 Dialog 창으로 연결하는 함수"""

        ## 예외처리 - SQL 서버와 연결되지 않은 상태에서 Input conditions 버튼을 누른 경우
        if self.cnxn is None:
            self.MessageBox_Open("SQL 서버가 연결되어 있지 않습니다.")
            return

        ## 예외처리 - 프로젝트가 선택되어 있지 않은 경우
        elif self.selected_project_id is None:
            self.MessageBox_Open("프로젝트가 선택되지 않았습니다.")
            return

        ## 예외처리 - 시나리오 유형을 선택하지 않은 경우
        elif self.selected_scenario_subclass_index == 0:
            self.MessageBox_Open("시나리오가 선택되지 않았습니다.")
            return

        else:
            ## 예외처리 - 접근 권한이 없는 프로젝트를 선택한 경우
            try:
                cursor = self.cnxn.cursor()
                sql = '''
                                         SELECT 											
                                                *
                                         FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                                    '''.format(field=self.selected_project_id)
                accountsname = pd.read_sql(sql, self.cnxn)

                if self.selected_scenario_subclass_index == 1:
                    self.Dialog4()

                elif self.selected_scenario_subclass_index == 2:
                    self.Dialog5()

                elif self.selected_scenario_subclass_index == 3:
                    self.Dialog6()

                elif self.selected_scenario_subclass_index == 4:
                    self.Dialog7()

                elif self.selected_scenario_subclass_index == 5:
                    self.Dialog8()

                elif self.selected_scenario_subclass_index == 6:
                    self.Dialog9()

                elif self.selected_scenario_subclass_index == 7:
                    self.Dialog10()

                elif self.selected_scenario_subclass_index == 8:
                    self.Dialog12()

                elif self.selected_scenario_subclass_index == 9:
                    self.Dialog13()

                elif self.selected_scenario_subclass_index == 10:
                    self.Dialog14()

                elif self.selected_scenario_subclass_index == 11:
                    self.Dialog15()

                elif self.selected_scenario_subclass_index == 12:
                    self.Dialog16()

                elif self.selected_scenario_subclass_index == 13:
                    self.Dialog17()

                gc.collect()

            except:
                self.MessageBox_Open("접근 권한이 없는 프로젝트 입니다.")

    ### 시나리오 1. 계정 사용빈도 N번 이하인 계정이 포함된 전표리스트
    def Dialog4(self):
        self.Addnew4 = AddForm()
        self.Addnew4.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew4.Acount))

        ### 상단 라벨
        Titlelabel4 = QLabel('1. 계정 사용빈도 N번 이하인 계정이 포함된 전표리스트\n')
        Titlelabel4.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(4)
        self.dialog4 = QDialog()
        self.dialog4.setStyleSheet('background-color: #2E2E38')
        self.dialog4.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        ### 계정 트리
        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog4)
        self.btn2.setStyleSheet('color:white; background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread4)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 버튼 2 - Close
        self.btnDialog = QPushButton('   Close', self.dialog4)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close4)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog4)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog4)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 라벨 1 - 사용빈도
        label_freq = QLabel('사용 빈도(N)* :', self.dialog4)
        label_freq.setStyleSheet('color: yellow;')
        font1 = label_freq.font()
        font1.setBold(True)
        label_freq.setFont(font1)

        ### LineEdit 1 - 사용 빈도
        self.D4_N = QLineEdit(self.dialog4)
        self.D4_N.setStyleSheet('background-color: white;')
        self.D4_N.setPlaceholderText('사용빈도를 입력하세요')

        ### 라벨 2 - 중요성 금액
        label_TE = QLabel('중요성 금액: ', self.dialog4)
        label_TE.setStyleSheet('color: white;')
        font2 = label_TE.font()
        font2.setBold(True)
        label_TE.setFont(font2)

        ### LineEdit 2 - 중요성 금액
        self.D4_TE = QLineEdit(self.dialog4)
        self.D4_TE.setStyleSheet('background-color: white;')
        self.D4_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 라벨 3 - 시트명
        labelSheet = QLabel('시나리오 번호* : ', self.dialog4)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### LineEdit 3 - 시트명
        self.D4_Sheet = QLineEdit(self.dialog4)
        self.D4_Sheet.setStyleSheet("background-color: white;")
        self.D4_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### 계정 선택 라벨
        label_tree = QLabel('특정 계정명 : ', self.dialog4)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog4)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog4)
        self.checkD = QCheckBox('Debit', self.dialog4)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog4)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog4)
        self.Auto = QCheckBox('자동', self.dialog4)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### LineEdit만 창 크기에 따라 확대/축소
        self.D4_N.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D4_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D4_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog4)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount4)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D4_Sheet, 1, 1)
        layout1.addWidget(label_freq, 2, 0)
        layout1.addWidget(self.D4_N, 2, 1)
        layout1.addWidget(label_TE, 3, 0)
        layout1.addWidget(self.D4_TE, 3, 1)
        layout1.addWidget(label_tree, 4, 0)
        layout1.addWidget(self.new_tree, 4, 1)
        layout1.addWidget(self.Addnew4.btnMid, 5, 1)
        layout1.addWidget(self.Addnew4.Acount, 6, 1)
        layout1.addWidget(self.Addnew4.sourceLabel, 7, 0)
        layout1.addWidget(self.Addnew4.source, 7, 1)
        layout1.addWidget(self.Addnew4.UserLabel, 8, 0)
        layout1.addWidget(self.Addnew4.User, 8, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch(2)
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel4)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew4.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog4.setLayout(main_layout)
        self.dialog4.setGeometry(100, 100, 1000, 600)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog4.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog4.setWindowTitle('Scenario1')
        self.dialog4.setWindowModality(Qt.NonModal)
        self.dialog4.show()

    ### 시나리오 2. 당기 생성된 계정리스트 추출
    def Dialog5(self):
        self.Addnew5 = AddForm()
        self.Addnew5.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew5.Acount))

        ### 상단 라벨
        Titlelabel5 = QLabel('2. 당기 생성된 계정리스트 추출\n')
        Titlelabel5.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(5)
        self.dialog5 = QDialog()
        self.dialog5.setStyleSheet('background-color: #2E2E38')
        self.dialog5.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog5)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog5)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog5)
        self.btn2.setStyleSheet('color:white; background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread5)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.resize(110, 30)

        ### 버튼 2 - Close (Non-SAP)
        self.btnDialog = QPushButton('Close', self.dialog5)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close5)
        font11 = self.btnDialog.font()
        font11.setBold(True)
        self.btnDialog.setFont(font11)
        self.btnDialog.resize(110, 30)

        ### 계정 트리
        cursor2 = self.cnxn.cursor()
        sql2 = '''
                         SELECT
                                *
                         FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA
                    '''.format(field=self.selected_project_id)
        accountsname2 = pd.read_sql(sql2, self.cnxn)

        self.new_tree2 = Form(self)
        self.new_tree2.tree.clear()
        accountType2 = accountsname2.AccountType.unique()
        accountType2.sort()
        for n, i in enumerate(accountType2):
            self.new_tree2.parent = QTreeWidgetItem(self.new_tree2.tree)

            self.new_tree2.parent.setText(0, "{}".format(i))
            self.new_tree2.parent.setFlags(self.new_tree2.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname2.AccountSubType[
                accountsname2.AccountType == accountType2[n]].unique()
            child_items.sort()
            for m, x in enumerate(child_items):
                self.new_tree2.child = QTreeWidgetItem(self.new_tree2.parent)

                self.new_tree2.child.setText(0, "{}".format(x))
                self.new_tree2.child.setFlags(self.new_tree2.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname2.AccountClass[accountsname2.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree2.grandchild = QTreeWidgetItem(self.new_tree2.child)

                    self.new_tree2.grandchild.setText(0, "{}".format(y))
                    self.new_tree2.grandchild.setFlags(
                        self.new_tree2.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname2[accountsname2.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree2.grandgrandchild = QTreeWidgetItem(self.new_tree2.grandchild)

                        self.new_tree2.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree2.grandgrandchild.setFlags(
                            self.new_tree2.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree2.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree2.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog5)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog5)
        self.checkD = QCheckBox('Debit', self.dialog5)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog5)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog5)
        self.Auto = QCheckBox('자동', self.dialog5)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### 라벨 2 - 시트명
        labelSheet = QLabel('시나리오 번호* : ', self.dialog5)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D5_Sheet = QLineEdit(self.dialog5)
        self.D5_Sheet.setStyleSheet("background-color: white;")
        self.D5_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### 라벨 3 - 계정 트리
        label_tree = QLabel('특정 계정명* : ', self.dialog5)
        label_tree.setStyleSheet("color: yellow;")
        font40 = label_tree.font()
        font40.setBold(True)
        label_tree.setFont(font40)

        ### 중요성 금액
        label_TE = QLabel('중요성 금액 : ', self.dialog5)
        label_TE.setStyleSheet("color: white;")
        font5 = label_TE.font()
        font5.setBold(True)
        label_TE.setFont(font5)
        self.D5_TE = QLineEdit(self.dialog5)
        self.D5_TE.setStyleSheet("background-color: white;")
        self.D5_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog5)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount5)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D5_Sheet, 1, 1)
        layout1.addWidget(label_tree, 2, 0)
        layout1.addWidget(self.new_tree2, 2, 1)
        layout1.addWidget(self.Addnew5.btnMid, 3, 1)
        layout1.addWidget(self.Addnew5.Acount, 4, 1)
        layout1.addWidget(label_TE, 5, 0)
        layout1.addWidget(self.D5_TE, 5, 1)
        layout1.addWidget(self.Addnew5.sourceLabel, 6, 0)
        layout1.addWidget(self.Addnew5.source, 6, 1)
        layout1.addWidget(self.Addnew5.UserLabel, 7, 0)
        layout1.addWidget(self.Addnew5.User, 7, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch(2)
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel5)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew5.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog5.setLayout(main_layout)
        self.dialog5.setGeometry(100, 100, 1000, 600)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog5.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog5.setWindowTitle('Scenario2')
        self.dialog5.setWindowModality(Qt.NonModal)
        self.dialog5.show()

    ### 시나리오 3. 결산일 전후 T일 입력 전표
    def Dialog6(self):
        self.Addnew6 = AddForm()
        self.Addnew6.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew6.Acount))

        ### 상단 라벨
        Titlelabel6 = QLabel('3. 결산일 전후 T일 입력 전표\n')
        Titlelabel6.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(6)
        self.dialog6 = QDialog()
        self.dialog6.setStyleSheet('background-color: #2E2E38')
        self.dialog6.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)
        accountsname = pd.read_sql(sql, self.cnxn)
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)

        self.new_tree.get_selected_leaves()

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog6)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread6)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton("   Close", self.dialog6)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close6)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)
        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog6)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog6)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog6)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog6)
        self.checkD = QCheckBox('Debit', self.dialog6)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog6)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog6)
        self.Auto = QCheckBox('자동', self.dialog6)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### 입력일
        label_p = QLabel('입력일 *     :            ', self.dialog6)
        label_p.setStyleSheet("color: yellow;")
        font11 = label_p.font()
        font11.setBold(True)
        label_p.setFont(font11)
        self.period1 = QLineEdit(self.dialog6)
        self.period1.setStyleSheet("background-color: white;")
        self.period1.setPlaceholderText('시작 시점을 입력하세요 yyyyMMdd')
        self.period2 = QLineEdit(self.dialog6)
        self.period2.setStyleSheet("background-color: white;")
        self.period2.setPlaceholderText('종료 시점을 입력하세요 yyyyMMdd')

        ### 계정 선택 라벨
        label_tree = QLabel('특정 계정명 : ', self.dialog6)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 중요성 금액
        label_TE = QLabel('중요성 금액 :           ', self.dialog6)
        label_TE.setStyleSheet("color: white;")
        font5 = label_TE.font()
        font5.setBold(True)
        label_TE.setFont(font5)
        self.D6_TE = QLineEdit(self.dialog6)
        self.D6_TE.setStyleSheet("background-color: white;")
        self.D6_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* :     ', self.dialog6)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D6_Sheet = QLineEdit(self.dialog6)
        self.D6_Sheet.setStyleSheet("background-color: white;")
        self.D6_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### UI를 맞추기 위한 비어있는 LineEdit
        temp_lineedit = QLineEdit(self.dialog6)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)

        ### LineEdit만 창 크기에 따라 확대/축소
        self.period1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.period2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D6_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D6_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog6)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount6)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### 입력일 Layout
        layout4 = QHBoxLayout()
        layout4.addWidget(label_p)
        layout4.addWidget(self.period1)
        layout4.addWidget(self.period2)

        ### 최상단 Layout
        layout0 = QGridLayout()
        layout0.addWidget(self.rbtn1, 0, 0)
        layout0.addWidget(self.rbtn2, 0, 1)
        layout0.addWidget(temp_lineedit, 0, 2)

        ### 시나리오 번호 입력 Layout
        layout1 = QGridLayout()
        layout1.addWidget(labelSheet, 0, 0)
        layout1.addWidget(self.D6_Sheet, 0, 1)

        ### 중간 Layout
        layout2 = QGridLayout()
        layout2.addWidget(label_TE, 0, 0)
        layout2.addWidget(self.D6_TE, 0, 1)
        layout2.addWidget(label_tree, 1, 0)
        layout2.addWidget(self.new_tree, 1, 1)
        layout2.addWidget(self.Addnew6.btnMid, 2, 1)
        layout2.addWidget(self.Addnew6.Acount, 3, 1)
        layout2.addWidget(self.Addnew6.sourceLabel, 4, 0)
        layout2.addWidget(self.Addnew6.source, 4, 1)
        layout2.addWidget(self.Addnew6.UserLabel, 5, 0)
        layout2.addWidget(self.Addnew6.User, 5, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout_btn = QHBoxLayout()
        layout_btn.addStretch()
        layout_btn.addStretch()
        layout_btn.addWidget(self.btnCount)
        layout_btn.addWidget(self.btn2)
        layout_btn.addWidget(self.btnDialog)
        layout_btn.setContentsMargins(-1, 10, -1, -1)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel6)
        main_layout.addLayout(layout0)
        main_layout.addLayout(layout1)
        main_layout.addLayout(layout4)
        main_layout.addLayout(layout2)
        main_layout.addLayout(self.Addnew6.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout_btn)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog6.setLayout(main_layout)
        self.dialog6.setGeometry(100, 100, 1000, 600)
        self.dialog6.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog6.setWindowTitle("Scenario3")
        self.dialog6.setWindowModality(Qt.NonModal)
        self.dialog6.show()

    ### 시나리오 4. 비영업일 전기/입력 전표
    def Dialog7(self):
        self.Addnew7 = AddForm()
        self.Addnew7.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew7.Acount))

        ### 상단 라벨
        Titlelabel7 = QLabel('4. 비영업일 전기/입력 전표\n')
        Titlelabel7.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(7)
        self.dialog7 = QDialog()
        self.dialog7.setStyleSheet('background-color: #2E2E38')
        self.dialog7.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()

        ### 차변 / 대변 체크 박스
        self.checkC = QCheckBox('Credit', self.dialog7)
        self.checkD = QCheckBox('Debit', self.dialog7)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        labelDC = QLabel('차변/대변 : ', self.dialog7)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog7)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread7)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.resize(110, 30)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton("   Close", self.dialog7)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close7)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)
        self.btnDialog.resize(110, 30)

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog7)
        self.rbtn1.setChecked(True)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog7)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 비영업일 입력
        labelDate = QLabel('비영업일 입력 :        ', self.dialog7)
        labelDate.setStyleSheet("color: white;")
        font3 = labelDate.font()
        font3.setBold(True)
        labelDate.setFont(font3)
        self.D7_Date = QTextEdit(self.dialog7)
        self.D7_Date.setStyleSheet("background-color: white;")
        self.D7_Date.setPlaceholderText('날짜를 추가해주세요 (법정 공휴일 및 주말은 포함되어 있습니다) \nex) 대체공휴일, 창립기념일, 근로자의 날')

        ### 특정 계정명
        label_tree = QLabel('특정 계정명 : ', self.dialog7)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 중요성 금액
        label_TE = QLabel('중요성 금액 : ', self.dialog7)
        label_TE.setStyleSheet("color: white;")
        font6 = label_TE.font()
        font6.setBold(True)
        label_TE.setFont(font6)
        self.D7_TE = QLineEdit(self.dialog7)
        self.D7_TE.setStyleSheet("background-color: white;")
        self.D7_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* : ', self.dialog7)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D7_Sheet = QLineEdit(self.dialog7)
        self.D7_Sheet.setStyleSheet("background-color: white;")
        self.D7_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### LineEdit만 창 크기에 따라 확대 / 축소
        self.D7_Date.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D7_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D7_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog7)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog7)
        self.Auto = QCheckBox('자동', self.dialog7)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### 전기일 / 입력일 선택 박스
        labelEntef = QLabel('전기일/입력일* : ', self.dialog7)
        labelEntef.setStyleSheet("color: yellow; font-weight : bold")
        self.Entry = QCheckBox('입력일', self.dialog7)
        self.Entry.setStyleSheet("color: white;")
        self.Effective = QCheckBox('전기일', self.dialog7)
        self.Effective.setStyleSheet("color: white;")

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog7)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount7)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D7_Sheet, 1, 1)

        ### 입력일 / 전기일 Layout
        layout2 = QHBoxLayout()
        layout2.addWidget(labelEntef)
        layout2.addWidget(self.Effective)
        layout2.addWidget(self.Entry)

        ### 중간 Layout
        layout3 = QGridLayout()
        layout3.addWidget(labelDate, 0, 0)
        layout3.addWidget(self.D7_Date, 0, 1)
        layout3.addWidget(label_TE, 1, 0)
        layout3.addWidget(self.D7_TE, 1, 1)
        layout3.addWidget(label_tree, 2, 0)
        layout3.addWidget(self.new_tree, 2, 1)
        layout3.addWidget(self.Addnew7.btnMid, 3, 1)
        layout3.addWidget(self.Addnew7.Acount, 4, 1)
        layout3.addWidget(self.Addnew7.sourceLabel, 5, 0)
        layout3.addWidget(self.Addnew7.source, 5, 1)
        layout3.addWidget(self.Addnew7.UserLabel, 6, 0)
        layout3.addWidget(self.Addnew7.User, 6, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout4 = QHBoxLayout()
        layout4.addStretch()
        layout4.addStretch()
        layout4.addWidget(self.btnCount)
        layout4.addWidget(self.btn2)
        layout4.addWidget(self.btnDialog)
        layout4.setContentsMargins(-1, 10, -1, -1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel7)
        main_layout.addLayout(layout1)
        main_layout.addLayout(layout2)
        main_layout.addLayout(layout3)
        main_layout.addLayout(self.Addnew7.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout4)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog7.setLayout(main_layout)
        self.dialog7.setGeometry(100, 100, 1000, 600)
        self.dialog7.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog7.setWindowTitle("Scenario4")
        self.dialog7.setWindowModality(Qt.NonModal)
        self.dialog7.show()

    ### 시나리오 5. 효력, 입력 일자 간 차이가 N일 이상인 전표
    def Dialog8(self):
        self.Addnew8 = AddForm()
        self.Addnew8.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew8.Acount))

        ### 상단 라벨
        Titlelabel8 = QLabel('5. 효력, 입력 일자 간 차이가 N일 이상인 전표\n')
        Titlelabel8.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(8)
        self.dialog8 = QDialog()
        self.dialog8.setStyleSheet('background-color: #2E2E38')
        self.dialog8.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()

        ### 차변 / 대변 체크 박스
        self.checkC = QCheckBox('Credit', self.dialog8)
        self.checkD = QCheckBox('Debit', self.dialog8)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        labelDC = QLabel('차변/대변 : ', self.dialog8)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog8)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog8)
        self.Auto = QCheckBox('자동', self.dialog8)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog8)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread8)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.resize(110, 30)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton("   Close", self.dialog8)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close8)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)
        self.btnDialog.resize(110, 30)

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog8)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog8)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### N일 입력
        labelDate = QLabel('N일* : ', self.dialog8)
        labelDate.setStyleSheet("color: yellow;")
        font1 = labelDate.font()
        font1.setBold(True)
        labelDate.setFont(font1)
        self.D8_N = QLineEdit(self.dialog8)
        self.D8_N.setStyleSheet("background-color: white;")
        self.D8_N.setPlaceholderText('N 값을 입력하세요')

        ### 특정 계정명 라벨
        label_tree = QLabel('특정 계정명 : ', self.dialog8)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 중요성 금액
        label_TE = QLabel('중요성금액 : ', self.dialog8)
        label_TE.setStyleSheet("color: white;")
        font4 = label_TE.font()
        font4.setBold(True)
        label_TE.setFont(font4)
        self.D8_TE = QLineEdit(self.dialog8)
        self.D8_TE.setStyleSheet("background-color: white;")
        self.D8_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* :      ', self.dialog8)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D8_Sheet = QLineEdit(self.dialog8)
        self.D8_Sheet.setStyleSheet("background-color: white;")
        self.D8_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### LineEdit만 창 크기에 따라 확대 / 축소
        self.D8_N.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D8_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D8_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ### UI를 맞추기 위한 비어있는 LineEdit
        temp_lineedit = QLineEdit(self.dialog8)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog8)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount8)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### 최상단 Layout
        layout0 = QGridLayout()
        layout0.addWidget(self.rbtn1, 0, 0)
        layout0.addWidget(self.rbtn2, 0, 1)
        layout0.addWidget(temp_lineedit, 0, 2)

        ### 중간 Layout
        layout1 = QGridLayout()
        layout1.addWidget(labelSheet, 0, 0)
        layout1.addWidget(self.D8_Sheet, 0, 1)
        layout1.addWidget(labelDate, 1, 0)
        layout1.addWidget(self.D8_N, 1, 1)
        layout1.addWidget(label_TE, 2, 0)
        layout1.addWidget(self.D8_TE, 2, 1)
        layout1.addWidget(label_tree, 3, 0)
        layout1.addWidget(self.new_tree, 3, 1)
        layout1.addWidget(self.Addnew8.btnMid, 4, 1)
        layout1.addWidget(self.Addnew8.Acount, 5, 1)
        layout1.addWidget(self.Addnew8.sourceLabel, 6, 0)
        layout1.addWidget(self.Addnew8.source, 6, 1)
        layout1.addWidget(self.Addnew8.UserLabel, 7, 0)
        layout1.addWidget(self.Addnew8.User, 7, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel8)
        main_layout.addLayout(layout0)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew8.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog8.setLayout(main_layout)
        self.dialog8.setGeometry(100, 100, 1000, 600)
        self.dialog8.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog8.setWindowTitle("Scenario5")
        self.dialog8.setWindowModality(Qt.NonModal)
        self.dialog8.show()

    ### 시나리오 6. 전표 작성 빈도수가 N회 이하인 작성자에 의한 생성된 전표
    def Dialog9(self):
        self.Addnew9 = AddForm()
        self.Addnew9.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew9.Acount))

        ### 상단 라벨
        Titlelabel9 = QLabel('6. 전표 작성 빈도수가 N회 이하인 작성자에 의한 생성된 전표\n')
        Titlelabel9.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가
        self.dialoglist.add(9)
        self.dialog9 = QDialog()
        groupbox = QGroupBox('접속 정보')

        ### 계정 트리
        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 디자인 설정
        self.dialog9.setStyleSheet('background-color: #2E2E38')
        self.dialog9.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog9)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread9)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton("  Close", self.dialog9)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close9)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        ### JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog9)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog9)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 작성빈도수
        label_N = QLabel('작성빈도수* : ', self.dialog9)
        label_N.setStyleSheet("color: yellow;")

        font1 = label_N.font()
        font1.setBold(True)
        label_N.setFont(font1)

        self.D9_N = QLineEdit(self.dialog9)
        self.D9_N.setStyleSheet("background-color: white;")
        self.D9_N.setPlaceholderText('작성 빈도수를 입력하세요')

        ###중요성 금액
        labelD9_TE = QLabel('중요성 금액 : ', self.dialog9)
        labelD9_TE.setStyleSheet("color: white;")

        font2 = labelD9_TE.font()
        font2.setBold(True)
        labelD9_TE.setFont(font2)

        self.D9_TE = QLineEdit(self.dialog9)
        self.D9_TE.setStyleSheet("background-color: white;")
        self.D9_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ###데이터 추출 버튼 & 창 닫기 버튼 사이즈 조절
        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        ### 특정 계정명 라벨
        label_tree = QLabel('특정 계정명 : ', self.dialog9)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* : ', self.dialog9)
        labelSheet.setStyleSheet("color: yellow;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D9_Sheet = QLineEdit(self.dialog9)
        self.D9_Sheet.setStyleSheet("background-color: white;")
        self.D9_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### 차변 / 대변 라벨
        labelDC = QLabel('차변/대변 : ', self.dialog9)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        ### 차변 / 대변 체크 박스
        self.checkC = QCheckBox('Credit', self.dialog9)
        self.checkD = QCheckBox('Debit', self.dialog9)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog9)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog9)
        self.Auto = QCheckBox('자동', self.dialog9)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        self.D9_N.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D9_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D9_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog9)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount9)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 중간 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D9_Sheet, 1, 1)
        layout1.addWidget(label_N, 2, 0)
        layout1.addWidget(self.D9_N, 2, 1)
        layout1.addWidget(labelD9_TE, 3, 0)
        layout1.addWidget(self.D9_TE, 3, 1)
        layout1.addWidget(label_tree, 4, 0)
        layout1.addWidget(self.new_tree, 4, 1)
        layout1.addWidget(self.Addnew9.btnMid, 5, 1)
        layout1.addWidget(self.Addnew9.Acount, 6, 1)
        layout1.addWidget(self.Addnew9.sourceLabel, 7, 0)
        layout1.addWidget(self.Addnew9.source, 7, 1)
        layout1.addWidget(self.Addnew9.UserLabel, 8, 0)
        layout1.addWidget(self.Addnew9.User, 8, 1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)

        layout2.setContentsMargins(-1, 10, -1, -1)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel9)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew9.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog9.setLayout(main_layout)
        self.dialog9.setGeometry(100, 100, 1000, 600)

        self.dialog9.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog9.setWindowTitle("Scenario6")
        self.dialog9.setWindowModality(Qt.NonModal)
        self.dialog9.show()

    ### 시나리오 7. 특정 전표입력자(W)에 의해 생성된 전표
    def Dialog10(self):
        self.Addnew10 = AddForm()
        self.Addnew10.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew10.Acount))

        ### 상단 라벨
        Titlelabel10 = QLabel('7. 특정 전표입력자(W)에 의해 생성된 전표\n')
        Titlelabel10.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(10)
        self.dialog10 = QDialog()
        self.dialog10.setStyleSheet('background-color: #2E2E38')
        self.dialog10.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))

        ### 계정 트리
        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog10)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog10)
        self.checkD = QCheckBox('Debit', self.dialog10)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog10)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog10)
        self.Auto = QCheckBox('자동', self.dialog10)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog10)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread10)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.resize(110, 30)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton("   Close", self.dialog10)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close10)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)
        self.btnDialog.resize(110, 30)

        ### JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog10)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog10)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 특정 전표입력자 입력
        labelKeyword = QLabel('전표입력자* : ', self.dialog10)
        labelKeyword.setStyleSheet("color: yellow;")
        font1 = labelKeyword.font()
        font1.setBold(True)
        labelKeyword.setFont(font1)

        ### 계정 선택 라벨
        label_tree = QLabel('특정 계정명 : ', self.dialog10)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 중요성 금액
        labelTE = QLabel('중요성 금액 : ', self.dialog10)
        labelTE.setStyleSheet("color: white;")
        font4 = labelTE.font()
        font4.setBold(True)
        labelTE.setFont(font4)
        self.D10_TE = QLineEdit(self.dialog10)
        self.D10_TE.setStyleSheet("background-color: white;")
        self.D10_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* : ', self.dialog10)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D10_Sheet = QLineEdit(self.dialog10)
        self.D10_Sheet.setStyleSheet("background-color: white;")
        self.D10_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### LineEdit만 창 크기에 따라 확대/축소
        self.D10_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D10_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog10)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount10)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D10_Sheet, 1, 1)
        layout1.addWidget(labelKeyword, 2, 0)
        layout1.addWidget(self.Addnew10.User, 2, 1)
        layout1.addWidget(labelTE, 3, 0)
        layout1.addWidget(self.D10_TE, 3, 1)
        layout1.addWidget(label_tree, 4, 0)
        layout1.addWidget(self.new_tree, 4, 1)
        layout1.addWidget(self.Addnew10.btnMid, 5, 1)
        layout1.addWidget(self.Addnew10.Acount, 6, 1)
        layout1.addWidget(self.Addnew10.sourceLabel, 7, 0)
        layout1.addWidget(self.Addnew10.source, 7, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel10)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew10.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog10.setLayout(main_layout)
        self.dialog10.setGeometry(100, 100, 1000, 600)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog10.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog10.setWindowTitle("Scenario7")
        self.dialog10.setWindowModality(Qt.NonModal)
        self.dialog10.show()

    ### 시나리오 8. 특정 계정(A)에 대한 상대계정 검토
    def Dialog12(self):
        self.Addnew12_A = AddForm()
        self.Addnew12_B = AddForm()
        self.Addnew12_A.btnMid.clicked.connect(lambda: self.AccountUpdate_A(self.Addnew12_A.Acount))
        self.Addnew12_B.btnMid.clicked.connect(lambda: self.AccountUpdate_B(self.Addnew12_B.Acount))

        ### 상단 라벨(8)
        TitlelabelMain = QLabel('8. 특정 계정(A)에 대한 상대계정 검토\n')
        TitlelabelMain.setStyleSheet("color: white; font-weight : bold")

        self.dialoglist.add(12)
        self.dialog12 = QDialog()
        self.dialog12.setStyleSheet('background-color: #2E2E38')
        self.dialog12.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        ### 상단 라벨(8-1)
        Titlelabel11 = QLabel('A계정의 상대계정이 B계정이 아닌 상대계정 리스트 추출\n')
        Titlelabel11.setStyleSheet("color: white; font-weight : bold")

        ### 계정 트리
        cursor1 = self.cnxn.cursor()
        sql1 = '''
                                 SELECT 											
                                        *
                                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                            '''.format(field=self.selected_project_id)

        accountsname1 = pd.read_sql(sql1, self.cnxn)

        ### 계정트리 - A, B
        self.new_tree1 = Form(self)
        self.new_tree2 = Form1(self)
        self.new_tree1.tree.clear()
        self.new_tree2.tree.clear()

        accountType1 = accountsname1.AccountType.unique()
        accountType1.sort()
        for n, i in enumerate(accountType1):
            self.new_tree1.parent = QTreeWidgetItem(self.new_tree1.tree)

            self.new_tree1.parent.setText(0, "{}".format(i))
            self.new_tree1.parent.setFlags(self.new_tree1.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)

            child_items1 = accountsname1.AccountSubType[accountsname1.AccountType == accountType1[n]].unique()
            child_items1.sort()

            for m, x in enumerate(child_items1):
                self.new_tree1.child = QTreeWidgetItem(self.new_tree1.parent)

                self.new_tree1.child.setText(0, "{}".format(x))
                self.new_tree1.child.setFlags(self.new_tree1.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)

                grandchild_items1 = accountsname1.AccountClass[accountsname1.AccountSubType == child_items1[m]].unique()
                grandchild_items1.sort()

                for o, y in enumerate(grandchild_items1):
                    self.new_tree1.grandchild = QTreeWidgetItem(self.new_tree1.child)
                    self.new_tree1.grandchild.setText(0, "{}".format(y))
                    self.new_tree1.grandchild.setFlags(
                        self.new_tree1.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name1 = accountsname1[accountsname1.AccountClass == grandchild_items1[o]].iloc[:, 2:4]
                    full_name1 = num_name1["GLAccountNumber"].map(str) + ' ' + num_name1["GLAccountName"]
                    full_name1.sort_values(inplace=True)

                    for z in full_name1:
                        self.new_tree1.grandgrandchild = QTreeWidgetItem(self.new_tree1.grandchild)

                        self.new_tree1.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree1.grandgrandchild.setFlags(
                            self.new_tree1.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree1.grandgrandchild.setCheckState(0, Qt.Unchecked)

        self.new_tree1.get_selected_leaves()

        cursor2 = self.cnxn.cursor()
        sql2 = '''
                                 SELECT 											
                                        *
                                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                            '''.format(field=self.selected_project_id)

        accountsname2 = pd.read_sql(sql2, self.cnxn)
        accountType2 = accountsname2.AccountType.unique()
        accountType2.sort()

        for n, i in enumerate(accountType2):
            self.new_tree2.parent = QTreeWidgetItem(self.new_tree2.tree)
            self.new_tree2.parent.setText(0, "{}".format(i))
            self.new_tree2.parent.setFlags(self.new_tree2.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)

            child_items2 = accountsname2.AccountSubType[accountsname2.AccountType == accountType2[n]].unique()
            child_items2.sort()

            for m, x in enumerate(child_items2):
                self.new_tree2.child = QTreeWidgetItem(self.new_tree2.parent)
                self.new_tree2.child.setText(0, "{}".format(x))
                self.new_tree2.child.setFlags(self.new_tree2.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)

                grandchild_items2 = accountsname2.AccountClass[accountsname2.AccountSubType == child_items2[m]].unique()
                grandchild_items2.sort()

                for o, y in enumerate(grandchild_items2):
                    self.new_tree2.grandchild = QTreeWidgetItem(self.new_tree2.child)
                    self.new_tree2.grandchild.setText(0, "{}".format(y))
                    self.new_tree2.grandchild.setFlags(
                        self.new_tree2.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name2 = accountsname2[accountsname2.AccountClass == grandchild_items2[o]].iloc[:, 2:4]
                    full_name2 = num_name2["GLAccountNumber"].map(str) + ' ' + num_name2["GLAccountName"]
                    full_name2.sort_values(inplace=True)

                    for z in full_name2:
                        self.new_tree2.grandgrandchild = QTreeWidgetItem(self.new_tree2.grandchild)
                        self.new_tree2.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree2.grandgrandchild.setFlags(
                            self.new_tree2.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree2.grandgrandchild.setCheckState(0, Qt.Unchecked)

        self.new_tree2.get_selected_leaves_1()

        ### 데이터 추출 버튼
        self.btn1 = QPushButton('   Extract Data', self.dialog12)
        self.btn1.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn1.clicked.connect(self.Thread12)
        font9 = self.btn1.font()
        font9.setBold(True)
        self.btn1.setFont(font9)

        ### 창 닫기 버튼
        self.btnDialog1 = QPushButton("   Close", self.dialog12)
        self.btnDialog1.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog1.clicked.connect(self.dialog_close12)
        font10 = self.btnDialog1.font()
        font10.setBold(True)
        self.btnDialog1.setFont(font10)
        self.btn1.resize(110, 30)
        self.btnDialog1.resize(110, 30)

        ### A 계정명 / 계정 코드 버튼
        labelAccount1 = QLabel('A 계정명/계정 코드* : ', self.dialog12)
        labelAccount1.setStyleSheet("color: yellow;")
        font3 = labelAccount1.font()
        font3.setBold(True)
        labelAccount1.setFont(font3)

        ### B 계정명 / 계정 코드 버튼
        labelAccount2 = QLabel('B 계정명/계정 코드 : ', self.dialog12)
        labelAccount2.setStyleSheet("color: white;")
        font3 = labelAccount2.font()
        font3.setBold(True)
        labelAccount2.setFont(font3)

        ### 중요성 금액
        labelD12_TE = QLabel('중요성 금액 : ', self.dialog12)
        labelD12_TE.setStyleSheet("color: white;")
        font3 = labelD12_TE.font()
        font3.setBold(True)
        labelD12_TE.setFont(font3)
        self.D12_TE = QLineEdit(self.dialog12)
        self.D12_TE.setStyleSheet("background-color: white;")
        self.D12_TE.setPlaceholderText('중요성 금액을 입력하세요')
        self.D12_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        ### A 차변 / 대변 체크 박스
        self.checkC1 = QCheckBox('Credit', self.dialog12)
        self.checkD1 = QCheckBox('Debit', self.dialog12)
        self.checkC1.setStyleSheet("color: white;")
        self.checkD1.setStyleSheet("color: white;")
        labelDC1 = QLabel('A 차변/대변 : ', self.dialog12)
        labelDC1.setStyleSheet("color: white;")
        font1 = labelDC1.font()
        font1.setBold(True)
        labelDC1.setFont(font1)

        ### B 차변 / 대변 체크 박스
        self.checkC2 = QCheckBox('Credit', self.dialog12)
        self.checkD2 = QCheckBox('Debit', self.dialog12)
        self.checkC2.setStyleSheet("color: white;")
        self.checkD2.setStyleSheet("color: white;")
        labelDC2 = QLabel('B 차변/대변 : ', self.dialog12)
        labelDC2.setStyleSheet("color: white;")
        font1 = labelDC2.font()
        font1.setBold(True)
        labelDC2.setFont(font1)

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog12)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog12)
        self.Auto = QCheckBox('자동', self.dialog12)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* : ', self.dialog12)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D12_Sheet = QLineEdit(self.dialog12)
        self.D12_Sheet.setStyleSheet("background-color: white;")
        self.D12_Sheet.setPlaceholderText('※ 입력 예시 : F01')
        self.D12_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        ### 기능영역 유/무
        self.checkF = QCheckBox('유', self.dialog12)
        labelBlank = QLabel('ㅤ', self.dialog12)
        self.checkF.setStyleSheet("color: white;")
        labelBlank.setStyleSheet("color: white;")
        labelFP = QLabel('기능영역 : ', self.dialog12)
        labelFP.setStyleSheet("color: white;")
        font11 = labelFP.font()
        font11.setBold(True)
        labelFP.setFont(font1)

        ### 기능영역 Layout
        sublayout_F = QHBoxLayout()
        sublayout_F.addWidget(labelFP)
        sublayout_F.addWidget(self.checkF)
        sublayout_F.addWidget(labelBlank)

        ### A 차대변 체크박스 Layout
        sublayout_CD1 = QHBoxLayout()
        sublayout_CD1.addWidget(labelDC1)
        sublayout_CD1.addWidget(self.checkD1)
        sublayout_CD1.addWidget(self.checkC1)

        ### B 차대변 체크박스 Layout
        sublayout_CD2 = QHBoxLayout()
        sublayout_CD2.addWidget(labelDC2)
        sublayout_CD2.addWidget(self.checkD2)
        sublayout_CD2.addWidget(self.checkC2)

        ### 수자동 체크박스 Layout
        sublayout_am = QHBoxLayout()
        sublayout_am.addWidget(labelManual)
        sublayout_am.addWidget(self.Manual)
        sublayout_am.addWidget(self.Auto)

        ### 중간 Layout
        sublayout12 = QGridLayout()
        sublayout12.addWidget(labelSheet, 0, 0)
        sublayout12.addWidget(self.D12_Sheet, 0, 1)
        sublayout12.addWidget(labelAccount1, 1, 0)
        sublayout12.addWidget(self.new_tree1, 1, 1)
        sublayout12.addWidget(self.Addnew12_A.btnMid, 2, 1)
        sublayout12.addWidget(self.Addnew12_A.Acount, 3, 1)
        sublayout12.addWidget(labelAccount2, 4, 0)
        sublayout12.addWidget(self.new_tree2, 4, 1)
        sublayout12.addWidget(self.Addnew12_B.btnMid, 5, 1)
        sublayout12.addWidget(self.Addnew12_B.Acount, 6, 1)
        sublayout12.addWidget(labelD12_TE, 7, 0)
        sublayout12.addWidget(self.D12_TE, 7, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        sublayout04 = QHBoxLayout()
        sublayout04.addStretch()
        sublayout04.addStretch()
        sublayout04.addWidget(self.btn1)
        sublayout04.addWidget(self.btnDialog1)

        ### Main Layout (8-1)
        main_layout3 = QVBoxLayout()
        main_layout3.addWidget(Titlelabel11)
        main_layout3.addLayout(sublayout12)
        main_layout3.addLayout(sublayout_CD1)
        main_layout3.addLayout(sublayout_CD2)
        main_layout3.addLayout(sublayout_am)
        main_layout3.addLayout(sublayout_F)
        main_layout3.addStretch()
        main_layout3.addLayout(sublayout04)

        ### Cursor문
        TitlelabelC = QLabel('상대계정 상세 내역 추출\n')
        TitlelabelC.setStyleSheet("color: white; font-weight : bold")

        self.AddnewC = AddForm()

        ### 중요성 금액
        label_TE = QLabel('중요성 금액 : ', self.dialog12)
        label_TE.setStyleSheet("color: white;")
        font14 = label_TE.font()
        font14.setBold(True)
        label_TE.setFont(font14)
        self.D12C_TE = QLineEdit(self.dialog12)
        self.D12C_TE.setStyleSheet("background-color: white;")
        self.D12C_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog12)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.ThreadC)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 창 닫기 버튼
        self.btnDialog2 = QPushButton("   Close", self.dialog12)
        self.btnDialog2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog2.clicked.connect(self.dialog_close12)
        font10 = self.btnDialog2.font()
        font10.setBold(True)
        self.btnDialog2.setFont(font10)
        self.btn2.resize(110, 30)
        self.btnDialog2.resize(110, 30)

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog12)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog12)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 기능영역
        self.checkF2 = QCheckBox('유', self.dialog12)
        self.checkF2.setStyleSheet("color: white;")
        labelBlank2 = QLabel('ㅤ', self.dialog12)
        labelBlank2.setStyleSheet("color: white;")
        labelFP2 = QLabel('기능영역 : ', self.dialog12)
        labelFP2.setStyleSheet("color: white;")
        font112 = labelFP2.font()
        font112.setBold(True)
        labelFP2.setFont(font12)

        ### 입력된 Cursor문
        labelCursortext = QLabel('입력된 Cursor : ', self.dialog12)
        labelCursortext.setStyleSheet("color: white;")
        font17 = labelCursortext.font()
        font17.setBold(True)
        labelCursortext.setFont(font17)
        self.Cursortext = QTextEdit(self.dialog12)
        self.Cursortext.setPlaceholderText('추출된 Cursor 조건이 표시됩니다')
        self.Cursortext.setReadOnly(True)
        self.Cursortext.setStyleSheet("background-color: white;")

        ### 커서 파일 위치
        labelCursor = QLabel('Cursor 파일 위치* : ', self.dialog12)
        labelCursor.setStyleSheet("color: yellow;")
        font3 = labelCursor.font()
        font3.setBold(True)
        labelCursor.setFont(font3)
        self.cursorCondition = QLineEdit(self.dialog12)
        self.cursorCondition.setStyleSheet("background-color: white;")
        self.cursorCondition.setPlaceholderText('Cursor 파일을 넣어주세요')
        self.cursorFile = QPushButton('File Open')
        self.cursorFile.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.cursorFile.clicked.connect(self.CursorFileOpen)
        font10 = self.cursorFile.font()
        font10.setBold(True)
        self.cursorFile.setFont(font10)

        ### 커서 시트 위치
        listCursor = QLabel('Cursor Sheet 위치* : ', self.dialog12)
        listCursor.setStyleSheet("color: yellow;")
        font13 = listCursor.font()
        font13.setBold(True)
        listCursor.setFont(font13)
        self.listCursor = QComboBox(self)
        self.listCursor.setStyleSheet("background-color: white;")

        ### 시나리오 번호
        labelSheetc = QLabel('시나리오 번호* : ', self.dialog12)
        labelSheetc.setStyleSheet("color: yellow;")
        font5 = labelSheetc.font()
        font5.setBold(True)
        labelSheetc.setFont(font5)
        self.D12_Sheetc = QLineEdit(self.dialog12)
        self.D12_Sheetc.setStyleSheet("background-color: white;")
        self.D12_Sheetc.setPlaceholderText('※ 입력 예시 : F01')

        ### 수동 / 자동 체크 박스
        labelManualC = QLabel('수동/자동 : ', self.dialog12)
        labelManualC.setStyleSheet("color: white; font-weight : bold")
        self.ManualC = QCheckBox('수동', self.dialog12)
        self.AutoC = QCheckBox('자동', self.dialog12)
        self.ManualC.setStyleSheet("color: white;")
        self.AutoC.setStyleSheet("color: white;")

        ### 수자동 체크박스 Layout
        sublayout_amC = QHBoxLayout()
        sublayout_amC.addWidget(labelManualC)
        sublayout_amC.addWidget(self.ManualC)
        sublayout_amC.addWidget(self.AutoC)

        ### 기능영역 체크박스 Layout
        sublayout000 = QHBoxLayout()
        sublayout000.addWidget(labelFP2)
        sublayout000.addWidget(self.checkF2)
        sublayout000.addWidget(labelBlank2)

        ### 최상단 Layout
        sublayout5 = QGridLayout()
        sublayout5.addWidget(self.rbtn1, 0, 0)
        sublayout5.addWidget(self.rbtn2, 0, 1)
        sublayout5.addWidget(labelSheetc, 1, 0)
        sublayout5.addWidget(self.D12_Sheetc, 1, 1)
        sublayout5.addWidget(labelCursor, 2, 0)
        sublayout5.addWidget(self.cursorCondition, 2, 1)
        sublayout5.addWidget(self.cursorFile, 2, 2)
        sublayout5.addWidget(listCursor, 3, 0)
        sublayout5.addWidget(self.listCursor, 3, 1)
        sublayout5.addWidget(label_TE, 4, 0)
        sublayout5.addWidget(self.D12C_TE, 4, 1)
        sublayout5.addWidget(labelCursortext, 5, 0)
        sublayout5.addWidget(self.Cursortext, 5, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        sublayout6 = QHBoxLayout()
        sublayout6.addStretch(2)
        sublayout6.addWidget(self.btn2)
        sublayout6.addWidget(self.btnDialog2)

        ### Main Layout
        main_layout2 = QVBoxLayout()
        main_layout2.addWidget(TitlelabelC)
        main_layout2.addLayout(sublayout5)
        main_layout2.addLayout(sublayout_amC)
        main_layout2.addLayout(sublayout000)
        main_layout2.addLayout(sublayout6)

        ### 탭 지정
        layout = QVBoxLayout()
        tabs = QTabWidget()
        tab3 = QWidget()  ### 시나리오 8-1
        tab2 = QWidget()  ### 시나리오 8-2(커서문)
        tab2.setLayout(main_layout2)
        tab3.setLayout(main_layout3)
        tabs.addTab(tab3, "Step1")
        tabs.addTab(tab2, "Step2")
        layout.addWidget(TitlelabelMain)
        layout.addWidget(tabs)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog12.setLayout(layout)
        self.dialog12.setGeometry(100, 100, 1000, 600)
        self.dialog12.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog12.setWindowTitle('Scenario8')
        self.dialog12.setWindowModality(Qt.NonModal)
        self.dialog12.show()

    ### 상대계정 Reference 파일을 읽어오는 함수
    def CursorFileOpen(self):
        self.listCursor.clear()
        fname = QFileDialog.getOpenFileName(self)
        if fname[0] == '':
            self.dialog12.activateWindow()
        else:
            self.cursorCondition.setText(fname[0])
            try:
                self.wb2 = pd.ExcelFile(fname[0])
                wbname = self.wb2.sheet_names
                for name in wbname:
                    self.listCursor.addItem(str(name))
                self.dialog12.activateWindow()
            except:
                self.MessageBox_Open("선택된 파일이 Excel 파일이 아닙니다.")

    ### 시나리오 9. 연속된 숫자로 끝나는 금액 검토
    def Dialog13(self):
        self.Addnew13 = AddForm()
        self.Addnew13.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew13.Acount))

        Titlelabel13 = QLabel('09. 연속된 숫자로 끝나는 금액 검토\n')
        Titlelabel13.setStyleSheet("color: white; font-weight : bold")

        self.dialoglist.add(13)
        self.dialog13 = QDialog()
        self.dialog13.setStyleSheet('background-color: #2E2E38')
        self.dialog13.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)
        ### 계정 트리
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog13)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog13)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog13)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread13)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 버튼 2 - Close
        self.btnDialog = QPushButton('Close', self.dialog13)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close13)

        font9 = self.btnDialog.font()
        font9.setBold(True)
        self.btnDialog.setFont(font9)

        ### 라벨 1 - 연속된 자릿수
        label_Continuous = QLabel('연속된 자릿수* : ', self.dialog13)
        label_Continuous.setStyleSheet("color: yellow;")
        font1 = label_Continuous.font()
        font1.setBold(True)
        label_Continuous.setFont(font1)

        ### Text Edit - 연속된 자릿수
        self.text_continuous = QTextEdit(self.dialog13)
        self.text_continuous.setAcceptRichText(False)
        self.text_continuous.setStyleSheet("background-color: white;")
        self.text_continuous.setPlaceholderText('연속된 자릿수를 입력하세요 (입력 예시: 3333, 666666)')

        ### 라벨 2 - 중요성 금액
        label_amount = QLabel('중요성 금액 : ', self.dialog13)
        label_amount.setStyleSheet("color: white;")

        font3 = label_amount.font()
        font3.setBold(True)
        label_amount.setFont(font3)

        ### Line Edit - 중요성 금액
        self.D13_TE = QLineEdit(self.dialog13)
        self.D13_TE.setStyleSheet("background-color: white;")
        self.D13_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 라벨 3 - 계정 트리
        label_tree = QLabel('특정 계정명 : ', self.dialog13)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelSheet = QLabel('시나리오 번호* : ', self.dialog13)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### Line Edit - 시트명
        self.D13_Sheet = QLineEdit(self.dialog13)
        self.D13_Sheet.setStyleSheet("background-color: white;")
        self.D13_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        labelDC = QLabel('차변/대변 : ', self.dialog13)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        ### 차변/대변 체크박스로 구현
        self.checkC = QCheckBox('Credit', self.dialog13)
        self.checkD = QCheckBox('Debit', self.dialog13)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        labelManual = QLabel('수동/자동 : ', self.dialog13)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog13)
        self.Auto = QCheckBox('자동', self.dialog13)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog13)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount13)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Layout - 다이얼로그 UI
        main_layout = QVBoxLayout()
        sublayout1 = QGridLayout()
        sublayout2 = QHBoxLayout()

        ### sublayout 배치 - 탭 삭제
        sublayout1.addWidget(self.rbtn1, 0, 0)
        sublayout1.addWidget(self.rbtn2, 0, 1)
        sublayout1.addWidget(labelSheet, 1, 0)
        sublayout1.addWidget(self.D13_Sheet, 1, 1)
        sublayout1.addWidget(label_Continuous, 2, 0)
        sublayout1.addWidget(self.text_continuous, 2, 1)
        sublayout1.addWidget(label_amount, 3, 0)
        sublayout1.addWidget(self.D13_TE, 3, 1)
        sublayout1.addWidget(label_tree, 4, 0)
        sublayout1.addWidget(self.new_tree, 4, 1)
        sublayout1.addWidget(self.Addnew13.btnMid, 5, 1)
        sublayout1.addWidget(self.Addnew13.Acount, 6, 1)
        sublayout1.addWidget(self.Addnew13.sourceLabel, 7, 0)
        sublayout1.addWidget(self.Addnew13.source, 7, 1)
        sublayout1.addWidget(self.Addnew13.UserLabel, 8, 0)
        sublayout1.addWidget(self.Addnew13.User, 8, 1)

        sublayout2.addStretch(2)
        sublayout2.addWidget(self.btnCount)
        sublayout2.addWidget(self.btn2)
        sublayout2.addWidget(self.btnDialog)

        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel13)
        main_layout.addLayout(sublayout1, stretch=4)
        main_layout.addLayout(self.Addnew13.sublayout1, stretch=4)
        main_layout.addLayout(layout_dc, stretch=4)
        main_layout.addLayout(layout_am, stretch=4)
        main_layout.addLayout(sublayout2, stretch=1)

        ### 공통 지정
        self.dialog13.setLayout(main_layout)
        self.dialog13.setGeometry(100, 100, 1000, 600)

        # ? 제거
        self.dialog13.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog13.setWindowTitle('Scenario9')
        self.dialog13.setWindowModality(Qt.NonModal)
        self.dialog13.show()

    ### 시나리오 10. 전표 description에 공란 또는 특정단어(key word)가 입력되어 있는 전표 리스트
    def Dialog14(self):
        self.Addnew14 = AddForm()
        self.Addnew14.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew14.Acount))

        ### 상단 라벨
        Titlelabel14 = QLabel('10. 전표 description에 공란 또는 특정단어(key word)가 입력되어 있는 전표 리스트\n')
        Titlelabel14.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(14)
        self.dialog14 = QDialog()
        self.dialog14.setStyleSheet('background-color: #2E2E38')
        self.dialog14.setWindowIcon(QIcon(self.resource_path("./EY_logo.png")))

        ### 계정 트리
        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog14)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread14)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton("   Close", self.dialog14)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close14)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        ### 데이터 추출 버튼 & 창 닫기 버튼 사이즈 조절
        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        ### JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog14)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog14)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 특정단어 (포함)
        labelKeyword = QLabel('Key Words* : ', self.dialog14)
        labelKeyword.setStyleSheet("color: yellow;")
        font1 = labelKeyword.font()
        font1.setBold(True)
        labelKeyword.setFont(font1)

        ### 특정단어 (제외)
        labelKeyword2 = QLabel('Except Key Words : ', self.dialog14)
        labelKeyword2.setStyleSheet("color: white;")
        font3 = labelKeyword2.font()
        font3.setBold(True)
        labelKeyword2.setFont(font3)

        ### 특정단어 (포함) QLineEdit
        self.D14_Key = QLineEdit(self.dialog14)
        self.D14_Key.setStyleSheet("background-color: white;")
        self.D14_Key.setPlaceholderText('검색할 단어를 입력하세요(구분자:",")')

        ### 특정단어 (제외) QLineEdit & Activate 체크박스
        self.D14_Key2 = QLineEdit(self.dialog14)
        self.D14_Key2.setStyleSheet("background-color: white;")
        self.D14_Key2.setPlaceholderText('제외할 단어를 입력하세요(구분자:",")')
        self.D14_Key2C = QCheckBox('Activate')
        self.D14_Key2C.setStyleSheet("color: white; font-weight: bold")

        ### 중요성 금액
        labelD14_TE = QLabel('중요성 금액 : ', self.dialog14)
        labelD14_TE.setStyleSheet("color: white;")

        font2 = labelD14_TE.font()
        font2.setBold(True)
        labelD14_TE.setFont(font2)

        self.D14_TE = QLineEdit(self.dialog14)
        self.D14_TE.setStyleSheet("background-color: white;")
        self.D14_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 특정 계정명 라벨
        label_tree = QLabel('특정 계정 : ', self.dialog14)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* : ', self.dialog14)
        labelSheet.setStyleSheet("color: yellow;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D14_Sheet = QLineEdit(self.dialog14)
        self.D14_Sheet.setStyleSheet("background-color: white;")
        self.D14_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### 차변 / 대변 라벨
        labelDC = QLabel('차변/대변 : ', self.dialog14)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        ### 수동 / 자동 체크 박스
        self.checkC = QCheckBox('Credit', self.dialog14)
        self.checkD = QCheckBox('Debit', self.dialog14)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog14)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog14)
        self.Auto = QCheckBox('자동', self.dialog14)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        self.D14_Key.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D14_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D14_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        self.D14_Key2C.stateChanged.connect(self.D14_LabelC)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog14)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount14)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 중간 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D14_Sheet, 1, 1)
        layout1.addWidget(labelKeyword, 2, 0)
        layout1.addWidget(self.D14_Key, 2, 1)
        layout1.addWidget(labelKeyword2, 3, 0)
        layout1.addWidget(self.D14_Key2, 3, 1)
        layout1.addWidget(self.D14_Key2C, 4, 0)
        layout1.addWidget(labelD14_TE, 5, 0)
        layout1.addWidget(self.D14_TE, 5, 1)
        layout1.addWidget(label_tree, 6, 0)
        layout1.addWidget(self.new_tree, 6, 1)
        layout1.addWidget(self.Addnew14.btnMid, 7, 1)
        layout1.addWidget(self.Addnew14.Acount, 8, 1)
        layout1.addWidget(self.Addnew14.sourceLabel, 9, 0)
        layout1.addWidget(self.Addnew14.source, 9, 1)
        layout1.addWidget(self.Addnew14.UserLabel, 10, 0)
        layout1.addWidget(self.Addnew14.User, 10, 1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel14)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew14.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog14.setLayout(main_layout)
        self.dialog14.setGeometry(100, 100, 1000, 600)

        ### 상단 ? 삭제, 창 라벨 설정
        self.dialog14.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog14.setWindowTitle("Scenario10")
        self.dialog14.setWindowModality(Qt.NonModal)
        self.dialog14.show()

    ### 시나리오 11. 증빙일과 전기일의 회계기간이 다른 전표 추출 및 검토
    def Dialog15(self):
        self.Addnew15 = AddForm()
        self.Addnew15.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew15.Acount))

        ### 상단 라벨
        Titlelabel15 = QLabel('11. 증빙일과 전기일의 회계기간이 다른 전표 추출 및 검토\n')
        Titlelabel15.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(15)
        self.dialog15 = QDialog()
        self.dialog15.setStyleSheet('background-color: #2E2E38')
        self.dialog15.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                         SELECT 											
                                *
                         FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                    '''.format(field=self.selected_project_id)
        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)
            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog15)
        self.btn2.setStyleSheet('color:white; background-image : url(./bar.png)')
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.clicked.connect(self.Thread15)

        ### 버튼 2 - Close
        self.btnDialog = QPushButton('   Close', self.dialog15)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close15)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog15)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog15)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### LineEdit 2 - 중요성 금액
        label_TE = QLabel('중요성 금액 : ', self.dialog15)
        label_TE.setStyleSheet("color: white;")
        font1 = label_TE.font()
        font1.setBold(True)
        label_TE.setFont(font1)
        self.D15_TE = QLineEdit(self.dialog15)
        self.D15_TE.setStyleSheet('background-color: white;')
        self.D15_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 라벨 3 - 시트명
        labelSheet = QLabel('시나리오 번호* : ', self.dialog15)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### LineEdit 3 - 시트명
        self.D15_Sheet = QLineEdit(self.dialog15)
        self.D15_Sheet.setStyleSheet("background-color: white;")
        self.D15_Sheet.setPlaceholderText('※ 입력 예시 : F01')
        label_tree = QLabel('특정 계정명 :           ', self.dialog15)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog15)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog15)
        self.checkD = QCheckBox('Debit', self.dialog15)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog15)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog15)
        self.Auto = QCheckBox('자동', self.dialog15)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### LineEdit만 창 크기에 따라 확대/축소
        self.D15_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D15_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog15)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount15)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D15_Sheet, 1, 1)
        layout1.addWidget(label_TE, 2, 0)
        layout1.addWidget(self.D15_TE, 2, 1)
        layout1.addWidget(label_tree, 3, 0)
        layout1.addWidget(self.new_tree, 3, 1)
        layout1.addWidget(self.Addnew15.btnMid, 4, 1)
        layout1.addWidget(self.Addnew15.Acount, 5, 1)
        layout1.addWidget(self.Addnew15.sourceLabel, 6, 0)
        layout1.addWidget(self.Addnew15.source, 6, 1)
        layout1.addWidget(self.Addnew15.UserLabel, 7, 0)
        layout1.addWidget(self.Addnew15.User, 7, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch(2)
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel15)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew15.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog15.setLayout(main_layout)
        self.dialog15.setGeometry(100, 100, 1000, 600)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog15.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog15.setWindowTitle('Scenario11')
        self.dialog15.setWindowModality(Qt.NonModal)
        self.dialog15.show()

    ### 시나리오 12. 차/대변 합계가 중요성금액 이상인 전표
    def Dialog16(self):
        self.Addnew16 = AddForm()
        self.Addnew16.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew16.Acount))

        ### 상단 라벨
        Titlelabel16 = QLabel('12. 차/대변 합계가 중요성금액 이상인 전표\n')
        Titlelabel16.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(16)
        self.dialog16 = QDialog()
        self.dialog16.setStyleSheet('background-color: #2E2E38')
        self.dialog16.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                                 SELECT 											
                                        *
                                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                            '''.format(field=self.selected_project_id)
        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)
            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()

        ### 데이터 추출 버튼
        self.btn2 = QPushButton('   Extract Data', self.dialog16)
        self.btn2.setStyleSheet('color:white; background-image : url(./bar.png)')
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.clicked.connect(self.Thread16)

        ### 창 닫기 버튼
        self.btnDialog = QPushButton('   Close', self.dialog16)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close16)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        ### JE Line / JE 선택 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog16)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog16)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 중요성 금액
        label_TE = QLabel('중요성 금액* : ', self.dialog16)
        label_TE.setStyleSheet("color: yellow;")
        font1 = label_TE.font()
        font1.setBold(True)
        label_TE.setFont(font1)
        self.D16_TE = QLineEdit(self.dialog16)
        self.D16_TE.setStyleSheet('background-color: white;')
        self.D16_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 전표입력일
        labelDate = QLabel('전표입력일 :        ', self.dialog16)
        labelDate.setStyleSheet("color: white;")
        font3 = labelDate.font()
        font3.setBold(True)
        labelDate.setFont(font3)
        self.period1 = QLineEdit(self.dialog16)
        self.period1.setStyleSheet("background-color: white;")
        self.period1.setPlaceholderText('시작 시점을 입력하세요 yyyyMMdd')
        self.period2 = QLineEdit(self.dialog16)
        self.period2.setStyleSheet("background-color: white;")
        self.period2.setPlaceholderText('종료 시점을 입력하세요 yyyyMMdd')

        ### 시나리오 번호
        labelSheet = QLabel('시나리오 번호* : ', self.dialog16)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)
        self.D16_Sheet = QLineEdit(self.dialog16)
        self.D16_Sheet.setStyleSheet("background-color: white;")
        self.D16_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### 특정 계정명
        label_tree = QLabel('특정 계정명 : ', self.dialog16)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog16)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog16)
        self.checkD = QCheckBox('Debit', self.dialog16)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog16)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog16)
        self.Auto = QCheckBox('자동', self.dialog16)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### LineEdit만 창 크기에 따라 확대/축소
        self.D16_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D16_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog16)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount16)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 입력일 Layout
        layout_date = QHBoxLayout()
        layout_date.addWidget(self.period1)
        layout_date.addWidget(self.period2)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D16_Sheet, 1, 1)
        layout1.addWidget(label_TE, 2, 0)
        layout1.addWidget(self.D16_TE, 2, 1)
        layout1.addWidget(labelDate, 3, 0)
        layout1.addLayout(layout_date, 3, 1)
        layout1.addWidget(label_tree, 4, 0)
        layout1.addWidget(self.new_tree, 4, 1)
        layout1.addWidget(self.Addnew16.btnMid, 6, 1)
        layout1.addWidget(self.Addnew16.Acount, 7, 1)
        layout1.addWidget(self.Addnew16.sourceLabel, 8, 0)
        layout1.addWidget(self.Addnew16.source, 8, 1)
        layout1.addWidget(self.Addnew16.UserLabel, 9, 0)
        layout1.addWidget(self.Addnew16.User, 9, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch(2)
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel16)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew16.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog16.setLayout(main_layout)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog16.setGeometry(100, 100, 1000, 600)
        self.dialog16.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog16.setWindowTitle('Scenario12')
        self.dialog16.setWindowModality(Qt.NonModal)
        self.dialog16.show()

    ### 시나리오 13. 전표 입력자와 승인자가 동일한 전표
    def Dialog17(self):
        self.Addnew17 = AddForm()
        self.Addnew17.btnMid.clicked.connect(lambda: self.AccountUpdate(self.Addnew17.Acount))

        ### 상단 라벨
        Titlelabel17 = QLabel('13. 전표 입력자와 승인자가 동일한 전표\n')
        Titlelabel17.setStyleSheet("color: white; font-weight : bold")

        ### 다이얼로그 추가 및 디자인 설정
        self.dialoglist.add(17)
        self.dialog17 = QDialog()
        self.dialog17.setStyleSheet('background-color: #2E2E38')
        self.dialog17.setWindowIcon(QIcon(self.resource_path('./EY_logo.png')))

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                                 SELECT 											
                                        *
                                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                            '''.format(field=self.selected_project_id)
        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)
            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Unchecked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog17)
        self.btn2.setStyleSheet('color:white; background-image : url(./bar.png)')
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.clicked.connect(self.Thread17)

        ### 버튼 2 - Close
        self.btnDialog = QPushButton('   Close', self.dialog17)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close17)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line (Result)', self.dialog17)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE (Journals)', self.dialog17)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### LineEdit 2 - 중요성 금액
        label_TE = QLabel('중요성 금액 : ', self.dialog17)
        label_TE.setStyleSheet("color: white;")
        font1 = label_TE.font()
        font1.setBold(True)
        label_TE.setFont(font1)
        self.D17_TE = QLineEdit(self.dialog17)
        self.D17_TE.setStyleSheet('background-color: white;')
        self.D17_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 라벨 3 - 시트명
        labelSheet = QLabel('시나리오 번호* : ', self.dialog17)
        labelSheet.setStyleSheet("color: yellow;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### LineEdit 3 - 시트명
        self.D17_Sheet = QLineEdit(self.dialog17)
        self.D17_Sheet.setStyleSheet("background-color: white;")
        self.D17_Sheet.setPlaceholderText('※ 입력 예시 : F01')
        label_tree = QLabel('특정 계정명 : ', self.dialog17)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        ### 차변 / 대변 체크 박스
        labelDC = QLabel('차변/대변 : ', self.dialog17)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog17)
        self.checkD = QCheckBox('Debit', self.dialog17)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")

        ### 수동 / 자동 체크 박스
        labelManual = QLabel('수동/자동 : ', self.dialog17)
        labelManual.setStyleSheet("color: white; font-weight : bold")
        self.Manual = QCheckBox('수동', self.dialog17)
        self.Auto = QCheckBox('자동', self.dialog17)
        self.Manual.setStyleSheet("color: white;")
        self.Auto.setStyleSheet("color: white;")

        ### LineEdit만 창 크기에 따라 확대/축소
        self.D17_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.D17_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        ################################# Line Count 버튼 ##################################
        self.btnCount = QPushButton("  Line Count", self.dialog17)
        self.btnCount.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnCount.clicked.connect(self.lineCount17)
        fontlc = self.btnCount.font()
        fontlc.setBold(True)
        self.btnCount.setFont(fontlc)
        self.btnCount.resize(110, 30)

        ### 최상단 Layout
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelSheet, 1, 0)
        layout1.addWidget(self.D17_Sheet, 1, 1)
        layout1.addWidget(label_TE, 2, 0)
        layout1.addWidget(self.D17_TE, 2, 1)
        layout1.addWidget(label_tree, 3, 0)
        layout1.addWidget(self.new_tree, 3, 1)
        layout1.addWidget(self.Addnew17.btnMid, 4, 1)
        layout1.addWidget(self.Addnew17.Acount, 5, 1)
        layout1.addWidget(self.Addnew17.sourceLabel, 6, 0)
        layout1.addWidget(self.Addnew17.source, 6, 1)
        layout1.addWidget(self.Addnew17.UserLabel, 7, 0)
        layout1.addWidget(self.Addnew17.User, 7, 1)

        ### 데이터 추출 / 창 닫기 버튼 Layout
        layout2 = QHBoxLayout()
        layout2.addStretch(2)
        layout2.addWidget(self.btnCount)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)

        ### 차대변 체크박스 Layout
        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkD)
        layout_dc.addWidget(self.checkC)

        ### 수자동 체크박스 Layout
        layout_am = QHBoxLayout()
        layout_am.addWidget(labelManual)
        layout_am.addWidget(self.Manual)
        layout_am.addWidget(self.Auto)

        ### Main Layout
        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(Titlelabel17)
        main_layout.addLayout(layout1)
        main_layout.addLayout(self.Addnew17.sublayout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout_am)
        main_layout.addLayout(layout2)
        self.dialog17.setLayout(main_layout)
        self.dialog17.setGeometry(100, 100, 1000, 600)

        ### Main Layout 창 크기 조절, 상단 ? 삭제, 창 라벨 설정
        self.dialog17.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog17.setWindowTitle('Scenario13')
        self.dialog17.setWindowModality(Qt.NonModal)
        self.dialog17.show()

    ############################################## Line Count 함수 ###############################################
    def lineCount4(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew4.SegmentBox1,
                                                                           self.Addnew4.SegmentBox2,
                                                                           self.Addnew4.SegmentBox3,
                                                                           self.Addnew4.SegmentBox4,
                                                                           self.Addnew4.SegmentBox5,
                                                                           self.Addnew4.UserDefine1,
                                                                           self.Addnew4.UserDefine2,
                                                                           self.Addnew4.UserDefine3,
                                                                           self.Addnew4.User, self.Addnew4.source,
                                                                           self.Manual, self.Auto)
        self.temp_N = self.D4_N.text()
        self.temp_TE = self.D4_TE.text()

        if self.Addnew4.Acount.toPlainText() == '':
            self.checked_account4 = ''

        else:
            self.checked_account4 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew4.Acount.toPlainText() + ')'

        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
            self.debitcredit = ''
        elif self.checkD.isChecked():  # Credit 이 0
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        ### 예외처리 1 - 필수값 입력 누락
        if self.temp_N == '':
            self.alertbox_open()

        ### 쿼리 연동
        else:
            if self.temp_TE == '': self.temp_TE = 0
            if self.check_account(self.checked_account4) != False:
                try:
                    int(self.temp_N)
                    float(self.temp_TE)

                    cursor = self.cnxn.cursor()
                    ### JE Line - Result
                    if self.rbtn1.isChecked():
                        sql  = '''
                                        SET NOCOUNT ON
                                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                                    GROUP BY CoA.GLAccountNumber
                                                        SELECT	COUNT(*) as cnt	
                                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                                        AND JournalEntries.JELINEID = Details.JENumberID 
                                                        AND JournalEntries.GLAccountNumber IN 				
                                                            (			
                                                                SELECT DISTINCT JournalEntries.GLAccountNumber			
                                                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                                                [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                                WHERE JournalEntries.JELINEID = Details.JENumberID 
                                                                GROUP BY JournalEntries.GLAccountNumber
                                                                HAVING COUNT(JournalEntries.GLAccountNumber) <= {N}		
                                                            ) 
                                                        AND ABS(JournalEntries.Amount) >= {TE}
                                                        {Account}
                                                        {NewSQL}
                                                        {AutoManual}
                                                        {DebitCredit}				
                                                        DROP TABLE #TMPCOA
                                                    '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                               N=self.temp_N,
                                                               Account=self.checked_account4, NewSQL=self.NewSQL,
                                                               AutoManual=self.ManualAuto,
                                                               DebitCredit=self.debitcredit)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    ### JE - Journals
                    elif self.rbtn2.isChecked():
                        sql = '''
                                        SET NOCOUNT ON
                                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                                    GROUP BY CoA.GLAccountNumber
                                                    SELECT	COUNT(*) as cnt	
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA
                                                    , [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                                    AND JournalEntries.JELINEID = Details.JENumberID 
                                                    AND Details.JEIdentifierID IN 
                                                        (				
                                                        SELECT DISTINCT Details.JEIdentifierID			
                                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,			
                    	                                    [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details				
                                                        WHERE JournalEntries.JELINEID = Details.JENumberID
                                                        AND JournalEntries.GLAccountNumber IN 			
                                                                (	
                                                                SELECT DISTINCT JournalEntries.GLAccountNumber			
                                                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                                                [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                                WHERE JournalEntries.JELINEID = Details.JENumberID 
                                                                GROUP BY JournalEntries.GLAccountNumber
                                                                HAVING COUNT(JournalEntries.GLAccountNumber) <= {N}
                                                                ) 
                                                        AND ABS(JournalEntries.Amount) >= {TE}
                                                        {Account}
                                                        {NewSQL}
                                                        {AutoManual}
                                                        {DebitCredit}
                                                        ) 	
                                                    DROP TABLE #TMPCOA
                                            '''.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                                                       Account=self.checked_account4, NewSQL=self.NewSQL,
                                                       AutoManual=self.ManualAuto,
                                                       DebitCredit=self.debitcredit)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    buttonReply = QMessageBox.information(self, '라인 수 확인','라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>', QMessageBox.Ok)
                    if buttonReply == QMessageBox.Ok: self.dialog4.activateWindow()

                ### 예외처리 5 - 필수 입력값 타입 오류
                except ValueError:
                    try:
                        int(self.temp_N)
                        try:
                            float(self.temp_TE)
                        except:
                            self.alertbox_open2('중요성금액')
                    except:
                        try:
                            float(self.temp_TE)
                            self.alertbox_open2('계정사용 빈도수')
                        except:
                            self.alertbox_open2('계정사용 빈도수와 중요성금액')

    def lineCount5(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew5.SegmentBox1,
                                                                           self.Addnew5.SegmentBox2,
                                                                           self.Addnew5.SegmentBox3,
                                                                           self.Addnew5.SegmentBox4,
                                                                           self.Addnew5.SegmentBox5,
                                                                           self.Addnew5.UserDefine1,
                                                                           self.Addnew5.UserDefine2,
                                                                           self.Addnew5.UserDefine3,
                                                                           self.Addnew5.User, self.Addnew5.source,
                                                                           self.Manual, self.Auto)

        ### 인풋 값 변수로 받아오기
        self.temp_TE = self.D5_TE.text()  ### 중요성금액

        ##Unselect all의 경우
        if self.Addnew5.Acount.toPlainText() == '':
            self.checked_account5 = "AND JournalEntries.GLAccountNumber IN ('')"  ###당기 생성 계정이 없는 경우 고려

        ##Select all이나 일부 체크박스가 선택된 경우
        else:
            self.checked_account5 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew5.Acount.toPlainText() + ')'

        ### 예외처리 1 - 필수값 입력 누락
        if self.checked_account5 == '':
            self.alertbox_open()

        else:
            if self.temp_TE == '':
                self.temp_TE = 0

            ##Checked_account의 유효성 체크
            if self.check_account(self.checked_account5) == False:
                return

            try:
                float(self.temp_TE)

                if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                        not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                    self.debitcredit = ''
                elif self.checkD.isChecked():  # Credit 이 0
                    self.debitcredit = 'AND JournalEntries.Credit = 0'
                elif self.checkC.isChecked():  # Debit 이 0
                    self.debitcredit = 'AND JournalEntries.Debit = 0'

                ### 쿼리 연동
                cursor = self.cnxn.cursor()
                ### JE Line
                if self.rbtn1.isChecked():
                    sql_query = """
                                    SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) as cnt            
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 											
                                    AND ABS(JournalEntries.Amount) >= {TE} 				
                                    {Account}
                                    {DebitCredit}
                                    {NewSQL}	
                                    {AutoManual}					
                                    DROP TABLE #TMPCOA				
                                            """.format(field=self.selected_project_id, TE=self.temp_TE,
                                                       Account=self.checked_account5, DebitCredit=self.debitcredit,
                                                       NewSQL=self.NewSQL,
                                                       AutoManual=self.ManualAuto)

                    self.dataframe = pd.read_sql(sql_query, self.cnxn)

                ### JE
                elif self.rbtn2.isChecked():
                    sql_query = """
                                    SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) as cnt		
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 							
                                    AND Details.JEIdentifierID IN				
                                            (		
                                             SELECT DISTINCT Details.JEIdentifierID		
                                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                             AND ABS(JournalEntries.Amount) >= {TE}	
                                             {Account}
                                             {DebitCredit}
                                             {NewSQL}
                                             {AutoManual}	
                                            )					
                                    DROP TABLE #TMPCOA				                                                                       
                                            """.format(field=self.selected_project_id, TE=self.temp_TE,
                                                       Account=self.checked_account5, DebitCredit=self.debitcredit,
                                                       NewSQL=self.NewSQL,
                                                       AutoManual=self.ManualAuto)
                    self.dataframe = pd.read_sql(sql_query, self.cnxn)

                buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                      '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                      QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

            ### 예외처리 5 - 필수 입력값 타입 오류
            except ValueError:
                self.alertbox_open2('중요성금액')

    def lineCount6(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew6.SegmentBox1,
                                                                           self.Addnew6.SegmentBox2,
                                                                           self.Addnew6.SegmentBox3,
                                                                           self.Addnew6.SegmentBox4,
                                                                           self.Addnew6.SegmentBox5,
                                                                           self.Addnew6.UserDefine1,
                                                                           self.Addnew6.UserDefine2,
                                                                           self.Addnew6.UserDefine3,
                                                                           self.Addnew6.User, self.Addnew6.source,
                                                                           self.Manual, self.Auto)

        ### 중요성 금액
        self.temp_TE = self.D6_TE.text()

        ### 필수 입력값 누락 검토
        if self.period1.text() == ''  or self.period2.text() == '':
            self.alertbox_open()

        else:
            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '':
                self.temp_TE = 0

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew6.Acount.toPlainText() == '':
                self.checked_account6 = ''
            else:
                self.checked_account6 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew6.Acount.toPlainText() + ')'

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account6) != False:
                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.temp_TE)

                    ### 시작/종료 날짜 정수로 입력했는지 확인
                    int(self.period1.text())
                    int(self.period2.text())

                    ### 시작/종료 시점 쿼리문에 적용할 수 있도록 변환
                    self.tempDate1 = "'" + self.period1.text() + "'"
                    self.tempDate2 = "'" + self.period2.text() + "'"

                    ### 시점 자릿수 확인(' 포함 10자리 여부 확인)
                    if len(str(self.tempDate1)) != 10:
                        self.alertbox_open19()
                    elif len(str(self.tempDate2)) != 10:
                        self.alertbox_open19()
                    else:
                        cursor = self.cnxn.cursor()

                        ### JE Line 추출
                        if self.rbtn1.isChecked():
                            sql = '''
                                                SET NOCOUNT ON				
                                                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                                GROUP BY CoA.GLAccountNumber				
                                                SELECT	COUNT(*) as cnt
                                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                    #TMPCOA,			
                                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                                AND JournalEntries.JELINEID = Details.JENumberID			
                                                AND JournalEntries.EntryDate >= {period1}				
                                                AND JournalEntries.EntryDate <= {period2}				
                                                AND ABS(JournalEntries.Amount) >= {TE}			
                                                {Account}			
                                                {NewSQL}				
                                                {DebitCredit}				
                                                {AutoManual}							
                                                DROP TABLE #TMPCOA				
                                            '''.format(field=self.selected_project_id, Account=self.checked_account6,
                                                       TE=self.temp_TE,
                                                       period1=str(self.tempDate1), period2=str(self.tempDate2),
                                                       NewSQL=self.NewSQL, DebitCredit=self.debitcredit,
                                                       AutoManual=self.ManualAuto)

                            self.dataframe = pd.read_sql(sql, self.cnxn)

                        ### JE 추출
                        elif self.rbtn2.isChecked():
                            sql = '''
                                               SET NOCOUNT ON				
                                                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                                GROUP BY CoA.GLAccountNumber				
                                                SELECT	COUNT(*) as cnt			
                                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                    #TMPCOA,			
                                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                                AND JournalEntries.JELINEID = Details.JENumberID 							
                                                AND Details.JEIdentifierID IN				
                                                        (		
                                                         SELECT DISTINCT Details.JEIdentifierID		
                                                         FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                         WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                                         AND JournalEntries.EntryDate >= {period1}	
                                                         AND JournalEntries.EntryDate <= {period2}	
                                                         AND ABS(JournalEntries.Amount) >= {TE}	
                                                         {Account}	
                                                         {NewSQL}		
                                                         {DebitCredit}		
                                                         {AutoManual}		
                                                        )						
                                                DROP TABLE #TMPCOA						
                                            '''.format(field=self.selected_project_id, Account=self.checked_account6,
                                                       TE=self.temp_TE,
                                                       period1=str(self.tempDate1), period2=str(self.tempDate2),
                                                       NewSQL=self.NewSQL, DebitCredit=self.debitcredit,
                                                       AutoManual=self.ManualAuto)

                            self.dataframe = pd.read_sql(sql, self.cnxn)

                        buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                              '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                              QMessageBox.Ok)

                        if buttonReply == QMessageBox.Ok: self.dialog6.activateWindow()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        float(self.temp_TE)  ### 중요성 금액이 실수가 아닌 경우
                    except:
                        try:
                            int(self.period1.text())
                            int(self.period2.text())
                            self.alertbox_open2('중요성 금액')
                        except:
                            self.alertbox_open2('입력일과 중요성 금액')  ### 중요성 금액과 입력일의 형식이 잘못되었을 경우
                    try:
                        int(self.period1.text())
                        int(self.period2.text())
                    except:
                        try:
                            float(self.temp_TE)
                            self.alertbox_open2('입력일')  ### 입력일의 형식이 잘못되었을 경우
                        except:
                            self.alertbox_open2('입력일과 중요성 금액')  ### 중요성 금액과 입력일의 형식이 잘못되었을 경우

    def lineCount7(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew7.SegmentBox1,
                                                                           self.Addnew7.SegmentBox2,
                                                                           self.Addnew7.SegmentBox3,
                                                                           self.Addnew7.SegmentBox4,
                                                                           self.Addnew7.SegmentBox5,
                                                                           self.Addnew7.UserDefine1,
                                                                           self.Addnew7.UserDefine2,
                                                                           self.Addnew7.UserDefine3,
                                                                           self.Addnew7.User, self.Addnew7.source,
                                                                           self.Manual, self.Auto)

        ### 중요성 금액
        self.temp_TE = self.D7_TE.text()

        if (self.Entry.isChecked() and self.Effective.isChecked()) or (
                not (self.Entry.isChecked()) and not (self.Effective.isChecked())):
            self.alertbox_open21()

        else:
            self.holiday = []  # 공휴일 리스트
            self.holiday_str = []  # 공휴일, 주말
            self.realDate_List = []  # 중복 제거 전 SQL 쿼리에 들어갈 리스트
            self.realDate_List_final = []  # SQL 쿼리에 들어갈 리스트

            ### 공휴일 추가
            self.holiday = [pytimekr.holidays(i) for i in range(2021, 2023)]
            for i in range(len(self.holiday)):
                for d in range(0, len(self.holiday[i])):
                    self.date_str = self.holiday[i][d].strftime('%Y-%m-%d')
                    self.holiday_str.append(self.date_str)

            ### 주말 추가
            self.start_date = date(2020, 1, 1)
            self.end_date = date(2022, 12, 31)
            self.delta = timedelta(days=1)
            while self.start_date <= self.end_date:
                if self.start_date.weekday() == 5 or self.start_date.weekday() == 6:
                    self.a = self.start_date.strftime('%Y-%m-%d')
                    self.holiday_str.append(self.a)
                self.start_date += self.delta

            ### 공휴일, 주말 yyyyMMdd 형식에 맞게 변환
            for i in range(0, len(self.holiday_str)):
                self.tempDate = []
                self.tempDate = str(self.holiday_str[i]).split('-')
                self.realDate = self.tempDate[0] + self.tempDate[1] + self.tempDate[2]
                self.realDate_List.append(self.realDate)

            ### 사용자 입력 일자 추가
            if self.D7_Date.toPlainText() != '':
                self.user_date = self.D7_Date.toPlainText().split(',')

            else:
                self.user_date = ''

            for a in self.user_date:
                a = a.strip()
                try:
                    int(a)
                    if len(a) == 8:  ### 날짜 형식이 yyyyMMdd일 경우만 추가
                        b = a
                        self.realDate_List.append(b)
                    else:
                        self.alertbox_open19()  ### 형식이 올바르지 않은 경우, 팝업
                        return
                except:
                    self.alertbox_open19()
                    return

            ### 날짜 중복 제거 완료 (self.realDate_List_final)
            self.realDate_List_final = set(self.realDate_List)

            ### 쿼리문에 적용할 수 있게끔 변환
            self.checked_date = ''
            for i in self.realDate_List_final:
                self.checked_date = self.checked_date + ',' + '\'' + i + '\''

            self.checked_date = self.checked_date[1:]

            self.checked_effective = 'AND JournalEntries.EffectiveDate IN (' + self.checked_date + ')'
            self.checked_entry = 'AND JournalEntries.EntryDate IN (' + self.checked_date + ')'

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():  # Debit 이 0
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 입력일을 선택했을 시, EntryDate 조건문
            if self.Entry.isChecked():
                self.tempState = self.checked_entry

            ### 전기일을 선택했을 시, EffectiveDate 조건문
            elif self.Effective.isChecked():
                self.tempState = self.checked_effective

            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '': self.temp_TE = 0

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew7.Acount.toPlainText() == '':
                self.checked_account7 = ''
            else:
                self.checked_account7 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew7.Acount.toPlainText() + ')'

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account7) != False:
                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.temp_TE)

                    ### Count 쿼리문
                    cursor = self.cnxn.cursor()

                    ### JE Line 추출
                    if self.rbtn1.isChecked():
                        sql = '''
                                           SET NOCOUNT ON				
                                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                            GROUP BY CoA.GLAccountNumber
                                            SELECT COUNT(*) as cnt	
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                #TMPCOA,			
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                            AND JournalEntries.JELINEID = Details.JENumberID 							
                                            {Date}				
                                            AND ABS(JournalEntries.Amount) >= {TE}		
                                            {Account}			
                                            {NewSQL}				
                                            {DebitCredit}				
                                            {AutoManual}											
                                            DROP TABLE #TMPCOA				
                                       '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                  Date=self.tempState,
                                                  Account=self.checked_account7, NewSQL=self.NewSQL,
                                                  AutoManual=self.ManualAuto,
                                                  DebitCredit=self.debitcredit)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    ### JE 추출
                    elif self.rbtn2.isChecked():
                        sql = '''
                                        SET NOCOUNT ON				
                                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                        GROUP BY CoA.GLAccountNumber				
                                        SELECT COUNT(*) as cnt	
                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                            #TMPCOA,			
                                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                        AND JournalEntries.JELINEID = Details.JENumberID 								
                                        AND Details.JEIdentifierID IN				
                                                (		
                                                 SELECT DISTINCT Details.JEIdentifierID		
                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                 WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                                 {Date}	
                                                 AND ABS(JournalEntries.Amount) >= {TE}	
                                                 {Account}		
                                                 {NewSQL}		
                                                 {DebitCredit}		
                                                 {AutoManual}		
                                                )						
                                        DROP TABLE #TMPCOA				
                                       '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                  Date=self.tempState,
                                                  Account=self.checked_account7, NewSQL=self.NewSQL,
                                                  AutoManual=self.ManualAuto,
                                                  DebitCredit=self.debitcredit)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                          '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                          QMessageBox.Ok)

                    if buttonReply == QMessageBox.Ok: self.dialog7.activateWindow()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    ### 중요성 금액이 실수가 아닌 경우
                    self.alertbox_open2('중요성 금액')

    def lineCount8(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew8.SegmentBox1,
                                                                           self.Addnew8.SegmentBox2,
                                                                           self.Addnew8.SegmentBox3,
                                                                           self.Addnew8.SegmentBox4,
                                                                           self.Addnew8.SegmentBox5,
                                                                           self.Addnew8.UserDefine1,
                                                                           self.Addnew8.UserDefine2,
                                                                           self.Addnew8.UserDefine3,
                                                                           self.Addnew8.User, self.Addnew8.source,
                                                                           self.Manual, self.Auto)

        ### N일
        self.tempN = self.D8_N.text()
        ### 중요성 금액
        self.temp_TE = self.D8_TE.text()

        ### 필수 입력값 누락 검토
        if self.tempN == '' :
            self.alertbox_open()

        else:
            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '':
                self.temp_TE = 0

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew8.Acount.toPlainText() == '':
                self.checked_account8 = ''
            else:
                self.checked_account8 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew8.Acount.toPlainText() + ')'

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account8) != False:
                try:
                    ### N 정수값인지 확인
                    int(self.tempN)

                    ### 중요성 금액 실수값인지 확인
                    float(self.temp_TE)

                    ### N값이 0이상 70만 이하인지 확인(비정상적인 값 예외처리)
                    if int(self.tempN) < 0 or int(self.tempN) > 700000:
                        self.alertbox_open13()
                        int('False')

                    else:
                        self.realNDate = int(self.tempN)
                        cursor = self.cnxn.cursor()

                        ### JE Line 추출
                        if self.rbtn1.isChecked():
                            sql = '''
                                                    SET NOCOUNT ON				
                                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                                    GROUP BY CoA.GLAccountNumber
                                                    SELECT COUNT(*) as cnt		
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                        #TMPCOA,			
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                                    AND JournalEntries.JELINEID = Details.JENumberID 						
                                                    AND ABS(DATEDIFF(dd, JournalEntries.EntryDate ,JournalEntries.EffectiveDate)) >= {N}			
                                                    AND ABS(JournalEntries.Amount) >= {TE}			
                                                    {Account}				
                                                    {NewSQL}				
                                                    {DebitCredit}
                                                    {AutoManual}								
                                                    DROP TABLE #TMPCOA				
                                                '''.format(field=self.selected_project_id, N=self.realNDate,
                                                           TE=self.temp_TE,
                                                           Account=self.checked_account8, AutoManual=self.ManualAuto,
                                                           NewSQL=self.NewSQL,
                                                           DebitCredit=self.debitcredit)

                            self.dataframe = pd.read_sql(sql, self.cnxn)

                        elif self.rbtn2.isChecked():

                            sql = '''
                                                    SET NOCOUNT ON				
                                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                                    GROUP BY CoA.GLAccountNumber				
                                                    SELECT	COUNT(*) as cnt	
                                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                        #TMPCOA,			
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                                    AND JournalEntries.JELINEID = Details.JENumberID 						
                                                    AND Details.JEIdentifierID IN				
                                                            (		
                                                             SELECT DISTINCT Details.JEIdentifierID		
                                                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                                             AND ABS(DATEDIFF(dd, JournalEntries.EntryDate ,JournalEntries.EffectiveDate)) >= {N}
                                                             AND ABS(JournalEntries.Amount) >= {TE}
                                                             {Account}	
                                                             {NewSQL}	
                                                             {DebitCredit}
                                                             {AutoManual}
                                                            )				
                                                    DROP TABLE #TMPCOA				
                                                    '''.format(field=self.selected_project_id, N=self.realNDate,
                                                               TE=self.temp_TE,
                                                               Account=self.checked_account8,
                                                               AutoManual=self.ManualAuto,
                                                               NewSQL=self.NewSQL,
                                                               DebitCredit=self.debitcredit)

                            self.dataframe = pd.read_sql(sql, self.cnxn)

                        buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                              '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                              QMessageBox.Ok)

                        if buttonReply == QMessageBox.Ok: self.dialog8.activateWindow()


                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        int(self.tempN)
                        try:
                            float(self.temp_TE)  ### 중요성 금액이 실수가 아닌 경우
                        except:
                            self.alertbox_open2('중요성금액')
                    except:
                        try:
                            float(self.temp_TE)
                            self.alertbox_open2('N')  ### N이 정수가 아닌 경우
                        except:
                            self.alertbox_open2('중요성금액과 N')  ### 중요성 금액과 N의 형식이 잘못된 경우

    def lineCount9(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew9.SegmentBox1,
                                                                           self.Addnew9.SegmentBox2,
                                                                           self.Addnew9.SegmentBox3,
                                                                           self.Addnew9.SegmentBox4,
                                                                           self.Addnew9.SegmentBox5,
                                                                           self.Addnew9.UserDefine1,
                                                                           self.Addnew9.UserDefine2,
                                                                           self.Addnew9.UserDefine3,
                                                                           self.Addnew9.User, self.Addnew9.source,
                                                                           self.Manual, self.Auto)
        self.tempN = self.D9_N.text()  # 전표 작성 빈도수 N회
        self.tempTE = self.D9_TE.text()  # 중요성 금액

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew9.Acount.toPlainText() == '':
            self.checked_account9 = ''

        else:
            self.checked_account9 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew9.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.tempN == '':
            self.alertbox_open()

        ### 중요성 금액 미입력시 0원
        else:
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account9) != False:
                try:
                    ### N 정수값인지 확인
                    int(self.tempN)
                    ### 중요성 금액 실수값인지 확인
                    float(self.tempTE)

                    ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
                    if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                            not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                        self.debitcredit = ''

                    ### Debit을 선택했을 시, Credit이 0원
                    elif self.checkD.isChecked():
                        self.debitcredit = 'AND JournalEntries.Credit = 0'
                    ### Credit을 선택했을 시, Debit이 0원
                    elif self.checkC.isChecked():
                        self.debitcredit = 'AND JournalEntries.Debit = 0'

                    ### 쿼리 연동
                    cursor = self.cnxn.cursor()
                    if self.rbtn1.isChecked():  # JE Line- Result
                        sql = '''
                                        SET NOCOUNT ON				
                                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                        GROUP BY CoA.GLAccountNumber				
                                        SELECT COUNT(*) AS cnt	
                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                            #TMPCOA,			
                                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                        AND JournalEntries.JELINEID = Details.JENumberID 							
                                        AND JournalEntries.PreparerID IN				
                                                (		
                                                 SELECT DISTINCT JournalEntries.PreparerID		
                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details		
                                                 WHERE JournalEntries.JELINEID = Details.JENumberID 
                                                 GROUP BY JournalEntries.PreparerID		
                                                 HAVING COUNT(JournalEntries.PreparerID) <= {N}
                                                )		
                                        AND ABS(JournalEntries.Amount) >= {TE}
                                        {Account}
                                        {DebitCredit}
                                        {NewSQL}
                                        {AutoManual}				
                                        DROP TABLE #TMPCOA				
                                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                                   DebitCredit=self.debitcredit,
                                                   Account=self.checked_account9, NewSQL=self.NewSQL,
                                                   AutoManual=self.ManualAuto)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    elif self.rbtn2.isChecked():  # JE- Journals
                        sql = '''
                                        SET NOCOUNT ON				
                                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                        GROUP BY CoA.GLAccountNumber				
                                        SELECT COUNT(*) AS cnt
                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                            #TMPCOA,			
                                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                        AND JournalEntries.JELINEID = Details.JENumberID 				
                                        AND Details.JEIdentifierID IN				
                                                (		
                                                 SELECT DISTINCT Details.JEIdentifierID		
                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                 WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                                 AND JournalEntries.PreparerID IN		
                                                        (
                                                         SELECT DISTINCT JournalEntries.PreparerID		
                                                         FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details		
                                                         WHERE JournalEntries.JELINEID = Details.JENumberID 
                                                         GROUP BY JournalEntries.PreparerID		
                                                         HAVING COUNT(JournalEntries.PreparerID) <= {N}
                                                        )
                                                AND ABS(JournalEntries.Amount) >= {TE} 
                                                {Account}
                                                {DebitCredit}
                                                {NewSQL}
                                                {AutoManual}		
                                                )						
                                        DROP TABLE #TMPCOA				
                                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                                   DebitCredit=self.debitcredit,
                                                   Account=self.checked_account9, NewSQL=self.NewSQL,
                                                   AutoManual=self.ManualAuto)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                          '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                          QMessageBox.Ok)
                    if buttonReply == QMessageBox.Ok: self.dialog9.activateWindow()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        int(self.tempN)
                        try:
                            float(self.tempTE)
                        except:
                            self.alertbox_open4('중요성금액을 숫자로만 입력해주시기 바랍니다.')  # 중요성금액이 실수가 아닌 경우
                    except:
                        try:
                            float(self.tempTE)
                            self.alertbox_open4('작성빈도수를 숫자로만 입력해주시기 바랍니다.')  # 작성빈도수가 정수가 아닌 경우
                        except:
                            self.alertbox_open4('작성빈도수와 중요성금액을 숫자로만 입력해주시기 바랍니다.')  # 중요성금액이 실수가 아니고 작성빈도수가 정수가 아닌 경우

    def lineCount10(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew10.SegmentBox1,
                                                                           self.Addnew10.SegmentBox2,
                                                                           self.Addnew10.SegmentBox3,
                                                                           self.Addnew10.SegmentBox4,
                                                                           self.Addnew10.SegmentBox5,
                                                                           self.Addnew10.UserDefine1,
                                                                           self.Addnew10.UserDefine2,
                                                                           self.Addnew10.UserDefine3,
                                                                           self.Addnew10.User, self.Addnew10.source,
                                                                           self.Manual, self.Auto)
        self.tempTE = self.D10_TE.text()

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew10.Acount.toPlainText() == '':
            self.checked_account10 = ''

        else:
            self.checked_account10 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew10.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.Addnew10.User.text() == '':
            self.alertbox_open()

        else:
            ### 중요성 금액 미입력시 0원
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력값 검토
            if self.check_account(self.checked_account10) != False:
                try:
                    float(self.tempTE)
                    if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                            not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                        self.debitcredit = ''
                    elif self.checkD.isChecked():
                        self.debitcredit = 'AND JournalEntries.Credit = 0'
                    elif self.checkC.isChecked():  # Debit 이 0
                        self.debitcredit = 'AND JournalEntries.Debit = 0'

                    cursor = self.cnxn.cursor()

                    ### JE Line 추출
                    if self.rbtn1.isChecked():

                        sql = '''
                                                     SET NOCOUNT ON
                                                     SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                                     FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                                     GROUP BY CoA.GLAccountNumber
                                                     SELECT	COUNT(*) as cnt		
                                                   FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                                                   [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                                   WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                                   AND JournalEntries.JELINEID = Details.JENumberID 				        	
                                                   AND ABS(JournalEntries.Amount) >= {TE} 
                                                   {Account}
                                                   {NewSQL}
                                                   {AutoManual}
                                                   {DebitCredit}
                                                   DROP TABLE #TMPCOA			
                                                '''.format(field=self.selected_project_id, TE=self.tempTE,
                                                           Account=self.checked_account10,
                                                           NewSQL=self.NewSQL, AutoManual=self.ManualAuto,
                                                           DebitCredit=self.debitcredit)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    ### JE 추출
                    elif self.rbtn2.isChecked():

                        sql = '''
                                                       SET NOCOUNT ON
                                                       SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                                       FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                                       GROUP BY CoA.GLAccountNumber
                                                       SELECT COUNT(*) as cnt
                                                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                                                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                                       WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                                       AND JournalEntries.JELINEID = Details.JENumberID 
                                                       AND Details.JEIdentifierID IN 		
                                                                        (	
                                                                        SELECT DISTINCT Details.JEIdentifierID	
                                                                        FROM  [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                                                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                                        WHERE JournalEntries.JELINEID = Details.JENumberID 
                                                                        AND ABS(JournalEntries.Amount) >= {TE}	
                                                                        {Account}
                                                                        {NewSQL}
                                                                        {AutoManual}
                                                                        {DebitCredit}
                                                                        ) 
                                                       DROP TABLE #TMPCOA			
                                                '''.format(field=self.selected_project_id, TE=self.tempTE,
                                                           Account=self.checked_account10,
                                                           NewSQL=self.NewSQL, AutoManual=self.ManualAuto,
                                                           DebitCredit=self.debitcredit)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                          '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                          QMessageBox.Ok)
                    if buttonReply == QMessageBox.Ok:self.dialog10.activateWindow()

                except ValueError:
                    self.alertbox_open4("중요성금액 값을 숫자로만 입력해주시기 바랍니다.")

    def lineCount13(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew13.SegmentBox1,
                                                                           self.Addnew13.SegmentBox2,
                                                                           self.Addnew13.SegmentBox3,
                                                                           self.Addnew13.SegmentBox4,
                                                                           self.Addnew13.SegmentBox5,
                                                                           self.Addnew13.UserDefine1,
                                                                           self.Addnew13.UserDefine2,
                                                                           self.Addnew13.UserDefine3,
                                                                           self.Addnew13.User, self.Addnew13.source,
                                                                           self.Manual, self.Auto)

        self.temp_Continuous = self.text_continuous.toPlainText()  # 필수
        self.temp_Continuous = str(self.temp_Continuous).strip()
        self.temp_TE = self.D13_TE.text()

        ##Unselect all의 경우
        if self.Addnew13.Acount.toPlainText() == '':
            self.checked_account13 = ''

        ##Select all이나 일부 체크박스가 선택된 경우
        else:
            self.checked_account13 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew13.Acount.toPlainText() + ')'

        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
            self.debitcredit = ''
        elif self.checkD.isChecked():
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        ### 예외처리 1 - 필수값 누락
        if self.temp_Continuous == '':
            self.alertbox_open()

        else:
            if self.temp_TE == '':
                self.temp_TE = 0

            ##Checked_account의 유효성 체크
            if self.check_account(self.checked_account13) == False:
                return

            try:
                float(self.temp_TE)
            except ValueError:
                self.alertbox_open2('중요성금액')
                return

            try:
                self.temp_Continuous = re.sub(r"[\s]+", '', self.temp_Continuous)
            except:
                self.MessageBox_Open("연속된 자릿수 입력이 잘못되었습니다.")
                return

            self.temp_Continuous = self.temp_Continuous.split(',')

            for i in range(len(self.temp_Continuous)):
                ### 예외처리 3 - 숫자가 아닌 값 입력한 경우
                try:
                    int(self.temp_Continuous[i])
                except ValueError:
                    self.alertbox_open2("연속된 자릿수")
                    return

            self.filter_Continuous = ""
            for x in self.temp_Continuous:
                self.filter_Continuous += f"OR RIGHT(FLOOR(JournalEntries.Amount), {len(x)}) IN (\'{x}\')"
            self.filter_Continuous = 'AND (' + self.filter_Continuous[3:] + ')'

            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                self.debitcredit = ''
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'
            elif self.checkC.isChecked():  # Debit 이 0
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 쿼리 연동
            cursor = self.cnxn.cursor()

            ### JE Line
            if self.rbtn1.isChecked():
                sql_query = '''
                                        SET NOCOUNT ON				
                                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                        GROUP BY CoA.GLAccountNumber				
                                        SELECT COUNT(*) as cnt
                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                            #TMPCOA,			
                                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                        AND JournalEntries.JELINEID = Details.JENumberID 							
                                        {Continuous} 		
                                        AND ABS(JournalEntries.Amount) >= {TE}			
                                        {Account}			
                                        {NewSQL}
                                        {DebitCredit}
                                        {AutoManual}								
                                        DROP TABLE #TMPCOA				
                                                '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                           Account=self.checked_account13,
                                                           DebitCredit=self.debitcredit,
                                                           NewSQL=self.NewSQL,
                                                           AutoManual=self.ManualAuto,
                                                           Continuous=self.filter_Continuous)

                self.dataframe = pd.read_sql(sql_query, self.cnxn)

                ### JE - Journals
            elif self.rbtn2.isChecked():
                sql_query = '''
                                            SET NOCOUNT ON				
                                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                            GROUP BY CoA.GLAccountNumber				
                                            SELECT COUNT(*) AS cnt
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                #TMPCOA,			
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                            AND JournalEntries.JELINEID = Details.JENumberID 							
                                            AND Details.JEIdentifierID IN				
                                                    (		
                                                     SELECT DISTINCT Details.JEIdentifierID		
                                                     FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                     WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                                     {Continuous}		
                                                     AND ABS(JournalEntries.Amount) >= {TE}		
                                                     {Account}	
                                                     {NewSQL}		
                                                     {DebitCredit}
                                                     {AutoManual}	
                                                    )				
                                            DROP TABLE #TMPCOA				
                                                    '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                               Account=self.checked_account13,
                                                               DebitCredit=self.debitcredit,
                                                               NewSQL=self.NewSQL,
                                                               AutoManual=self.ManualAuto,
                                                               Continuous=self.filter_Continuous)

                self.dataframe = pd.read_sql(sql_query, self.cnxn)

            buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                  '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                  QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog13.activateWindow()

    def lineCount14(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew14.SegmentBox1,
                                                                           self.Addnew14.SegmentBox2,
                                                                           self.Addnew14.SegmentBox3,
                                                                           self.Addnew14.SegmentBox4,
                                                                           self.Addnew14.SegmentBox5,
                                                                           self.Addnew14.UserDefine1,
                                                                           self.Addnew14.UserDefine2,
                                                                           self.Addnew14.UserDefine3,
                                                                           self.Addnew14.User, self.Addnew14.source,
                                                                           self.Manual, self.Auto)

        self.tempTE = self.D14_TE.text()  # 중요성 금액

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew14.Acount.toPlainText() == '':
            self.checked_account14 = ''

        else:
            self.checked_account14 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew14.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.D14_Key.text().strip() == '':
            self.alertbox_open()

        ### 제외 키워드를 activate한 상태에서 제외 키워드를 입력하지 않을 경우, 경고창 생성
        elif self.D14_Key2C.isChecked() and self.D14_Key2.text().strip() == '':
            self.alertbox_open6()

        ### 콤마(,) 구분자를 이용하여 전표 적요 특정단어 입력 (포함 단어)
        else:
            self.baseKey = self.D14_Key.text().split(',')
            self.baseKey_clean = []
            for a in self.baseKey:
                a = a.strip()
                if a.upper() == '[NULL]':
                    b = "((JournalEntries.JEDescription LIKE '' OR JournalEntries.JEDescription LIKE ' ' OR JournalEntries.JEDescription IS NULL)" \
                        "AND (JournalEntries.JELineDescription LIKE '' OR JournalEntries.JELineDescription LIKE ' ' OR JournalEntries.JELineDescription IS NULL))"
                elif a == '':
                    continue
                else:
                    b = "(JournalEntries.JEDescription LIKE N'%" + a + "%' OR JournalEntries.JELineDescription LIKE N'%" + a + "%')"
                self.baseKey_clean.append(b)

            ### 콤마(,) 구분자를 이용하여 전표 적요 특정단어 입력 (제외 단어)
            self.baseKey2 = self.D14_Key2.text().split(',')
            self.baseKey2_clean = []
            if self.D14_Key2C.isChecked():
                for a in self.baseKey2:
                    a = a.strip()
                    if a.upper() == '[NULL]':
                        b = "(NOT (JournalEntries.JEDescription LIKE '' OR JournalEntries.JEDescription LIKE ' ' OR JournalEntries.JEDescription IS NULL)" \
                            "OR NOT (JournalEntries.JELineDescription LIKE '' OR JournalEntries.JELineDescription LIKE ' ' OR JournalEntries.JELineDescription IS NULL))"
                    elif a == '':
                        continue
                    else:
                        b = "(NOT(JournalEntries.JEDescription LIKE N'%" + a + "%' OR JournalEntries.JELineDescription LIKE N'%" + a + "%'))"
                    self.baseKey2_clean.append(b)
                self.tempKey = 'AND (' + str('OR '.join(self.baseKey_clean)) + ') AND (' + str(
                    ' AND '.join(self.baseKey2_clean)) + ')'

            else:
                self.tempKey = 'AND (' + str(' OR '.join(self.baseKey_clean)) + ')'

            ### 중요성 금액 미입력시 0원
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account14) != False:

                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.tempTE)

                    ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
                    if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                            not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                        self.debitcredit = ''

                    ### Debit을 선택했을 시, Credit이 0원
                    elif self.checkD.isChecked():
                        self.debitcredit = 'AND JournalEntries.Credit = 0'
                    ### Credit을 선택했을 시, Debit이 0원
                    elif self.checkC.isChecked():
                        self.debitcredit = 'AND JournalEntries.Debit = 0'

                    if self.D14_Key2C.isChecked():
                        tempword = ", " + str(self.baseKey2) + "이/가 제외"
                    else:
                        tempword = ''

                    cursor = self.cnxn.cursor()

                    if self.rbtn1.isChecked():  # JE Line- Result

                        sql = '''
                                    SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) as cnt
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 
                                    AND ABS(JournalEntries.Amount) >= {TE}
                                    {KEY}
                                    {Account} 		
                                    {NewSQL}				
                                    {DebitCredit}			
                                    {AutoManual}								
                                    DROP TABLE #TMPCOA				
                                    '''.format(field=self.selected_project_id, KEY=self.tempKey, TE=self.tempTE,
                                               DebitCredit=self.debitcredit,
                                               Account=self.checked_account14, NewSQL=self.NewSQL,
                                               AutoManual=self.ManualAuto)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    elif self.rbtn2.isChecked():  # JE- Journals

                        sql = '''
                                    SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) AS cnt
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 				
                                    AND Details.JEIdentifierID IN				
                                            (		
                                             SELECT DISTINCT Details.JEIdentifierID		
                                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                             AND ABS(JournalEntries.Amount) >= {TE} 
                                             {KEY}
                                             {Account} 		
                                             {NewSQL}				
                                             {DebitCredit}			
                                             {AutoManual}		
                                             )					
                                    DROP TABLE #TMPCOA				
                                    '''.format(field=self.selected_project_id, KEY=self.tempKey, TE=self.tempTE,
                                               DebitCredit=self.debitcredit,
                                               Account=self.checked_account14, NewSQL=self.NewSQL,
                                               AutoManual=self.ManualAuto)

                        self.dataframe = pd.read_sql(sql, self.cnxn)

                    buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                          '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                          QMessageBox.Ok)
                    if buttonReply == QMessageBox.Ok:self.dialog14.activateWindow()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        float(self.tempTE)
                    except:
                        self.alertbox_open4('중요성금액 값을 숫자로만 입력해주시기 바랍니다.')  # 중요성금액이 실수가 아닌 경우

    def lineCount15(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew15.SegmentBox1,
                                                                           self.Addnew15.SegmentBox2,
                                                                           self.Addnew15.SegmentBox3,
                                                                           self.Addnew15.SegmentBox4,
                                                                           self.Addnew15.SegmentBox5,
                                                                           self.Addnew15.UserDefine1,
                                                                           self.Addnew15.UserDefine2,
                                                                           self.Addnew15.UserDefine3,
                                                                           self.Addnew15.User, self.Addnew15.source,
                                                                           self.Manual, self.Auto)
        self.tempTE = self.D15_TE.text()  # 중요성 금액

        ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
            self.debitcredit = ''
        elif self.checkD.isChecked():
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew15.Acount.toPlainText() == '':
            self.checked_account15 = ''

        else:
            self.checked_account15 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew15.Acount.toPlainText() + ')'

        if self.tempTE == '': self.tempTE = 0

        ### 계정 입력 값 검토
        if self.check_account(self.checked_account15) != False:

            try:
                ### 중요성 금액 실수값인지 확인
                float(self.tempTE)
                cursor = self.cnxn.cursor()
                ### JE Line
                if self.rbtn1.isChecked():
                    sql = '''
                                SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) as cnt	       
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 						
                                    AND Month(JournalEntries.UserDefined1) <> Month(JournalEntries.EffectiveDate) 				
                                    AND ABS(JournalEntries.Amount) >= {TE} 				
                                    {Account}					
                                    {NewSQL} 			
                                    {AutoManual}
                                    {DebitCredit}	  									
                                    DROP TABLE #TMPCOA						
                                '''.format(field=self.selected_project_id, TE=self.tempTE,
                                           Account=self.checked_account15, NewSQL=self.NewSQL,
                                           AutoManual=self.ManualAuto,
                                           DebitCredit=self.debitcredit)

                    self.dataframe = pd.read_sql(sql, self.cnxn)

                ### JE
                elif self.rbtn2.isChecked():

                    sql = '''
                                SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) as cnt		 
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 					
                                    AND Details.JEIdentifierID IN				
                                            (		
                                             SELECT DISTINCT Details.JEIdentifierID		
                                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                             AND Month(JournalEntries.UserDefined1) <> Month(JournalEntries.EffectiveDate) 		
                                             AND ABS(JournalEntries.Amount) >= {TE} 	
                                             {Account} 	
                                             {NewSQL}
                                             {AutoManual}
                                             {DebitCredit}	
                                            )		
                                    DROP TABLE #TMPCOA						
                                '''.format(field=self.selected_project_id, TE=self.tempTE,
                                           Account=self.checked_account15, NewSQL=self.NewSQL,
                                           AutoManual=self.ManualAuto,
                                           DebitCredit=self.debitcredit)
                    self.dataframe = pd.read_sql(sql, self.cnxn)

                buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                      '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                      QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:self.dialog15.activateWindow()

            except ValueError:
                self.alertbox_open4("중요성금액 값을 숫자로만 입력해주시기 바랍니다.")


    def lineCount16(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew16.SegmentBox1,
                                                                           self.Addnew16.SegmentBox2,
                                                                           self.Addnew16.SegmentBox3,
                                                                           self.Addnew16.SegmentBox4,
                                                                           self.Addnew16.SegmentBox5,
                                                                           self.Addnew16.UserDefine1,
                                                                           self.Addnew16.UserDefine2,
                                                                           self.Addnew16.UserDefine3,
                                                                           self.Addnew16.User, self.Addnew16.source,
                                                                           self.Manual, self.Auto)

        ### 중요성 금액
        self.temp_TE = self.D16_TE.text()

        ### 필수 입력값 누락 검토
        if self.temp_TE.strip() == '' :
            self.alertbox_open()

        else:

            ### 시작일이 yyyyMMdd 형식이 아닌 경우 - 숫자가 아닌 경우
            if not self.period1.text().strip().isdigit() and self.period1.text().strip() != '':
                self.alertbox_open19();
                return

            ### 종료일이 yyyyMMdd 형식이 아닌 경우 - 숫자가 아닌 경우
            if not self.period2.text().strip().isdigit() and self.period2.text().strip() != '':
                self.alertbox_open19();
                return

            ### 시작일이 yyyyMMdd 형식이 아닌 경우 - 8자리가 아닌 경우
            if len(self.period1.text().strip()) != 8 and len(self.period1.text().strip()) != 0:
                self.alertbox_open19();
                return

            ### 종료일이 yyyyMMdd 형식이 아닌 경우 - 8자리가 아닌 경우
            if len(self.period2.text().strip()) != 8 and len(self.period2.text().strip()) != 0:
                self.alertbox_open19();
                return

            self.EntryDate = ''
            self.subEntryDate = ''
            ### 시작일이 입력된 경우
            if self.period1.text().strip() != '':
                self.EntryDate += 'AND JournalEntries.EntryDate >= ' + "'" + self.period1.text().strip() + "'"
                self.subEntryDate += 'AND JournalEntries1.EntryDate >= ' + "'" + self.period1.text().strip() + "'"
            ### 종료일이 입력된 경우
            if self.period2.text().strip() != '':
                self.EntryDate += 'AND JournalEntries.EntryDate <= ' + "'" + self.period2.text().strip() + "'"
                self.subEntryDate += 'AND JournalEntries1.EntryDate <= ' + "'" + self.period2.text().strip() + "'"

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew16.Acount.toPlainText() == '':
                self.checked_account16 = ''

            else:
                self.checked_account16 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew16.Acount.toPlainText() + ')'

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            self.sub_checked_account16 = re.sub('JournalEntries.', 'JournalEntries1.', self.checked_account16)
            self.sub_NewSQL = re.sub('JournalEntries.', 'JournalEntries1.', self.NewSQL)
            self.sub_debitcredit = re.sub('JournalEntries.', 'JournalEntries1.', self.debitcredit)
            self.sub_ManualAuto = re.sub('Details.', 'Details1.', self.ManualAuto)

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account16) != False:
                try:
                    ### 중요성 금액 실수값인지 확인
                    if float(self.temp_TE) == 0.:
                        self.alertbox_open7();
                        return
                    cursor = self.cnxn.cursor()

                    ### JE Line 추출
                    if self.rbtn1.isChecked():
                        if self.debitcredit != '':

                            sql = '''
                                            SET NOCOUNT ON				
                                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                            GROUP BY CoA.GLAccountNumber;			
                                            SELECT COUNT(*) AS cnt
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                #TMPCOA,			
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			

                                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                            AND JournalEntries.JELINEID = Details.JENumberID 				

                                            {Account}			
                                            {Date}
                                            {NewSQL}					
                                            {DebitCredit}								
                                            {AutoManual}	
                                            AND (				
                                                 SELECT SUM(ABS(JournalEntries1.Amount))			
                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                                      [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                                 WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                                 AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                                 {SubAccount}
                                                 {SubDate}
                                                 {SubNewSQL}
                                                 {SubDebitCredit}
                                                 {SubAutoManual}			 
                                                 GROUP BY Details1.JEIdentifierID			
                                                ) >= {TE}	-- 중요성 금액(이상으로)							
                                            DROP TABLE #TMPCOA										
                                            '''.format(field=self.selected_project_id, Account=self.checked_account16,
                                                       TE=self.temp_TE, Date=self.EntryDate,
                                                       NewSQL=self.NewSQL, DebitCredit=self.debitcredit,
                                                       AutoManual=self.ManualAuto,
                                                       SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate,
                                                       SubNewSQL=self.sub_NewSQL,
                                                       SubDebitCredit=self.sub_debitcredit,
                                                       SubAutoManual=self.sub_ManualAuto)
                            self.dataframe = pd.read_sql(sql, self.cnxn)
                            buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                                  '라인 수 : ' + str(
                                                                      self.dataframe['cnt'].loc[0]) + '<br>',
                                                                  QMessageBox.Ok)

                            if buttonReply == QMessageBox.Ok: self.dialog16.activateWindow()

                        else:

                            sql = """
                                    		SET NOCOUNT ON				
                                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                            GROUP BY CoA.GLAccountNumber;
                                            
                                            SELECT COUNT(*) AS cnt
                                            FROM ( 
                                                (
                                                            SELECT				
                                                                JournalEntries.BusinessUnit AS 회사코드			
                                                                , JournalEntries.JENumber AS 전표번호			
                                                                , JournalEntries.JELineNumber AS 전표라인번호			
                                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,						
                                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    
                                                            WHERE JournalEntries.JELINEID = Details.JENumberID 				
                                                            {Account}			
                                                            {Date}
                                                            {NewSQL}												
                                                            {AutoManual}
                                                            AND JournalEntries.Credit = 0
                                                            AND (				
                                                                 SELECT SUM(ABS(JournalEntries1.Amount))			
                                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                                                      [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                                                 WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                                                 AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                                                 {SubAccount}
                                                                 {SubDate}
                                                                 {SubNewSQL}
                                                                 {SubAutoManual}
                                                                 AND JournalEntries1.Credit = 0					 
                                                                 GROUP BY Details1.JEIdentifierID			
                                                                ) >= {TE}	-- 중요성 금액(이상으로)			
                                                            )			
                                                            Union
                                                            (
                                                            SELECT				
                                                                JournalEntries.BusinessUnit AS 회사코드			
                                                                , JournalEntries.JENumber AS 전표번호			
                                                                , JournalEntries.JELineNumber AS 전표라인번호			
                                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,						
                                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    
                                                            WHERE JournalEntries.JELINEID = Details.JENumberID 				
                                    
                                                            {Account}			
                                                            {Date}
                                                            {NewSQL}												
                                                            {AutoManual}
                                                            AND JournalEntries.Debit = 0				
                                                            AND (				
                                                                 SELECT SUM(ABS(JournalEntries1.Amount))			
                                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                                                      [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                                                 WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                                                 AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                                                 {SubAccount}
                                                                 {SubDate}
                                                                 {SubNewSQL}
                                                                 {SubAutoManual}
                                                                 AND JournalEntries1.Debit = 0						 
                                                                 GROUP BY Details1.JEIdentifierID			
                                                                ) >= {TE}	-- 중요성 금액(이상으로)			
                                    
                                                            )
                                            ) AS A
                                    
                                                    DROP TABLE #TMPCOA	
                                            """.format(field=self.selected_project_id, Account=self.checked_account16,
                                                       TE=self.temp_TE, Date=self.EntryDate,
                                                       NewSQL=self.NewSQL,
                                                       AutoManual=self.ManualAuto,
                                                       SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate,
                                                       SubNewSQL=self.sub_NewSQL,
                                                       SubAutoManual=self.sub_ManualAuto)

                            self.dataframe = pd.read_sql(sql, self.cnxn)
                            buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                                  '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                                  QMessageBox.Ok)

                            if buttonReply == QMessageBox.Ok: self.dialog16.activateWindow()

                    ### JE 추출
                    elif self.rbtn2.isChecked():

                        if self.debitcredit != '':
                            sql = '''
                                            SET NOCOUNT ON				
                                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                            GROUP BY CoA.GLAccountNumber;			
                                            SELECT	COUNT(*) as cnt		
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                #TMPCOA,			
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			

                                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                            AND JournalEntries.JELINEID = Details.JENumberID 	
                                            AND Details.JEIdentifierID IN
                                                (			
                                                 SELECT DISTINCT Details.JEIdentifierID		
                                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                 WHERE JournalEntries.JELINEID = Details.JENumberID 	
                                                 {Account}			
                                                 {Date}
                                                 {NewSQL}					
                                                 {DebitCredit}								
                                                 {AutoManual}	
                                                 AND (				
                                                     SELECT SUM(ABS(JournalEntries1.Amount))			
                                                     FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                                          [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                                     WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                                     AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                                     {SubAccount}
                                                     {SubDate}
                                                     {SubNewSQL}
                                                     {SubDebitCredit}
                                                     {SubAutoManual}			 
                                                     GROUP BY Details1.JEIdentifierID			
                                                     ) >= {TE}	-- 중요성 금액(이상으로)			
                                                )			
                                            DROP TABLE #TMPCOA										
                                            '''.format(field=self.selected_project_id, Account=self.checked_account16,
                                                       TE=self.temp_TE, Date=self.EntryDate,
                                                       NewSQL=self.NewSQL, DebitCredit=self.debitcredit,
                                                       AutoManual=self.ManualAuto,
                                                       SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate,
                                                       SubNewSQL=self.sub_NewSQL,
                                                       SubDebitCredit=self.sub_debitcredit,
                                                       SubAutoManual=self.sub_ManualAuto)
                        else:
                            sql = """
                                    		SET NOCOUNT ON				
                                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                            GROUP BY CoA.GLAccountNumber;			

                                            SELECT COUNT(*) AS cnt
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                #TMPCOA,			
                                                [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			

                                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                            AND JournalEntries.JELINEID = Details.JENumberID 				

                                            AND Details.JEIdentifierID IN
                                                (
                                                    (
                                                    SELECT DISTINCT Details.JEIdentifierID				

                                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			

                                                    WHERE JournalEntries.JELINEID = Details.JENumberID 				

                                                    {Account}			
                                                    {Date}
                                                    {NewSQL}												
                                                    {AutoManual}
                                                    AND JournalEntries.Credit = 0
                                                    AND (				
                                                         SELECT SUM(ABS(JournalEntries1.Amount))			
                                                         FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                                              [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                                         WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                                         AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                                         {SubAccount}
                                                         {SubDate}
                                                         {SubNewSQL}
                                                         {SubAutoManual}
                                                         AND JournalEntries1.Credit = 0					 
                                                         GROUP BY Details1.JEIdentifierID			
                                                        ) >= {TE}	-- 중요성 금액(이상으로)			 
                                                    )			
                                                Union
                                                    (
                                                    SELECT DISTINCT Details.JEIdentifierID		

                                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,			
                                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			

                                                    WHERE JournalEntries.JELINEID = Details.JENumberID 				

                                                    {Account}			
                                                    {Date}
                                                    {NewSQL}												
                                                    {AutoManual}
                                                    AND JournalEntries.Debit = 0				

                                                    AND (				
                                                         SELECT SUM(ABS(JournalEntries1.Amount))			
                                                         FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                                              [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                                         WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                                         AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                                         {SubAccount}
                                                         {SubDate}
                                                         {SubNewSQL}
                                                         {SubAutoManual}
                                                         AND JournalEntries1.Debit = 0						 
                                                         GROUP BY Details1.JEIdentifierID			
                                                        ) >= {TE}	-- 중요성 금액(이상으로)			        
                                                    )
                                                )				
                    		                DROP TABLE #TMPCOA			
                                            """.format(field=self.selected_project_id, Account=self.checked_account16,
                                                       TE=self.temp_TE, Date=self.EntryDate,
                                                       NewSQL=self.NewSQL,
                                                       AutoManual=self.ManualAuto,
                                                       SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate,
                                                       SubNewSQL=self.sub_NewSQL,
                                                       SubAutoManual=self.sub_ManualAuto)

                        self.dataframe = pd.read_sql(sql, self.cnxn)
                        buttonReply = QMessageBox.information(self, '라인 수 확인',
                                                              '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                                              QMessageBox.Ok)

                        if buttonReply == QMessageBox.Ok: self.dialog16.activateWindow()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    self.alertbox_open2('중요성 금액')  ### 중요성 금액이 실수가 아닌 경우


    def lineCount17(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew17.SegmentBox1,
                                                                           self.Addnew17.SegmentBox2,
                                                                           self.Addnew17.SegmentBox3,
                                                                           self.Addnew17.SegmentBox4,
                                                                           self.Addnew17.SegmentBox5,
                                                                           self.Addnew17.UserDefine1,
                                                                           self.Addnew17.UserDefine2,
                                                                           self.Addnew17.UserDefine3,
                                                                           self.Addnew17.User, self.Addnew17.source,
                                                                           self.Manual, self.Auto)
        self.temp_TE = self.D17_TE.text()

        ##Unselect all의 경우
        if self.Addnew17.Acount.toPlainText() == '':
            self.checked_account17 = ''

        ##Select all이나 일부 체크박스가 선택된 경우
        else:
            self.checked_account17 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew17.Acount.toPlainText() + ')'

        if self.temp_TE == '':
                self.temp_TE = 0

        ##Checked_account의 유효성 체크
        if self.check_account(self.checked_account17) == False:
            return

        ## 예외 처리 - 중요성금액이 양수가 아닌 경우
        try:
            float(self.temp_TE)
        except ValueError:
            self.alertbox_open2('중요성금액')
            return

        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
            self.debitcredit = ''
        elif self.checkD.isChecked():
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        cursor = self.cnxn.cursor()

        ### JE Line
        if self.rbtn1.isChecked():
            sql_query = """
                            SET NOCOUNT ON				
                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                            GROUP BY CoA.GLAccountNumber				
                            SELECT COUNT(*) as cnt
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                #TMPCOA,			
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                            AND JournalEntries.JELINEID = Details.JENumberID 							
                            AND JournalEntries.PreparerID = JournalEntries.ApproverID			
                            AND ABS(JournalEntries.Amount) >= {TE} 		
                            {Account}				
                            {NewSQL}				
                            {DebitCredit}
                            {AutoManual}								
                            DROP TABLE #TMPCOA				            
                    """.format(field=self.selected_project_id, TE=self.temp_TE, Account=self.checked_account17,
                               DebitCredit=self.debitcredit, NewSQL=self.NewSQL,
                               AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### JE
        elif self.rbtn2.isChecked():
            sql_query = """
                                    SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT COUNT(*) as cnt
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 							
                                    AND Details.JEIdentifierID IN				
                                            (		
                                             SELECT DISTINCT Details.JEIdentifierID		
                                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                             WHERE JournalEntries.JELINEID = Details.JENumberID 
                                             AND JournalEntries.PreparerID = JournalEntries.ApproverID 
                                             AND ABS(JournalEntries.Amount) >= {TE}	
                                             {Account}	
                                             {NewSQL}			
                                             {DebitCredit}
                                             {AutoManual}			
                                                )					
                                    DROP TABLE #TMPCOA				           
                                """.format(field=self.selected_project_id, TE=self.temp_TE,
                                           Account=self.checked_account17,
                                           DebitCredit=self.debitcredit, NewSQL=self.NewSQL,
                                           AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        buttonReply = QMessageBox.information(self, '라인 수 확인',
                                              '라인 수 : ' + str(self.dataframe['cnt'].loc[0]) + '<br>',
                                              QMessageBox.Ok)

        if buttonReply == QMessageBox.Ok: self.dialog17.activateWindow()

    ### 제외 키워드 활성화 반영 함수
    def D14_LabelC(self, state):
        if state == 0:
            self.D14_Key2.clear()
            self.D14_Key2.setReadOnly(True)
        else:
            self.D14_Key2.setReadOnly(False)

    ### dialog4(시나리오 1번) 창 닫는 함수
    def dialog_close4(self):
        self.dialog4.close()

    ### dialog5(시나리오 2번) 창 닫는 함수
    def dialog_close5(self):
        self.dialog5.close()

    ### dialog6(시나리오 3번) 창 닫는 함수
    def dialog_close6(self):
        self.dialog6.close()

    ### dialog7(시나리오 4번) 창 닫는 함수
    def dialog_close7(self):
        self.dialog7.close()

    ### dialog8(시나리오 5번) 창 닫는 함수
    def dialog_close8(self):
        self.dialog8.close()

    ### dialog9(시나리오 6번) 창 닫는 함수
    def dialog_close9(self):
        self.dialog9.close()

    ### dialog10(시나리오 7번) 창 닫는 함수
    def dialog_close10(self):
        self.dialog10.close()

    ### dialog12(시나리오 8번) 창 닫는 함수
    def dialog_close12(self):
        self.dialog12.close()

    ### dialog13(시나리오 9번) 창 닫는 함수
    def dialog_close13(self):
        self.dialog13.close()

    ### dialog14(시나리오 10번) 창 닫는 함수
    def dialog_close14(self):
        self.dialog14.close()

    ### dialog15(시나리오 11번) 창 닫는 함수
    def dialog_close15(self):
        self.dialog15.close()

    ### dialog16(시나리오 12번) 창 닫는 함수
    def dialog_close16(self):
        self.dialog16.close()

    ### dialog17(시나리오 13번) 창 닫는 함수
    def dialog_close17(self):
        self.dialog17.close()

    ### Main 중앙에 데이터 프레임 출력 함수
    def Show_DataFrame_Group(self):
        tables = QGroupBox('데이터')
        self.setStyleSheet('QGroupBox  {color: white;}')
        font6 = tables.font()
        font6.setBold(True)
        tables.setFont(font6)
        box = QBoxLayout(QBoxLayout.TopToBottom)

        self.viewtable = QTableView(self)

        box.addWidget(self.viewtable)
        tables.setLayout(box)

        return tables

    ### 소요 시간 측정을 위한 타이머
    def Timer(self):
        self.secondTimer = 0
        self.timerVar.start()

    ### 소요 시간을 보여주는 함수
    def printTime(self):
        self.secondTimer += 1
        elapsetime = "Elapsed time : " + str(int(self.secondTimer / 3600)) + "h " + str(
            int(self.secondTimer / 60)) + "m " + str(
            self.secondTimer % 60) + "s"
        self.progressLabel.setText(elapsetime)

    ### 소요 시간 측정 창 닫기 버튼 클릭 시 모든 창을 닫게 하는 함수
    def pClose(self):
        for a in self.dialoglist:
            if a == 4:
                self.dialog4.close()
            elif a == 5:
                self.dialog5.close()
            elif a == 6:
                self.dialog6.close()
            elif a == 7:
                self.dialog7.close()
            elif a == 8:
                self.dialog8.close()
            elif a == 9:
                self.dialog9.close()
            elif a == 10:
                self.dialog10.close()
            elif a == 12:
                self.dialog12.close()
            elif a == 13:
                self.dialog13.close()
            elif a == 14:
                self.dialog14.close()
            elif a == 15:
                self.dialog15.close()
            elif a == 16:
                self.dialog16.close()
            elif a == 17:
                self.dialog17.close()
        self.Action.close()
        sys.exit()

    ### 소요 시간을 보여주는 창
    def doAction(self):
        self.Timer()
        self.Action = QDialog()
        self.Action.setStyleSheet('background-color : black;')
        lbl_img = QLabel()
        label = QLabel('Now Loading')
        label.setStyleSheet("font : bold 14pt; color: white;")
        self.progressLabel = QLabel("Elapsed time : 0h 0m 0s")
        self.progressLabel.setStyleSheet("font : bold 8pt; color: grey;")
        pixmap = QPixmap(self.resource_path('./Loading.png'))
        lbl_img.setPixmap(pixmap)

        self.pclosebtn = QPushButton("Close", self.Action)
        self.pclosebtn.setStyleSheet('color:white; background-image : url(./bar.png) ;')
        self.pclosebtn.clicked.connect(self.pClose)

        self.progressBar = QProgressBar()
        self.progressBar.setRange(0, 0)
        self.progressBar.setStyleSheet("QProgressBar::chunk "
                                       "{"
                                       "background-color: yellow;"
                                       "}")
        sub_layout = QHBoxLayout()
        sub_layout.addWidget(lbl_img)

        sub_layout2 = QVBoxLayout()
        sub_layout2.addWidget(label)
        sub_layout2.addWidget(self.progressLabel)
        sub_layout2.setAlignment(Qt.AlignCenter)

        sub_layout3 = QHBoxLayout()
        sub_layout3.addStretch(2)
        sub_layout3.addWidget(self.pclosebtn)

        sub_layout.addLayout(sub_layout2)
        main_layout = QVBoxLayout()
        main_layout.addLayout(sub_layout)
        main_layout.addWidget(self.progressBar)
        main_layout.addLayout(sub_layout3)

        self.Action.setLayout(main_layout)
        self.Action.setGeometry(700, 400, 400, 220)
        self.Action.setWindowFlags(Qt.FramelessWindowHint)
        self.Action.setWindowModality(Qt.NonModal)
        self.Action.show()

    ### 결과값과 관련한 팝업 함수 (시나리오 1번)
    def doneAction4(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가 ' + str(self.temp_N) + '회 이하인 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가 ' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe)) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(
                    self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가 ' + str(self.temp_N) + '회 이하인 전표가 '
                                                      + str(len(self.dataframe)) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog4.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe)) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog4.activateWindow()

        self.th4.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 2번)
    def doneAction5(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():

                buttonReply = QMessageBox.information(self, '라인수 추출', '- 당기('
                                                      + str(self.pname_year) + ')에 생성된 계정을 사용한 전표가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)



            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:

                buttonReply = QMessageBox.information(self, '라인수 추출', '- 당기('
                                                      + str(self.pname_year) + ')에 생성된 계정을 사용한 전표가 '
                                                      + str(len(self.dataframe))
                                                      + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(
                    self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 당기('
                                                      + str(self.pname_year) + ')에 생성된 계정을 사용한 전표가 '
                                                      + str(len(self.dataframe))
                                                      + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 당기('
                                                      + str(self.pname_year) + ')에 생성된 계정을 사용한 전표가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.pname_year) + ')에 생성된 계정을 사용한 전표가 '
                                                      + str(len(self.dataframe))
                                                      + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

        self.th5.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 3번)
    def doneAction6(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 시작 시점 : " + str(self.period1.text()) + " 종료 시점 : " + str(
                                                          self.period2.text())
                                                      + "에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 시작 시점 : " + str(self.period1.text()) + " 종료 시점 : " + str(
                                                          self.period2.text())
                                                      + "에 입력된 전표가 " + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 시작 시점 : " + str(self.period1.text()) + " 종료 시점 : " + str(
                                                          self.period2.text())
                                                      + "에 입력된 전표가 " + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog6.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 시작 시점 : " + str(self.period1.text()) + " 종료 시점 : " + str(
                                                          self.period2.text())
                                                      + "에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 시작 시점 : " + str(self.period1.text()) + " 종료 시점 : " + str(
                                                          self.period2.text())
                                                      + "에 입력된 전표가 " + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog6.activateWindow()

        self.th6.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 4번)
    def doneAction7(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog7.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog7.activateWindow()

        self.th7.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 5번)
    def doneAction8(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog8.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog8.activateWindow()

        self.th8.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 6번)
    def doneAction9(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### 추출 데이터가 존재하지 않을 경우
        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[전표작성 빈도수: " + str(self.tempN) + "," + " 중요성금액: " + str(
                self.tempTE) + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            model_refer = DataFrameModel(self.dataframe_refer)
            self.viewtable.setModel(model)
            ### JE Line 기준
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            ### JE 기준
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            ### JE Line 기준
            if self.rbtn1.isChecked():

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                      + "회 이하인 작성자에 의해 생성된 전표가 "
                                                      + str(len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액("
                                                      + str(self.tempTE) + ")을 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            ### JE 기준
            else:

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                      + "회 이하인 작성자에 의해 생성된 전표가 "
                                                      + str(len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액("
                                                      + str(self.tempTE) + ")을 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog9.activateWindow()

        ### 추출 데이터가 300건 초과일 경우
        else:
            ### JE Line 기준
            if self.rbtn1.isChecked():

                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                          + "회 이하인 작성자에 의해 생성된 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                          + "회 이하인 작성자에 의해 생성된 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog9.activateWindow()
            ### JE 기준
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                      + "회 이하인 작성자에 의해 생성된 전표가 "
                                                      + str(
                    len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액("
                                                      + str(
                    self.tempTE) + ")을 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog9.activateWindow()
        self.th9.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 7번)
    def doneAction10(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### 추출 데이터가 존재하지 않을 경우
        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': [" 중요성금액: " + str(
                    self.tempTE) + "] 라인수 " + str(
                    len(self.dataframe)) + "개입니다"]})

            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            ### JE Line 기준 추출 시
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE 기준 추출 시
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            if self.rbtn1.isChecked():

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "을 적용하였습니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "을 적용하였습니다.<br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog10.activateWindow()

        else:
            if self.rbtn1.isChecked():

                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(
                        self.tempTE) + "을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(self.tempTE) + "을 적용하였습니다.<br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog10.activateWindow()
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "을 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog10.activateWindow()
        self.th10.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 8-1번)
    def doneAction12(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### 추출 데이터가 존재하지 않을 경우
        elif 'No Data' in self.dataframe.columns.tolist():
            buttonReply = QMessageBox.information(self, "라인수 추출",
                                                  "[중요성 금액: " + str(self.temp_TE) +
                                                  "] 라인수 " + str(len(self.dataframe) - 1) + "개입니다",
                                                  QMessageBox.Ok)

        else:
            buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(
                self.temp_TE) + "] 라인수 " + str(len(self.dataframe)) + "개입니다", QMessageBox.Ok)

        if buttonReply == QMessageBox.Ok: self.dialog12.activateWindow()

        self.th12.join()

    @pyqtSlot(str)
    ### 결과값과 관련한 팝업 함수 (시나리오 8-2번)
    def doneActionC(self, cursortext):
        self.Action.close()
        self.timerVar.stop()
        self.Cursortext.setText(cursortext)

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### 추출 데이터가 존재하지 않을 경우
        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ['No Cursor']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(self.temp_TE) + "] 라인수 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(self.temp_TE) + "] 라인수 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok:
                self.dialog12.activateWindow()

        ### 추출 데이터가 300건 초과일 경우
        elif len(self.dataframe) > 300:
            if self.rbtn1.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(self.temp_TE) + "] 라인수 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> 추가 필터링이 필요해보입니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            elif self.rbtn2.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(self.temp_TE) + "] 라인수 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog12.activateWindow()

        else:
            if self.rbtn1.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(self.temp_TE) + "] 라인수 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            elif self.rbtn2.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(self.temp_TE) + "] 라인수 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다.<br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog12.activateWindow()
        self.thC.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 9번)
    def doneAction13(self):
        self.Action.close()
        self.timerVar.stop()

        ### 예외처리 3 - 최대 추출 라인수
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자' + str(self.temp_Continuous) + '로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자' + str(self.temp_Continuous) + '로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe)) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(
                                                          self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자' + str(self.temp_Continuous) + '로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe)) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog13.activateWindow()

        elif self.rbtn2.isChecked():
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자' + str(self.temp_Continuous) + '로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자' + str(self.temp_Continuous) + '로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe)) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')를 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog13.activateWindow()

        self.th13.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 10번)
    def doneAction14(self):
        self.Action.close()
        self.timerVar.stop()

        ### 제외단어 Activate 체크 유무에 따른 tempword
        if self.D14_Key2C.isChecked():
            tempword = ", " + str(self.baseKey2) + "이/가 제외"
        else:
            tempword = ''

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### 추출 데이터가 존재하지 않을 경우
        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[전표 적요 특정단어: " + str(self.baseKey) + "," + " 중요성금액: " + str(
                self.tempTE) + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            ### JE Line 기준
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                      + str(self.baseKey) + "이/가 포함"
                                                      + tempword + "된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                      + ")을 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            ### JE 기준
            else:
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                      + str(self.baseKey) + "이/가 포함"
                                                      + tempword + "된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                      + ")을 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog14.activateWindow()

        ### 추출 데이터가 300건 초과일 경우
        else:
            ### JE Line 기준
            if self.rbtn1.isChecked():
                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                          + str(self.baseKey) + "이/가 포함"
                                                          + tempword + "된 전표가 "
                                                          + str(len(self.dataframe))
                                                          + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br>  [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                          + str(self.baseKey) + "이/가 포함"
                                                          + tempword + "된 전표가 "
                                                          + str(len(self.dataframe))
                                                          + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog14.activateWindow()
            ### JE 기준
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                      + str(self.baseKey) + "이/가 포함"
                                                      + tempword + "된 전표가 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                      + ")을 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog14.activateWindow()
        self.th14.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 11번)
    def doneAction15(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### 추출 데이터가 존재하지 않을 경우
        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': [" 중요성금액: " + str(
                    self.tempTE) + "] 라인수 " + str(
                    len(self.dataframe)) + "개입니다"]})

            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            if self.rbtn1.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "을 적용하였습니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "을 적용하였습니다.<br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog15.activateWindow()

        else:
            if self.rbtn1.isChecked():
                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 " + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(
                        self.tempTE) + "을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(self.tempTE) + "을 적용하였습니다.<br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog15.activateWindow()
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "을 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog15.activateWindow()
        self.th15.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 12번)
    def doneAction16(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "차/대변 합계가 중요성 금액(" + str(self.temp_TE) + ")원 이상인 전표가 " + str(
                                                          len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "차/대변 합계가 중요성 금액(" + str(self.temp_TE) + ")원 이상인 전표가 " + str(
                                                          len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "차/대변 합계가 중요성 금액(" + str(self.temp_TE) + ")원 이상인 전표가 " + str(
                                                          len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog16.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "차/대변 합계가 중요성 금액(" + str(self.temp_TE) + ")원 이상인 전표가 " + str(
                                                          len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "차/대변 합계가 중요성 금액(" + str(self.temp_TE) + ")원 이상인 전표가 " + str(
                                                          len(self.dataframe))
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.temp_TE)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog16.activateWindow()

        self.th16.join()

    ### 결과값과 관련한 팝업 함수 (시나리오 13번)
    def doneAction17(self):
        self.Action.close()
        self.timerVar.stop()

        ### 결과값이 50만건 초과일 경우
        if len(self.dataframe) > 500000:
            self.alertbox_open3()

        ### JE Line 기준 추출 시
        elif self.rbtn1.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 전표입력자와 승인자가 동일한 전표를 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            ### 추출 데이터가 300건 초과일 경우
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 전표입력자와 승인자가 동일한 전표를 검토한 결과 '
                                                      + str(len(self.dataframe)) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(
                                                          self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 전표입력자와 승인자가 동일한 전표를 검토한 결과 '
                                                      + str(len(self.dataframe)) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog17.activateWindow()

        ### JE 기준 추출 시
        elif self.rbtn2.isChecked():

            ### 추출 데이터가 존재하지 않을 경우
            if 'No Data' in self.dataframe.columns.tolist():
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 전표입력자와 승인자가 동일한 전표를 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 전표입력자와 승인자가 동일한 전표를 검토한 결과 '
                                                      + str(len(self.dataframe)) + ' 건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE) + ')를 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog17.activateWindow()

        self.th17.join()

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 1번)
    def Thread4(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew4.SegmentBox1,
                                                                           self.Addnew4.SegmentBox2,
                                                                           self.Addnew4.SegmentBox3,
                                                                           self.Addnew4.SegmentBox4,
                                                                           self.Addnew4.SegmentBox5,
                                                                           self.Addnew4.UserDefine1,
                                                                           self.Addnew4.UserDefine2,
                                                                           self.Addnew4.UserDefine3,
                                                                           self.Addnew4.User, self.Addnew4.source,
                                                                           self.Manual, self.Auto)
        self.temp_N = self.D4_N.text()
        self.temp_TE = self.D4_TE.text()
        self.tempSheet = self.D4_Sheet.text()

        if self.Addnew4.Acount.toPlainText() == '':
            self.checked_account4 = ''

        else:
            self.checked_account4 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew4.Acount.toPlainText() + ')'

        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
            self.debitcredit = ''
        elif self.checkD.isChecked():  # Credit 이 0
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        ### 예외처리 1 - 필수값 입력 누락
        if self.temp_N == '' or self.tempSheet == '':
            self.alertbox_open()

        ### 예외처리 2 - 시트명 중복 확인 (JE Line)
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        ### 예외처리 3 - 시트명 중복 확인 (JE)
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        ### 쿼리 연동
        else:
            if self.temp_TE == '': self.temp_TE = 0
            if self.check_account(self.checked_account4) != False:
                try:
                    int(self.temp_N)
                    float(self.temp_TE)

                    self.doAction()
                    self.th4 = Thread(target=self.extButtonClicked4)
                    self.th4.daemon = True
                    self.th4.start()

                ### 예외처리 5 - 필수 입력값 타입 오류
                except ValueError:
                    try:
                        int(self.temp_N)
                        try:
                            float(self.temp_TE)
                        except:
                            self.alertbox_open2('중요성금액')
                    except:
                        try:
                            float(self.temp_TE)
                            self.alertbox_open2('계정사용 빈도수')
                        except:
                            self.alertbox_open2('계정사용 빈도수와 중요성금액')

    def ChangeInt(self, row):
        try:
            return str(int(row))
        except:
            return str(row)

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 2번)
    def Thread5(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew5.SegmentBox1,
                                                                           self.Addnew5.SegmentBox2,
                                                                           self.Addnew5.SegmentBox3,
                                                                           self.Addnew5.SegmentBox4,
                                                                           self.Addnew5.SegmentBox5,
                                                                           self.Addnew5.UserDefine1,
                                                                           self.Addnew5.UserDefine2,
                                                                           self.Addnew5.UserDefine3,
                                                                           self.Addnew5.User, self.Addnew5.source,
                                                                           self.Manual, self.Auto)

        ### 인풋 값 변수로 받아오기
        self.tempSheet = self.D5_Sheet.text()  # 필수값 ###시트명
        self.temp_TE = self.D5_TE.text()  ### 중요성금액

        ##Unselect all의 경우
        if self.Addnew5.Acount.toPlainText() == '':
            self.checked_account5 = "AND JournalEntries.GLAccountNumber IN ('')"  ###당기 생성 계정이 없는 경우 고려

        ##Select all이나 일부 체크박스가 선택된 경우
        else:
            self.checked_account5 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew5.Acount.toPlainText() + ')'

        ### 예외처리 1 - 필수값 입력 누락
        if self.tempSheet == '' or self.checked_account5 == '':
            self.alertbox_open()

        ### 예외처리 2 - 시트명 중복 확인 (JE Line)
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        ### 예외처리 3 - 시트명 중복 확인 (JE)
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            if self.temp_TE == '':
                self.temp_TE = 0

            ##Checked_account의 유효성 체크
            if self.check_account(self.checked_account5) == False:
                return

            try:
                float(self.temp_TE)

                if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                        not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                    self.debitcredit = ''
                elif self.checkD.isChecked():  # Credit 이 0
                    self.debitcredit = 'AND JournalEntries.Credit = 0'
                elif self.checkC.isChecked():  # Debit 이 0
                    self.debitcredit = 'AND JournalEntries.Debit = 0'

                self.doAction()
                self.th5 = Thread(target=self.extButtonClicked5)
                self.th5.daemon = True
                self.th5.start()

            ### 예외처리 5 - 필수 입력값 타입 오류
            except ValueError:
                self.alertbox_open2('중요성금액')

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 3번)
    def Thread6(self):

        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew6.SegmentBox1,
                                                                           self.Addnew6.SegmentBox2,
                                                                           self.Addnew6.SegmentBox3,
                                                                           self.Addnew6.SegmentBox4,
                                                                           self.Addnew6.SegmentBox5,
                                                                           self.Addnew6.UserDefine1,
                                                                           self.Addnew6.UserDefine2,
                                                                           self.Addnew6.UserDefine3,
                                                                           self.Addnew6.User, self.Addnew6.source,
                                                                           self.Manual, self.Auto)

        ### 중요성 금액
        self.temp_TE = self.D6_TE.text()

        ### 시나리오 번호
        self.tempSheet = self.D6_Sheet.text()

        ### 필수 입력값 누락 검토
        if self.period1.text() == '' or self.tempSheet == '' or self.period2.text() == '':
            self.alertbox_open()

        ### Result 시나리오 번호(시트명) 중복 검토
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        ### Journals 시나리오 번호(시트명) 중복 검토
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '':
                self.temp_TE = 0

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew6.Acount.toPlainText() == '':
                self.checked_account6 = ''
            else:
                self.checked_account6 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew6.Acount.toPlainText() + ')'

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account6) != False:
                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.temp_TE)

                    ### 시작/종료 날짜 정수로 입력했는지 확인
                    int(self.period1.text())
                    int(self.period2.text())

                    ### 시작/종료 시점 쿼리문에 적용할 수 있도록 변환
                    self.tempDate1 = "'" + self.period1.text() + "'"
                    self.tempDate2 = "'" + self.period2.text() + "'"

                    ### 시점 자릿수 확인(' 포함 10자리 여부 확인)
                    if len(str(self.tempDate1)) != 10:
                        self.alertbox_open19()
                    elif len(str(self.tempDate2)) != 10:
                        self.alertbox_open19()
                    else:
                        self.doAction()
                        self.th6 = Thread(target=self.extButtonClicked6)
                        self.th6.daemon = True
                        self.th6.start()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        float(self.temp_TE)  ### 중요성 금액이 실수가 아닌 경우
                    except:
                        try:
                            int(self.period1.text())
                            int(self.period2.text())
                            self.alertbox_open2('중요성 금액')
                        except:
                            self.alertbox_open2('입력일과 중요성 금액')  ### 중요성 금액과 입력일의 형식이 잘못되었을 경우
                    try:
                        int(self.period1.text())
                        int(self.period2.text())
                    except:
                        try:
                            float(self.temp_TE)
                            self.alertbox_open2('입력일')  ### 입력일의 형식이 잘못되었을 경우
                        except:
                            self.alertbox_open2('입력일과 중요성 금액')  ### 중요성 금액과 입력일의 형식이 잘못되었을 경우

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 4번)
    def Thread7(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew7.SegmentBox1,
                                                                           self.Addnew7.SegmentBox2,
                                                                           self.Addnew7.SegmentBox3,
                                                                           self.Addnew7.SegmentBox4,
                                                                           self.Addnew7.SegmentBox5,
                                                                           self.Addnew7.UserDefine1,
                                                                           self.Addnew7.UserDefine2,
                                                                           self.Addnew7.UserDefine3,
                                                                           self.Addnew7.User, self.Addnew7.source,
                                                                           self.Manual, self.Auto)

        ### 중요성 금액
        self.temp_TE = self.D7_TE.text()

        ### 시나리오 번호
        self.tempSheet = self.D7_Sheet.text()

        ### 필수 입력값 누락 검토
        if self.tempSheet == '':
            self.alertbox_open()

        ### Result 시나리오 번호(시트명) 중복 검토
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        ### Journals 시나리오 번호(시트명) 중복 검토
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        ### 입력일 / 전기일 모두 선택/미선택 여부 검토
        elif (self.Entry.isChecked() and self.Effective.isChecked()) or (
                not (self.Entry.isChecked()) and not (self.Effective.isChecked())):
            self.alertbox_open21()

        else:
            self.holiday = []  # 공휴일 리스트
            self.holiday_str = []  # 공휴일, 주말
            self.realDate_List = []  # 중복 제거 전 SQL 쿼리에 들어갈 리스트
            self.realDate_List_final = []  # SQL 쿼리에 들어갈 리스트

            ### 공휴일 추가
            self.holiday = [pytimekr.holidays(i) for i in range(2021, 2023)]
            for i in range(len(self.holiday)):
                for d in range(0, len(self.holiday[i])):
                    self.date_str = self.holiday[i][d].strftime('%Y-%m-%d')
                    self.holiday_str.append(self.date_str)

            ### 주말 추가
            self.start_date = date(2020, 1, 1)
            self.end_date = date(2022, 12, 31)
            self.delta = timedelta(days=1)
            while self.start_date <= self.end_date:
                if self.start_date.weekday() == 5 or self.start_date.weekday() == 6:
                    self.a = self.start_date.strftime('%Y-%m-%d')
                    self.holiday_str.append(self.a)
                self.start_date += self.delta

            ### 공휴일, 주말 yyyyMMdd 형식에 맞게 변환
            for i in range(0, len(self.holiday_str)):
                self.tempDate = []
                self.tempDate = str(self.holiday_str[i]).split('-')
                self.realDate = self.tempDate[0] + self.tempDate[1] + self.tempDate[2]
                self.realDate_List.append(self.realDate)

            ### 사용자 입력 일자 추가
            if self.D7_Date.toPlainText() != '':
                self.user_date = self.D7_Date.toPlainText().split(',')

            else:
                self.user_date = ''

            for a in self.user_date:
                a = a.strip()
                try:
                    int(a)
                    if len(a) == 8:  ### 날짜 형식이 yyyyMMdd일 경우만 추가
                        b = a
                        self.realDate_List.append(b)
                    else:
                        self.alertbox_open19()  ### 형식이 올바르지 않은 경우, 팝업
                        return
                except:
                    self.alertbox_open19()
                    return

            ### 날짜 중복 제거 완료 (self.realDate_List_final)
            self.realDate_List_final = set(self.realDate_List)

            ### 쿼리문에 적용할 수 있게끔 변환
            self.checked_date = ''
            for i in self.realDate_List_final:
                self.checked_date = self.checked_date + ',' + '\'' + i + '\''

            self.checked_date = self.checked_date[1:]

            self.checked_effective = 'AND JournalEntries.EffectiveDate IN (' + self.checked_date + ')'
            self.checked_entry = 'AND JournalEntries.EntryDate IN (' + self.checked_date + ')'

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():  # Debit 이 0
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 입력일을 선택했을 시, EntryDate 조건문
            if self.Entry.isChecked():
                self.tempState = self.checked_entry

            ### 전기일을 선택했을 시, EffectiveDate 조건문
            elif self.Effective.isChecked():
                self.tempState = self.checked_effective

            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '': self.temp_TE = 0

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew7.Acount.toPlainText() == '':
                self.checked_account7 = ''
            else:
                self.checked_account7 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew7.Acount.toPlainText() + ')'

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account7) != False:
                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.temp_TE)

                    self.doAction()
                    self.th7 = Thread(target=self.extButtonClicked7)
                    self.th7.daemon = True
                    self.th7.start()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    ### 중요성 금액이 실수가 아닌 경우
                    self.alertbox_open2('중요성 금액')

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 5번)
    def Thread8(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew8.SegmentBox1,
                                                                           self.Addnew8.SegmentBox2,
                                                                           self.Addnew8.SegmentBox3,
                                                                           self.Addnew8.SegmentBox4,
                                                                           self.Addnew8.SegmentBox5,
                                                                           self.Addnew8.UserDefine1,
                                                                           self.Addnew8.UserDefine2,
                                                                           self.Addnew8.UserDefine3,
                                                                           self.Addnew8.User, self.Addnew8.source,
                                                                           self.Manual, self.Auto)

        ### N일
        self.tempN = self.D8_N.text()
        ### 중요성 금액
        self.temp_TE = self.D8_TE.text()
        ### 시나리오 번호
        self.tempSheet = self.D8_Sheet.text()

        ### 필수 입력값 누락 검토
        if self.tempN == '' or self.tempSheet == '':
            self.alertbox_open()

        ### Result 시나리오 번호(시트명) 중복 검토
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        ### Journals 시나리오 번호(시트명) 중복 검토
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '':
                self.temp_TE = 0

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew8.Acount.toPlainText() == '':
                self.checked_account8 = ''
            else:
                self.checked_account8 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew8.Acount.toPlainText() + ')'

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account8) != False:
                try:
                    ### N 정수값인지 확인
                    int(self.tempN)

                    ### 중요성 금액 실수값인지 확인
                    float(self.temp_TE)

                    ### N값이 0이상 70만 이하인지 확인(비정상적인 값 예외처리)
                    if int(self.tempN) < 0 or int(self.tempN) > 700000:
                        self.alertbox_open13()
                        int('False')

                    else:
                        self.realNDate = int(self.tempN)
                        self.doAction()
                        self.th8 = Thread(target=self.extButtonClicked8)
                        self.th8.daemon = True
                        self.th8.start()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        int(self.tempN)
                        try:
                            float(self.temp_TE)  ### 중요성 금액이 실수가 아닌 경우
                        except:
                            self.alertbox_open2('중요성금액')
                    except:
                        try:
                            float(self.temp_TE)
                            self.alertbox_open2('N')  ### N이 정수가 아닌 경우
                        except:
                            self.alertbox_open2('중요성금액과 N')  ### 중요성 금액과 N의 형식이 잘못된 경우

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 6번)
    def Thread9(self):

        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew9.SegmentBox1,
                                                                           self.Addnew9.SegmentBox2,
                                                                           self.Addnew9.SegmentBox3,
                                                                           self.Addnew9.SegmentBox4,
                                                                           self.Addnew9.SegmentBox5,
                                                                           self.Addnew9.UserDefine1,
                                                                           self.Addnew9.UserDefine2,
                                                                           self.Addnew9.UserDefine3,
                                                                           self.Addnew9.User, self.Addnew9.source,
                                                                           self.Manual, self.Auto)
        self.tempN = self.D9_N.text()  # 전표 작성 빈도수 N회
        self.tempTE = self.D9_TE.text()  # 중요성 금액
        self.tempSheet = self.D9_Sheet.text()  # 시나리오 번호

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew9.Acount.toPlainText() == '':
            self.checked_account9 = ''

        else:
            self.checked_account9 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew9.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.tempN == '' or self.tempSheet == '':
            self.alertbox_open()

        ### Result & Reference 시나리오 번호(시트명) 중복 검토
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        ### Journals 시나리오 번호(시트명) 중복 검토
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        ### 중요성 금액 미입력시 0원
        else:
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account9) != False:
                try:
                    ### N 정수값인지 확인
                    int(self.tempN)
                    ### 중요성 금액 실수값인지 확인
                    float(self.tempTE)

                    ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
                    if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                            not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                        self.debitcredit = ''

                    ### Debit을 선택했을 시, Credit이 0원
                    elif self.checkD.isChecked():
                        self.debitcredit = 'AND JournalEntries.Credit = 0'
                    ### Credit을 선택했을 시, Debit이 0원
                    elif self.checkC.isChecked():
                        self.debitcredit = 'AND JournalEntries.Debit = 0'

                    self.doAction()
                    self.th9 = Thread(target=self.extButtonClicked9)
                    self.th9.daemon = True
                    self.th9.start()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        int(self.tempN)
                        try:
                            float(self.tempTE)
                        except:
                            self.alertbox_open4('중요성금액을 숫자로만 입력해주시기 바랍니다.')  # 중요성금액이 실수가 아닌 경우
                    except:
                        try:
                            float(self.tempTE)
                            self.alertbox_open4('작성빈도수를 숫자로만 입력해주시기 바랍니다.')  # 작성빈도수가 정수가 아닌 경우
                        except:
                            self.alertbox_open4('작성빈도수와 중요성금액을 숫자로만 입력해주시기 바랍니다.')  # 중요성금액이 실수가 아니고 작성빈도수가 정수가 아닌 경우

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 7번)
    def Thread10(self):

        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew10.SegmentBox1,
                                                                           self.Addnew10.SegmentBox2,
                                                                           self.Addnew10.SegmentBox3,
                                                                           self.Addnew10.SegmentBox4,
                                                                           self.Addnew10.SegmentBox5,
                                                                           self.Addnew10.UserDefine1,
                                                                           self.Addnew10.UserDefine2,
                                                                           self.Addnew10.UserDefine3,
                                                                           self.Addnew10.User, self.Addnew10.source,
                                                                           self.Manual, self.Auto)
        self.tempTE = self.D10_TE.text()
        self.tempSheet = self.D10_Sheet.text()

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew10.Acount.toPlainText() == '':
            self.checked_account10 = ''

        else:
            self.checked_account10 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew10.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.tempSheet == '':
            self.alertbox_open()

        elif self.Addnew10.User.text() == '':
            self.alertbox_open()

        ### 시트명 중복 확인
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            ### 중요성 금액 미입력시 0원
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력값 검토
            if self.check_account(self.checked_account10) != False:
                try:
                    float(self.tempTE)
                    if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                            not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                        self.debitcredit = ''
                    elif self.checkD.isChecked():
                        self.debitcredit = 'AND JournalEntries.Credit = 0'
                    elif self.checkC.isChecked():  # Debit 이 0
                        self.debitcredit = 'AND JournalEntries.Debit = 0'
                    self.doAction()
                    self.th10 = Thread(target=self.extButtonClicked10)
                    self.th10.daemon = True
                    self.th10.start()

                except ValueError:
                    self.alertbox_open4("중요성금액 값을 숫자로만 입력해주시기 바랍니다.")

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 8-1번)
    def Thread12(self):
        ## 수자동 선택 버튼을 모두 클릭하거나 모두 클릭하지 않은 경우
        if (self.Manual.isChecked() and self.Auto.isChecked()) or (
                not (self.Manual.isChecked()) and not (self.Auto.isChecked())):
            self.ManualAuto = ''

        ## 수동 버튼을 클릭한 경우
        elif self.Manual.isChecked():
            self.ManualAuto = "AND Details.SystemManualIndicator = 'Manual' "

        ## 자동 버튼을 클릭한 경우
        elif self.Auto.isChecked():
            self.ManualAuto = "AND Details.SystemManualIndicator = 'System' "

        ### 중요성 금액
        self.temp_TE = self.D12_TE.text()

        ### 시나리오 번호
        self.temp_Sheet = self.D12_Sheet.text()

        ## 예외 처리 - 필수 입력값 누락
        if self.temp_Sheet == '' or self.Addnew12_A.Acount.toPlainText() == 'AND LVL4.GL_Account_Number IN ()' or self.Addnew12_A.Acount.toPlainText() == '':
            self.alertbox_open()

        ## 예외 처리 - 중복된 시트명
        elif self.combo_sheet.findText(self.temp_Sheet + '_Reference') != -1:
            self.alertbox_open5()

        else:
            ### 중요성 금액 미입력시 0원
            if self.temp_TE == '':
                self.temp_TE = 0

            try:
                ### 중요성 금액 실수값인지 확인
                float(self.temp_TE)

                ## 예외 처리 - 기능영역이 존재하지 않음에도 기능영역을 체크한 경우
                if self.checkF.isChecked():

                    check_CoAsegment_query = """SELECT Segment01 FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts]""".format(
                        field=self.selected_project_id)
                    check_CoA = pd.read_sql(check_CoAsegment_query, self.cnxn)

                    if check_CoA.iloc[:, 0].isnull().sum() == len(check_CoA):
                        self.alertbox_open20()
                        return

                ### 계정 B 미입력 시, 계정 B 쿼리 조건문 삭제
                if self.Addnew12_B.Acount.toPlainText() == 'AND LVL4.Analysis_GL_Account_Number NOT IN ()' or self.Addnew12_B.Acount.toPlainText() == '':
                    self.checked_accountA = 'AND LVL4.GL_Account_Number IN (' + self.Addnew12_A.Acount.toPlainText() + ')'
                    self.checked_accountB = ''
                    self.tempStateB = ''

                ### 계정 B 입력시
                else:
                    self.checked_accountA = 'AND LVL4.GL_Account_Number IN (' + self.Addnew12_A.Acount.toPlainText() + ')'
                    self.checked_accountB = 'AND LVL4.Analysis_GL_Account_Number NOT IN (' + self.Addnew12_B.Acount.toPlainText() + ')'

                    ### 계정 B 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
                    if ((self.checkC2.isChecked()) and (self.checkD2.isChecked())) or (
                            not (self.checkC2.isChecked()) and not (self.checkD2.isChecked())):
                        self.tempStateB = 'AND LVL4.Analysis_Position IN (' + "'" + 'Credit' + "'" + "," + "'" + 'Debit' + "')"

                    ### 계정 B Credit 선택 시
                    elif self.checkC2.isChecked():
                        self.tempStateB = 'AND LVL4.Analysis_Position IN (' + "'" + 'Credit' + "')"

                    ### 계정 B Debit 선택 시
                    elif self.checkD2.isChecked():
                        self.tempStateB = 'AND LVL4.Analysis_Position IN (' + "'" + 'Debit' + "')"

                ### 계정 A 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
                if ((self.checkC1.isChecked()) and (self.checkD1.isChecked())) or (
                        not (self.checkC1.isChecked()) and not (self.checkD1.isChecked())):
                    self.tempStateA = 'AND LVL4.GL_Account_Position IN (' + "'" + 'Credit' + "'" + "," + "'" + 'Debit' + "'" + ')'

                ### 계정 A Credit 선택 시
                elif self.checkC1.isChecked():
                    self.tempStateA = 'AND LVL4.GL_Account_Position =' + "'" + 'Credit' + "'"

                ### 계정 A Debit 선택 시
                elif self.checkD1.isChecked():
                    self.tempStateA = 'AND LVL4.GL_Account_Position =' + "'" + 'Debit' + "'"

                ### 계정 A,B 입력 값 검토
                if self.check_account2(self.checked_accountA, self.checked_accountB) != False:
                    self.doAction()
                    self.th12 = Thread(target=self.extButtonClicked12)
                    self.th12.daemon = True
                    self.th12.start()

            ### 추가 예외처리 (팝업)
            except ValueError:
                self.alertbox_open2('중요성 금액')  ### 중요성 금액이 실수가 아닌 경우

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 8-2번)
    def ThreadC(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.AddnewC.SegmentBox1,
                                                                           self.AddnewC.SegmentBox2,
                                                                           self.AddnewC.SegmentBox3,
                                                                           self.AddnewC.SegmentBox4,
                                                                           self.AddnewC.SegmentBox5,
                                                                           self.AddnewC.UserDefine1,
                                                                           self.AddnewC.UserDefine2,
                                                                           self.AddnewC.UserDefine3,
                                                                           self.AddnewC.User, self.AddnewC.source,
                                                                           self.ManualC, self.AutoC)
        self.tempSheet = self.D12_Sheetc.text()
        self.cursorpath = self.cursorCondition.text()
        self.temp_TE = self.D12C_TE.text()

        if self.temp_TE == '':
            self.temp_TE = 0

        ## 예외 처리 - 필수 입력값 누락
        if self.listCursor.currentText() == '' or self.tempSheet == '' or self.cursorpath == '':
            self.alertbox_open()

        ## 예외 처리 - 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        ## 예외 처리 - 커서 지정 경로 상에 파일이 존재하지 않을 경우
        elif not os.path.isfile(self.cursorpath):
            self.MessageBox_Open("경로에 해당 파일이 존재하지 않습니다.")

        ## JE와 CoA 상에 기능영역이 존재하는 경우
        elif self.checkF2.isChecked():

            ## 예외 처리 - 기능영역이 존재하지 않음에도 기능영역을 체크한 경우
            if self.checkF.isChecked():

                check_CoAsegment_query = """SELECT Segment01 FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts]""".format(
                    field=self.selected_project_id)
                check_CoA = pd.read_sql(check_CoAsegment_query, self.cnxn)

                if check_CoA.iloc[:, 0].isnull().sum() == len(check_CoA):
                    self.alertbox_open20()
                    return

            try:
                float(self.temp_TE)
                self.wbC = self.wb2.parse(self.listCursor.currentText())

                ## 예외 처리 - 선택된 sheet가 커서 reference 시트가 아닌 경우
                if len(self.wbC.columns) != 17:
                    self.alertbox_open4('Cursor 필드가 존재하지 않습니다.')

                ## 예외 처리 - 비경상적 계정이 아무 것도 선택되지 않은 경우
                elif self.wbC['비경상적계정 선택여부'].notnull().any() == False:
                    self.alertbox_open4('Check된 조건이 없습니다.')

                ## 예외 처리 - 필수 커서 입력 값에 Null이 존재하는 경우
                elif self.wbC.iloc[:, [1, 5, 8, 12]].isnull().any().any():
                    self.alertbox_open4('필요 조건 필드를 충족하지 않습니다.')
                else:
                    self.doAction()
                    self.thC = Thread(target=self.extButtonClickedC)
                    self.thC.start()

            except:
                self.alertbox_open2('중요성금액')

        ## JE와 CoA 상에 기능영역이 존재하지 않는 경우
        else:
            try:
                float(self.temp_TE)
                self.wbC = self.wb2.parse(self.listCursor.currentText())

                ## 예외 처리 - 선택된 sheet가 커서 reference 시트가 아닌 경우
                if len(self.wbC.columns) != 15:
                    self.alertbox_open4('Cursor 필드가 존재하지 않습니다.')

                ## 예외 처리 - 비경상적 계정이 아무 것도 선택되지 않은 경우
                elif self.wbC['비경상적계정 선택여부'].notnull().any() == False:
                    self.alertbox_open4('Check된 조건이 없습니다.')

                ## 예외 처리 - 필수 커서 입력 값에 Null이 존재하는 경우
                elif self.wbC.iloc[:, [0, 4, 6, 10]].isnull().any().any():
                    self.alertbox_open4('필요 조건 필드를 충족하지 않습니다.')
                else:
                    self.doAction()
                    self.thC = Thread(target=self.extButtonClickedC)
                    self.thC.start()

            except:
                self.alertbox_open2('중요성금액')

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 9번)
    def Thread13(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew13.SegmentBox1,
                                                                           self.Addnew13.SegmentBox2,
                                                                           self.Addnew13.SegmentBox3,
                                                                           self.Addnew13.SegmentBox4,
                                                                           self.Addnew13.SegmentBox5,
                                                                           self.Addnew13.UserDefine1,
                                                                           self.Addnew13.UserDefine2,
                                                                           self.Addnew13.UserDefine3,
                                                                           self.Addnew13.User, self.Addnew13.source,
                                                                           self.Manual, self.Auto)

        self.temp_Continuous = self.text_continuous.toPlainText()  # 필수
        self.temp_Continuous = str(self.temp_Continuous).strip()
        self.temp_TE = self.D13_TE.text()
        self.tempSheet = self.D13_Sheet.text()  # 필수

        ##Unselect all의 경우
        if self.Addnew13.Acount.toPlainText() == '':
            self.checked_account13 = ''

        ##Select all이나 일부 체크박스가 선택된 경우
        else:
            self.checked_account13 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew13.Acount.toPlainText() + ')'

        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
            self.debitcredit = ''
        elif self.checkD.isChecked():
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        ### 예외처리 1 - 필수값 누락
        if self.temp_Continuous == '' or self.tempSheet == '':
            self.alertbox_open()

        ### 예외처리 2 - 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            if self.temp_TE == '':
                self.temp_TE = 0

            ##Checked_account의 유효성 체크
            if self.check_account(self.checked_account13) == False:
                return

            try:
                float(self.temp_TE)
            except ValueError:
                self.alertbox_open2('중요성금액')
                return

            try:
                self.temp_Continuous = re.sub(r"[\s]+", '', self.temp_Continuous)
            except:
                self.MessageBox_Open("연속된 자릿수 입력이 잘못되었습니다.")
                return

            self.temp_Continuous = self.temp_Continuous.split(',')

            for i in range(len(self.temp_Continuous)):
                ### 예외처리 3 - 숫자가 아닌 값 입력한 경우
                try:
                    int(self.temp_Continuous[i])
                except ValueError:
                    self.alertbox_open2("연속된 자릿수")
                    return

            self.filter_Continuous = ""
            for x in self.temp_Continuous:
                self.filter_Continuous += f"OR RIGHT(FLOOR(JournalEntries.Amount), {len(x)}) IN (\'{x}\')"
            self.filter_Continuous = 'AND (' + self.filter_Continuous[3:] + ')'

            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                self.debitcredit = ''
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'
            elif self.checkC.isChecked():  # Debit 이 0
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            self.doAction()
            self.th13 = Thread(target=self.extButtonClicked13)
            self.th13.daemon = True
            self.th13.start()

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 10번)
    def Thread14(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew14.SegmentBox1,
                                                                           self.Addnew14.SegmentBox2,
                                                                           self.Addnew14.SegmentBox3,
                                                                           self.Addnew14.SegmentBox4,
                                                                           self.Addnew14.SegmentBox5,
                                                                           self.Addnew14.UserDefine1,
                                                                           self.Addnew14.UserDefine2,
                                                                           self.Addnew14.UserDefine3,
                                                                           self.Addnew14.User, self.Addnew14.source,
                                                                           self.Manual, self.Auto)

        self.tempTE = self.D14_TE.text()  # 중요성 금액
        self.tempSheet = self.D14_Sheet.text()  # 시나리오 번호

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew14.Acount.toPlainText() == '':
            self.checked_account14 = ''

        else:
            self.checked_account14 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew14.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.tempSheet == '' or self.D14_Key.text().strip() == '':
            self.alertbox_open()

        ### 제외 키워드를 activate한 상태에서 제외 키워드를 입력하지 않을 경우, 경고창 생성
        elif self.D14_Key2C.isChecked() and self.D14_Key2.text().strip() == '':
            self.alertbox_open6()

        ### Result 시나리오 번호(시트명) 중복 검토
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        ### Journals 시나리오 번호(시트명) 중복 검토
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        ### 콤마(,) 구분자를 이용하여 전표 적요 특정단어 입력 (포함 단어)
        else:
            self.baseKey = self.D14_Key.text().split(',')
            self.baseKey_clean = []
            for a in self.baseKey:
                a = a.strip()
                if a.upper() == '[NULL]':
                    b = "((JournalEntries.JEDescription LIKE '' OR JournalEntries.JEDescription LIKE ' ' OR JournalEntries.JEDescription IS NULL)" \
                        "AND (JournalEntries.JELineDescription LIKE '' OR JournalEntries.JELineDescription LIKE ' ' OR JournalEntries.JELineDescription IS NULL))"
                elif a == '':
                    continue
                else:
                    b = "(JournalEntries.JEDescription LIKE N'%" + a + "%' OR JournalEntries.JELineDescription LIKE N'%" + a + "%')"
                self.baseKey_clean.append(b)

            ### 콤마(,) 구분자를 이용하여 전표 적요 특정단어 입력 (제외 단어)
            self.baseKey2 = self.D14_Key2.text().split(',')
            self.baseKey2_clean = []
            if self.D14_Key2C.isChecked():
                for a in self.baseKey2:
                    a = a.strip()
                    if a.upper() == '[NULL]':
                        b = "(NOT (JournalEntries.JEDescription LIKE '' OR JournalEntries.JEDescription LIKE ' ' OR JournalEntries.JEDescription IS NULL)" \
                            "OR NOT (JournalEntries.JELineDescription LIKE '' OR JournalEntries.JELineDescription LIKE ' ' OR JournalEntries.JELineDescription IS NULL))"
                    elif a == '':
                        continue
                    else:
                        b = "(NOT(JournalEntries.JEDescription LIKE N'%" + a + "%' OR JournalEntries.JELineDescription LIKE N'%" + a + "%'))"
                    self.baseKey2_clean.append(b)
                self.tempKey = 'AND (' + str('OR '.join(self.baseKey_clean)) + ') AND (' + str(
                    ' AND '.join(self.baseKey2_clean)) + ')'

            else:
                self.tempKey = 'AND (' + str(' OR '.join(self.baseKey_clean)) + ')'

            ### 중요성 금액 미입력시 0원
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account14) != False:

                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.tempTE)

                    ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
                    if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                            not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                        self.debitcredit = ''

                    ### Debit을 선택했을 시, Credit이 0원
                    elif self.checkD.isChecked():
                        self.debitcredit = 'AND JournalEntries.Credit = 0'
                    ### Credit을 선택했을 시, Debit이 0원
                    elif self.checkC.isChecked():
                        self.debitcredit = 'AND JournalEntries.Debit = 0'

                    self.doAction()
                    self.th14 = Thread(target=self.extButtonClicked14)
                    self.th14.daemon = True
                    self.th14.start()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    try:
                        float(self.tempTE)
                    except:
                        self.alertbox_open4('중요성금액 값을 숫자로만 입력해주시기 바랍니다.')  # 중요성금액이 실수가 아닌 경우

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 11번)
    def Thread15(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew15.SegmentBox1,
                                                                           self.Addnew15.SegmentBox2,
                                                                           self.Addnew15.SegmentBox3,
                                                                           self.Addnew15.SegmentBox4,
                                                                           self.Addnew15.SegmentBox5,
                                                                           self.Addnew15.UserDefine1,
                                                                           self.Addnew15.UserDefine2,
                                                                           self.Addnew15.UserDefine3,
                                                                           self.Addnew15.User, self.Addnew15.source,
                                                                           self.Manual, self.Auto)
        self.tempTE = self.D15_TE.text() # 중요성 금액
        self.tempSheet = self.D15_Sheet.text() # 시나리오 번호

        ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
        if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
            self.debitcredit = ''
        elif self.checkD.isChecked():
            self.debitcredit = 'AND JournalEntries.Credit = 0'
        elif self.checkC.isChecked():  # Debit 이 0
            self.debitcredit = 'AND JournalEntries.Debit = 0'

        ### 계정 미선택시 계정 조건 제거
        if self.Addnew15.Acount.toPlainText() == '':
            self.checked_account15 = ''

        else:
            self.checked_account15 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew15.Acount.toPlainText() + ')'

        ### 필수 입력값 누락 검토
        if self.tempSheet == '':
            self.alertbox_open()

        ### 시트명 중복 확인
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            if self.tempTE == '': self.tempTE = 0

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account15) != False:

                try:
                    ### 중요성 금액 실수값인지 확인
                    float(self.tempTE)
                    self.doAction()
                    self.th15 = Thread(target=self.extButtonClicked15)
                    self.th15.daemon = True
                    self.th15.start()

                except ValueError:
                    self.alertbox_open4("중요성금액 값을 숫자로만 입력해주시기 바랍니다.")

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 12번)
    def Thread16(self):
        ### Segment, UserDefine, 전표입력자, Source, 수자동 설정
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew16.SegmentBox1,
                                                                           self.Addnew16.SegmentBox2,
                                                                           self.Addnew16.SegmentBox3,
                                                                           self.Addnew16.SegmentBox4,
                                                                           self.Addnew16.SegmentBox5,
                                                                           self.Addnew16.UserDefine1,
                                                                           self.Addnew16.UserDefine2,
                                                                           self.Addnew16.UserDefine3,
                                                                           self.Addnew16.User, self.Addnew16.source,
                                                                           self.Manual, self.Auto)

        ### 중요성 금액
        self.temp_TE = self.D16_TE.text()

        ### 시나리오 번호
        self.tempSheet = self.D16_Sheet.text()

        ### 필수 입력값 누락 검토
        if self.temp_TE.strip() == '' or self.tempSheet == '':
            self.alertbox_open()

        ### Result 시나리오 번호(시트명) 중복 검토
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:

            ### 시작일이 yyyyMMdd 형식이 아닌 경우 - 숫자가 아닌 경우
            if not self.period1.text().strip().isdigit() and self.period1.text().strip() != '':
                self.alertbox_open19(); return

            ### 종료일이 yyyyMMdd 형식이 아닌 경우 - 숫자가 아닌 경우
            if not self.period2.text().strip().isdigit() and self.period2.text().strip() != '':
                self.alertbox_open19(); return

            ### 시작일이 yyyyMMdd 형식이 아닌 경우 - 8자리가 아닌 경우
            if len(self.period1.text().strip()) != 8 and len(self.period1.text().strip()) != 0:
                self.alertbox_open19(); return

            ### 종료일이 yyyyMMdd 형식이 아닌 경우 - 8자리가 아닌 경우
            if len(self.period2.text().strip()) != 8 and len(self.period2.text().strip()) != 0:
                self.alertbox_open19(); return

            self.EntryDate = ''
            self.subEntryDate = ''
            ### 시작일이 입력된 경우
            if self.period1.text().strip() != '':
                self.EntryDate += 'AND JournalEntries.EntryDate >= ' + "'" + self.period1.text().strip() + "'"
                self.subEntryDate += 'AND JournalEntries1.EntryDate >= ' + "'" + self.period1.text().strip() + "'"
            ### 종료일이 입력된 경우
            if self.period2.text().strip() != '':
                self.EntryDate += 'AND JournalEntries.EntryDate <= ' + "'" + self.period2.text().strip() + "'"
                self.subEntryDate += 'AND JournalEntries1.EntryDate <= ' + "'" + self.period2.text().strip() + "'"

            ### 계정 미선택시 계정 조건 제거
            if self.Addnew16.Acount.toPlainText() == '':
                self.checked_account16 = ''

            else:
                self.checked_account16 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew16.Acount.toPlainText() + ')'

            ### 차대변 체크박스 모두 선택 / 미선택 시, 차대변 조건 제거
            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):
                self.debitcredit = ''

            ### Debit을 선택했을 시, Credit이 0원
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'

            ### Credit을 선택했을 시, Debit이 0원
            elif self.checkC.isChecked():
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            self.sub_checked_account16 = re.sub('JournalEntries.', 'JournalEntries1.', self.checked_account16)
            self.sub_NewSQL = re.sub('JournalEntries.', 'JournalEntries1.', self.NewSQL)
            self.sub_debitcredit = re.sub('JournalEntries.', 'JournalEntries1.', self.debitcredit)
            self.sub_ManualAuto = re.sub('Details.', 'Details1.', self.ManualAuto)

            ### 계정 입력 값 검토
            if self.check_account(self.checked_account16) != False:
                try:
                    ### 중요성 금액 실수값인지 확인
                    if float(self.temp_TE) == 0.:
                        self.alertbox_open7(); return
                    self.doAction()
                    self.th16 = Thread(target=self.extButtonClicked16)
                    self.th16.daemon = True
                    self.th16.start()

                ### 추가 예외처리 (팝업)
                except ValueError:
                    self.alertbox_open2('중요성 금액')  ### 중요성 금액이 실수가 아닌 경우

    ### extraction버튼 클릭 시 유효성 확인 및 Thread 시작 (시나리오 13번)
    def Thread17(self):
        self.NewSQL, self.NewSelect, self.ManualAuto = self.NewQueryConcat(self.Addnew17.SegmentBox1,
                                                                           self.Addnew17.SegmentBox2,
                                                                           self.Addnew17.SegmentBox3,
                                                                           self.Addnew17.SegmentBox4,
                                                                           self.Addnew17.SegmentBox5,
                                                                           self.Addnew17.UserDefine1,
                                                                           self.Addnew17.UserDefine2,
                                                                           self.Addnew17.UserDefine3,
                                                                           self.Addnew17.User, self.Addnew17.source,
                                                                           self.Manual, self.Auto)
        self.temp_TE = self.D17_TE.text()
        self.tempSheet = self.D17_Sheet.text()

        ##Unselect all의 경우
        if self.Addnew17.Acount.toPlainText() == '':
            self.checked_account17 = ''

        ##Select all이나 일부 체크박스가 선택된 경우
        else:
            self.checked_account17 = 'AND JournalEntries.GLAccountNumber IN (' + self.Addnew17.Acount.toPlainText() + ')'

        ### 예외처리 1 - 필수값 누락
        if self.tempSheet == '':
            self.alertbox_open()

        ### 예외처리 2 - 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        ### Journals 시나리오 번호(시트명) 중복 검토
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:
            if self.temp_TE == '':
                self.temp_TE = 0

            ##Checked_account의 유효성 체크
            if self.check_account(self.checked_account17) == False:
                return

            ## 예외 처리 - 중요성금액이 양수가 아닌 경우
            try:
                float(self.temp_TE)
            except ValueError:
                self.alertbox_open2('중요성금액')
                return

            if (self.checkD.isChecked() and self.checkC.isChecked()) or (
                    not (self.checkD.isChecked()) and not (self.checkC.isChecked())):  # Credit 이 0
                self.debitcredit = ''
            elif self.checkD.isChecked():
                self.debitcredit = 'AND JournalEntries.Credit = 0'
            elif self.checkC.isChecked():  # Debit 이 0
                self.debitcredit = 'AND JournalEntries.Debit = 0'

            self.doAction()
            self.th17 = Thread(target=self.extButtonClicked17)
            self.th17.daemon = True
            self.th17.start()

    ### 저장된 시트 목록 중 선택한 시트에 해당하는 데이터 프레임을 보여주는 함수
    def Sheet_ComboBox_Selected(self, text):

        model = DataFrameModel(self.scenario_dic[text])
        self.viewtable.setModel(model)

    ### 저장된 시트 목록 중 사용자가 특정 시트를 삭제하고자 할 경우 실행되는 함수
    def RemoveSheetButton_Clicked(self):

        ### 예외 처리 - 삭제할 Sheet가 없는 경우
        if not self.combo_sheet:
            self.MessageBox_Open("삭제할 Sheet가 없습니다.")
            return

        ### Sheet 정보 삭제
        del self.scenario_dic[self.combo_sheet.currentText()]
        ### Query 정보 삭제
        self.my_query.drop(labels=[self.combo_sheet.currentText()], axis=0, inplace=True)
        ### Sheet name 콤보박스에서 해당 sheet 삭제
        self.combo_sheet.removeItem(self.combo_sheet.currentIndex())
        gc.collect()

        if not self.combo_sheet:
            self.dataframe = pd.DataFrame({'No Sheet': []})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
        else:
            model = DataFrameModel(self.scenario_dic[self.combo_sheet.currentText()])
            self.viewtable.setModel(model)

    ### 시트 저장 관련 기능을 담고 있는 그룹 박스
    def Save_Buttons_Group(self):
        groupbox = QGroupBox("저장")
        font_groupbox = groupbox.font()
        font_groupbox.setBold(True)
        groupbox.setFont(font_groupbox)
        self.setStyleSheet('QGroupBox  {color: white;}')

        ### RemoveSheet 버튼
        RemoveSheet_button = QPushButton('Remove Sheet')
        RemoveSheet_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        RemoveSheet_button.setStyleSheet('color:white;background-image : url(./bar.png)')
        font_RemoveSheet = RemoveSheet_button.font()
        font_RemoveSheet.setBold(True)
        RemoveSheet_button.setFont(font_RemoveSheet)

        ### Sheet Name 라벨
        label_sheet = QLabel("Sheet names: ", self)
        font_sheet = label_sheet.font()
        font_sheet.setBold(True)
        label_sheet.setFont(font_sheet)
        label_sheet.setStyleSheet('color:white;')

        ### 시나리오 Sheet를 표현할 콤보박스
        self.combo_sheet = QComboBox(self)

        ### Save 버튼
        export_file_button = QPushButton("Save", self)
        export_file_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        font_export_button = export_file_button.font()
        font_export_button.setBold(True)
        export_file_button.setFont(font_export_button)
        export_file_button.setStyleSheet('color:white;background-image : url(./bar.png)')

        ### 버튼 클릭 or 콤보박스 선택시 발생하는 시그널 함수들
        RemoveSheet_button.clicked.connect(self.RemoveSheetButton_Clicked)
        RemoveSheet_button.setShortcut("Ctrl+R")
        export_file_button.clicked.connect(self.saveFile)
        export_file_button.setShortcut("Ctrl+S")
        self.combo_sheet.activated[str].connect(self.Sheet_ComboBox_Selected)

        ### Layout 쌓기
        layout = QHBoxLayout()
        layout.addWidget(label_sheet, stretch=1)
        layout.addWidget(self.combo_sheet, stretch=4)
        layout.addWidget(RemoveSheet_button, stretch=1)
        layout.addWidget(export_file_button, stretch=1)
        groupbox.setLayout(layout)

        return groupbox

    def extButtonClicked4(self):
        cursor = self.cnxn.cursor()

        ### JE Line - Result
        if self.rbtn1.isChecked():

            sql_refer = """
                                SET NOCOUNT ON
                                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                GROUP BY CoA.GLAccountNumber
                                        SELECT 
                                            JournalEntries.GLAccountNumber AS 계정코드
                                            , MAX(#TMPCOA.GLAccountName) AS 계정명
                                            , COUNT(JournalEntries.GLAccountNumber) AS CNT
                                            , SUM(Debit) Sum_of_Debit
                                            , SUM(Credit) Sum_of_Credit				
                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                        AND JournalEntries.JELINEID = Details.JENumberID 
                                        AND JournalEntries.GLAccountNumber IN				
                                        (			
                                            SELECT DISTINCT JournalEntries.GLAccountNumber			
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                            WHERE JournalEntries.JELINEID = Details.JENumberID 
                                            GROUP BY JournalEntries.GLAccountNumber
                                            HAVING COUNT(JournalEntries.GLAccountNumber) <= {N}
                                        ) AND ABS(JournalEntries.Amount) >= {TE}
                                        {Account}
                                        {NewSQL}
                                        {AutoManual}
                                        {DebitCredit}
                                        GROUP BY JournalEntries.GLAccountNumber	
                                        ORDER BY JournalEntries.GLAccountNumber
                                        DROP TABLE #TMPCOA
                                    """.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                                               Account=self.checked_account4, NewSQL=self.NewSQL,
                                               AutoManual=self.ManualAuto, DebitCredit=self.debitcredit)

            ### JE Line - Refer
            sql_query = '''
                    SET NOCOUNT ON
                                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                GROUP BY CoA.GLAccountNumber
                                    SELECT				
                                        JournalEntries.BusinessUnit	AS 회사코드		
                                        , JournalEntries.JENumber AS 전표번호			
                                        , JournalEntries.JELineNumber AS 전표라인번호
                                        , JournalEntries.Year AS 회계연도	
                                        , JournalEntries.Period	AS 회계기간		
                                        , JournalEntries.EffectiveDate AS 전기일		
                                        , JournalEntries.EntryDate AS 입력일	
                                        , JournalEntries.Amount AS 금액	
                                        , JournalEntries.FunctionalCurrencyCode AS 통화	
                                        , JournalEntries.GLAccountNumber AS 계정코드	
                                        , #TMPCOA.GLAccountName AS 계정명
                                        , JournalEntries.Source AS 전표유형			
                                        , JournalEntries.PreparerID AS 입력자
                                        , JournalEntries.ApproverID AS 승인자
                                        , JournalEntries.JEDescription AS 전표헤더적요		
                                        , JournalEntries.JELineDescription AS 전표라인적요
                                        {NewSelect}			
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                    AND JournalEntries.JELINEID = Details.JENumberID 
                                    AND JournalEntries.GLAccountNumber IN 				
                                        (			
                                            SELECT DISTINCT JournalEntries.GLAccountNumber			
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                            WHERE JournalEntries.JELINEID = Details.JENumberID 
                                            GROUP BY JournalEntries.GLAccountNumber
                                            HAVING COUNT(JournalEntries.GLAccountNumber) <= {N}		
                                        ) 
                                    AND ABS(JournalEntries.Amount) >= {TE}
                                    {Account}
                                    {NewSQL}
                                    {AutoManual}
                                    {DebitCredit}
                                    ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                                    DROP TABLE #TMPCOA
                                '''.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                                           Account=self.checked_account4, NewSQL=self.NewSQL,
                                           NewSelect=self.NewSelect, AutoManual=self.ManualAuto,
                                           DebitCredit=self.debitcredit)

            self.dataframe_refer = pd.read_sql(sql_refer, self.cnxn)
            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### JE - Journals
        elif self.rbtn2.isChecked():
            sql_query = '''
                    SET NOCOUNT ON
                                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                GROUP BY CoA.GLAccountNumber
                                SELECT				
                                     JournalEntries.BusinessUnit AS 회사코드		
                                    , JournalEntries.JENumber AS 전표번호			
                                    , JournalEntries.JELineNumber AS 전표라인번호
                                    , JournalEntries.Year AS 회계연도	
                                    , JournalEntries.Period	AS 회계기간		
                                    , JournalEntries.EffectiveDate AS 전기일		
                                    , JournalEntries.EntryDate AS 입력일	
                                    , JournalEntries.Amount AS 금액	
                                    , JournalEntries.FunctionalCurrencyCode AS 통화	
                                    , JournalEntries.GLAccountNumber AS 계정코드	
                                    , #TMPCOA.GLAccountName AS 계정명
                                    , JournalEntries.Source AS 전표유형			
                                    , JournalEntries.PreparerID AS 입력자
                                    , JournalEntries.ApproverID AS 승인자
                                    , JournalEntries.JEDescription AS 전표헤더적요		
                                    , JournalEntries.JELineDescription AS 전표라인적요
                                    {NewSelect}			
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA
                                , [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                AND JournalEntries.JELINEID = Details.JENumberID 
                                AND Details.JEIdentifierID IN 
                                    (				
                                    SELECT DISTINCT Details.JEIdentifierID			
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,			
	                                    [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details				
                                    WHERE JournalEntries.JELINEID = Details.JENumberID
                                    AND JournalEntries.GLAccountNumber IN 			
                                            (	
                                            SELECT DISTINCT JournalEntries.GLAccountNumber			
                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                            WHERE JournalEntries.JELINEID = Details.JENumberID 
                                            GROUP BY JournalEntries.GLAccountNumber
                                            HAVING COUNT(JournalEntries.GLAccountNumber) <= {N}
                                            ) 
                                    AND ABS(JournalEntries.Amount) >= {TE}
                                    {Account}
                                    {NewSQL}
                                    {AutoManual}
                                    {DebitCredit}
                                    ) 	
                                ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                                DROP TABLE #TMPCOA
                        '''.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                                   Account=self.checked_account4, NewSQL=self.NewSQL,
                                   NewSelect=self.NewSelect, AutoManual=self.ManualAuto, DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Reference"] = [self.tempSheet + "_Reference", "Scenario01",
                                                                "---Filtered Result_1  Scenario01---\n" + sql_refer]

            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario01",
                                                             "---Filtered Result_2  Scenario01---\n" + sql_query]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario01",
                                                               "---Filtered JE  Scenario01---\n" + sql_query]

        ### 최대 추출 라인수
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate4.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[계정사용 빈도수: " + str(self.temp_N) + ","
                                                       + "중요성금액: " + str(self.temp_TE)
                                                       + '] 라인 수 ' + str(len(self.dataframe)) + '개입니다']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate4.closeApp.emit()

        else:
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                ### 시트 콤보박스에 저장
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate4.closeApp.emit()

    def extButtonClicked5(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()
        ### JE Line
        if self.rbtn1.isChecked():

            sql_query = """
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}            
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 											
                        AND ABS(JournalEntries.Amount) >= {TE} 				
                        {Account}
                        {DebitCredit}
                        {NewSQL}	
                        {AutoManual}		
                        ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA				
                                """.format(field=self.selected_project_id, TE=self.temp_TE,
                                           Account=self.checked_account5, DebitCredit=self.debitcredit,
                                           NewSQL=self.NewSQL,
                                           AutoManual=self.ManualAuto, NewSelect=self.NewSelect)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### JE
        elif self.rbtn2.isChecked():

            sql_query = """
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}		
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 							
                        AND Details.JEIdentifierID IN				
                                (		
                                 SELECT DISTINCT Details.JEIdentifierID		
                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                 WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                 AND ABS(JournalEntries.Amount) >= {TE}	
                                 {Account}
                                 {DebitCredit}
                                 {NewSQL}
                                 {AutoManual}	
                                )		
                        ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA				                                                                       
                                """.format(field=self.selected_project_id, TE=self.temp_TE,
                                           Account=self.checked_account5, DebitCredit=self.debitcredit,
                                           NewSQL=self.NewSQL,
                                           AutoManual=self.ManualAuto, NewSelect=self.NewSelect)
            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario02",
                                                             "---Filtered Result  Scenario02---\n" + sql_query]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals",
                                                               "Scenario02",
                                                               "---Filtered JE  Scenario02---\n" + sql_query]

        self.AccCode = re.sub("['|\s]", '', self.checked_account5)
        self.AccCode = self.AccCode[36:-1].split(',')

        ### 예외처리 5 - 최대 출력 라인 초과
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate5.closeApp.emit()

        ### 예외처리 6 - 데이터 미추출
        elif len(self.dataframe) == 0:

            if len(self.AccCode) == 1 and self.AccCode[0] == '':
                self.dataframe = pd.DataFrame({'No Data': ['[연도: ' + str(self.pname_year) + ','
                                                           + ' 계정코드: [당기생성계정 없음],'
                                                           + ' 라인수 ' + str(len(self.dataframe)) + '개 입니다.']})
            else:
                self.dataframe = pd.DataFrame({'No Data': ['[연도: ' + str(self.pname_year) + ','
                                                           + ' 계정코드: ' + str(self.AccCode) + ','
                                                           + ' 라인수 ' + str(len(self.dataframe)) + '개 입니다.']})

            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate5.closeApp.emit()

        else:
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate5.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 3번)
    def extButtonClicked6(self):
        cursor = self.cnxn.cursor()

        ### JE Line 추출
        if self.rbtn1.isChecked():
            sql = '''
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요						
                            {NewSelect}
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID			
                        AND JournalEntries.EntryDate >= {period1}				
                        AND JournalEntries.EntryDate <= {period2}				
                        AND ABS(JournalEntries.Amount) >= {TE}			
                        {Account}			
                        {NewSQL}				
                        {DebitCredit}				
                        {AutoManual}				
                        ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA				
                    '''.format(field=self.selected_project_id, Account=self.checked_account6, TE=self.temp_TE,
                               period1=str(self.tempDate1), period2=str(self.tempDate2),
                               NewSQL=self.NewSQL, DebitCredit=self.debitcredit, NewSelect=self.NewSelect,
                               AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE 추출
        elif self.rbtn2.isChecked():
            sql = '''
                       SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요
                            {NewSelect}			
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 							
                        AND Details.JEIdentifierID IN				
                                (		
                                 SELECT DISTINCT Details.JEIdentifierID		
                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                 WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                 AND JournalEntries.EntryDate >= {period1}	
                                 AND JournalEntries.EntryDate <= {period2}	
                                 AND ABS(JournalEntries.Amount) >= {TE}	
                                 {Account}	
                                 {NewSQL}		
                                 {DebitCredit}		
                                 {AutoManual}		
                                )		
                        ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA						
                    '''.format(field=self.selected_project_id, Account=self.checked_account6, TE=self.temp_TE,
                               period1=str(self.tempDate1), period2=str(self.tempDate2),
                               NewSQL=self.NewSQL, DebitCredit=self.debitcredit, NewSelect=self.NewSelect,
                               AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE Line 추출 시, 쿼리 저장
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario03",
                                                             "---Filtered Result  Scenario03---\n" + sql]
        ### JE 추출 시, 쿼리 저장
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario03",
                                                               "---Filtered JE  Scenario03---\n" + sql]

        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate6.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': ["- 시작 시점 : " + str(self.period1.text()) + " 종료 시점 : " + str(self.period2.text())
                             + "," + "중요성금액: " + str(self.temp_TE)
                             + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate6.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate6.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 4번)
    def extButtonClicked7(self):
        cursor = self.cnxn.cursor()

        ### JE Line 추출
        if self.rbtn1.isChecked():
            sql = '''
                       SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}		
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 							
                        {Date}				
                        AND ABS(JournalEntries.Amount) >= {TE}		
                        {Account}			
                        {NewSQL}				
                        {DebitCredit}				
                        {AutoManual}							
                        ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA				
                   '''.format(field=self.selected_project_id, TE=self.temp_TE, Date=self.tempState,
                              Account=self.checked_account7, NewSQL=self.NewSQL,
                              AutoManual=self.ManualAuto, NewSelect=self.NewSelect, DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE 추출
        elif self.rbtn2.isChecked():
            sql = '''
                    SET NOCOUNT ON				
                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                    GROUP BY CoA.GLAccountNumber				
                    SELECT				
                        JournalEntries.BusinessUnit AS 회사코드			
                        , JournalEntries.JENumber AS 전표번호			
                        , JournalEntries.JELineNumber AS 전표라인번호			
                        , JournalEntries.Year AS 회계연도			
                        , JournalEntries.Period AS 회계기간			
                        , JournalEntries.EffectiveDate AS 전기일			
                        , JournalEntries.EntryDate AS 입력일			
                        , JournalEntries.Amount AS 금액			
                        , JournalEntries.FunctionalCurrencyCode AS 통화			
                        , JournalEntries.GLAccountNumber AS 계정코드			
                        , #TMPCOA.GLAccountName AS 계정명			
                        , JournalEntries.Source AS 전표유형			
                        , JournalEntries.PreparerID AS 입력자			
                        , JournalEntries.ApproverID AS 승인자			
                        , JournalEntries.JEDescription AS 전표헤더적요			
                        , JournalEntries.JELineDescription AS 전표라인적요
                        {NewSelect}			
                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                        #TMPCOA,			
                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                    AND JournalEntries.JELINEID = Details.JENumberID 								
                    AND Details.JEIdentifierID IN				
                            (		
                             SELECT DISTINCT Details.JEIdentifierID		
                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                             {Date}	
                             AND ABS(JournalEntries.Amount) >= {TE}	
                             {Account}		
                             {NewSQL}		
                             {DebitCredit}		
                             {AutoManual}		
                            )		
                    ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                    DROP TABLE #TMPCOA				
                   '''.format(field=self.selected_project_id, TE=self.temp_TE, Date=self.tempState,
                              Account=self.checked_account7, NewSQL=self.NewSQL,
                              AutoManual=self.ManualAuto, NewSelect=self.NewSelect, DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE Line 추출 시, 쿼리 저장
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario04",
                                                             "---Filtered Result  Scenario04---\n" + sql]

        ### JE 추출 시, 쿼리 저장
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario04",
                                                               "---Filtered JE  Scenario04---\n" + sql]

        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate7.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[EffectiveDate/EntryDate: " + str(self.tempState) + ","
                                                       + "중요성금액: " + str(self.temp_TE)
                                                       + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate7.closeApp.emit()

        else:

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate7.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 5번)
    def extButtonClicked8(self):
        cursor = self.cnxn.cursor()

        ### JE Line 추출
        if self.rbtn1.isChecked():
            sql = '''
                            SET NOCOUNT ON				
                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                            GROUP BY CoA.GLAccountNumber
                            SELECT				
                                JournalEntries.BusinessUnit AS 회사코드			
                                , JournalEntries.JENumber AS 전표번호			
                                , JournalEntries.JELineNumber AS 전표라인번호			
                                , JournalEntries.Year AS 회계연도			
                                , JournalEntries.Period AS 회계기간			
                                , JournalEntries.EffectiveDate AS 전기일			
                                , JournalEntries.EntryDate AS 입력일			
                                , JournalEntries.Amount AS 금액			
                                , JournalEntries.FunctionalCurrencyCode AS 통화			
                                , JournalEntries.GLAccountNumber AS 계정코드			
                                , #TMPCOA.GLAccountName AS 계정명			
                                , JournalEntries.Source AS 전표유형			
                                , JournalEntries.PreparerID AS 입력자			
                                , JournalEntries.ApproverID AS 승인자			
                                , JournalEntries.JEDescription AS 전표헤더적요			
                                , JournalEntries.JELineDescription AS 전표라인적요			
                                {NewSelect}			
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                #TMPCOA,			
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                            AND JournalEntries.JELINEID = Details.JENumberID 						
                            AND ABS(DATEDIFF(dd, JournalEntries.EntryDate ,JournalEntries.EffectiveDate)) >= {N}			
                            AND ABS(JournalEntries.Amount) >= {TE}			
                            {Account}				
                            {NewSQL}				
                            {DebitCredit}
                            {AutoManual}				
                            ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                            DROP TABLE #TMPCOA				
                        '''.format(field=self.selected_project_id, N=self.realNDate, TE=self.temp_TE,
                                   Account=self.checked_account8, AutoManual=self.ManualAuto,
                                   NewSQL=self.NewSQL, NewSelect=self.NewSelect, DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        elif self.rbtn2.isChecked():

            sql = '''
                            SET NOCOUNT ON				
                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                            GROUP BY CoA.GLAccountNumber				
                            SELECT				
                                JournalEntries.BusinessUnit AS 회사코드			
                                , JournalEntries.JENumber AS 전표번호			
                                , JournalEntries.JELineNumber AS 전표라인번호			
                                , JournalEntries.Year AS 회계연도			
                                , JournalEntries.Period AS 회계기간			
                                , JournalEntries.EffectiveDate AS 전기일			
                                , JournalEntries.EntryDate AS 입력일			
                                , JournalEntries.Amount AS 금액			
                                , JournalEntries.FunctionalCurrencyCode AS 통화			
                                , JournalEntries.GLAccountNumber AS 계정코드			
                                , #TMPCOA.GLAccountName AS 계정명			
                                , JournalEntries.Source AS 전표유형			
                                , JournalEntries.PreparerID AS 입력자			
                                , JournalEntries.ApproverID AS 승인자			
                                , JournalEntries.JEDescription AS 전표헤더적요			
                                , JournalEntries.JELineDescription AS 전표라인적요			
                                {NewSelect}		
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                #TMPCOA,			
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                            AND JournalEntries.JELINEID = Details.JENumberID 						
                            AND Details.JEIdentifierID IN				
                                    (		
                                     SELECT DISTINCT Details.JEIdentifierID		
                                     FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                     WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                     AND ABS(DATEDIFF(dd, JournalEntries.EntryDate ,JournalEntries.EffectiveDate)) >= {N}
                                     AND ABS(JournalEntries.Amount) >= {TE}
                                     {Account}	
                                     {NewSQL}	
                                     {DebitCredit}
                                     {AutoManual}
                                    )	
                            ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                            DROP TABLE #TMPCOA				
                            '''.format(field=self.selected_project_id, N=self.realNDate, TE=self.temp_TE,
                                       Account=self.checked_account8, AutoManual=self.ManualAuto,
                                       NewSQL=self.NewSQL, NewSelect=self.NewSelect, DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE Line 추출 시, 쿼리 저장
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario05",
                                                             "---Filtered Result  Scenario05---\n" + sql]

        ### JE 추출 시, 쿼리 저장
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario05",
                                                               "---Filtered JE  Scenario05---\n" + sql]

        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate8.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': ["[Effective Date와 Entry Date 간 차이: " + str(int(self.realNDate))
                             + "," + "중요성금액: " + str(self.temp_TE)
                             + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate8.closeApp.emit()

        else:

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate8.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 6번)
    def extButtonClicked9(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        if self.rbtn1.isChecked():  # JE Line- Result
            sql = '''
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}		
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 							
                        AND JournalEntries.PreparerID IN				
                                (		
                                 SELECT DISTINCT JournalEntries.PreparerID		
                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details		
                                 WHERE JournalEntries.JELINEID = Details.JENumberID 
                                 GROUP BY JournalEntries.PreparerID		
                                 HAVING COUNT(JournalEntries.PreparerID) <= {N}
                                )		
                        AND ABS(JournalEntries.Amount) >= {TE}
                        {Account}
                        {DebitCredit}
                        {NewSQL}
                        {AutoManual}
                        ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA				
                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                   DebitCredit=self.debitcredit,
                                   Account=self.checked_account9, NewSQL=self.NewSQL,
                                   NewSelect=self.NewSelect, AutoManual=self.ManualAuto)

            # Reference
            sql_refer = '''
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT JournalEntries.PreparerID AS 전표입력자ID,					
                                MAX(Users.FullName) AS 입력자명, 			
                                MAX(Users.Title) AS 직급, 			
                                MAX(Users.Department) AS 부서,			
                                COUNT(JournalEntries.PreparerID) AS CNT
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details,			
                            [{field}_Import_Dim].[dbo].[pbcUser] AS Users			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 				
                        AND Users.UserName = JournalEntries.PreparerID				
                        AND JournalEntries.PreparerID IN				
                                (		
                                 SELECT DISTINCT JournalEntries.PreparerID		
                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details		
                                 WHERE JournalEntries.JELINEID = Details.JENumberID 
                                 GROUP BY JournalEntries.PreparerID		
                                 HAVING COUNT(JournalEntries.PreparerID) <= {N}
                                )		
                        AND ABS(JournalEntries.Amount) >= {TE}
                        {Account}
                        {DebitCredit}
                        {NewSQL}
                        {AutoManual}
                        GROUP BY JournalEntries.PreparerID			
                        ORDER BY JournalEntries.PreparerID				
                        DROP TABLE #TMPCOA				
                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                   DebitCredit=self.debitcredit,
                                   Account=self.checked_account9, NewSQL=self.NewSQL,
                                   AutoManual=self.ManualAuto)

            self.dataframe_refer = pd.read_sql(sql_refer, self.cnxn)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        elif self.rbtn2.isChecked():  # JE- Journals
            sql = '''
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber				
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 				
                        AND Details.JEIdentifierID IN				
                                (		
                                 SELECT DISTINCT Details.JEIdentifierID		
                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                 WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                 AND JournalEntries.PreparerID IN		
                                        (
                                         SELECT DISTINCT JournalEntries.PreparerID		
                                         FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details		
                                         WHERE JournalEntries.JELINEID = Details.JENumberID 
                                         GROUP BY JournalEntries.PreparerID		
                                         HAVING COUNT(JournalEntries.PreparerID) <= {N}
                                        )
                                AND ABS(JournalEntries.Amount) >= {TE} 
                                {Account}
                                {DebitCredit}
                                {NewSQL}
                                {AutoManual}		
                                )		
                        ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA				
                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                   DebitCredit=self.debitcredit,
                                   Account=self.checked_account9, NewSQL=self.NewSQL,
                                   NewSelect=self.NewSelect, AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### Reference 추출 시, 쿼리 저장
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Reference"] = [self.tempSheet + "_Reference", "Scenario06",
                                                                "---Filtered Result_1  Scenario06---\n" + sql_refer]
            ### JE Line- Result 추출 시, 쿼리 저장
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario06",
                                                             "---Filtered Result_2  Scenario06---\n" + sql]
        ### JE- Journals 추출 시, 쿼리 저장
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario06",
                                                               "---Filtered JE  Scenario06---\n" + sql]
        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate9.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.communicate9.closeApp.emit()

        ### 0건이 아닐 경우, 시트&데이터 추출
        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate9.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 7번)
    def extButtonClicked10(self):
        cursor = self.cnxn.cursor()

        ### JE Line 추출
        if self.rbtn1.isChecked():

            sql = '''
                                 SET NOCOUNT ON
                                 SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                 FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                 GROUP BY CoA.GLAccountNumber
                                 SELECT			
                                        JournalEntries.BusinessUnit AS 회사코드			
                                        , JournalEntries.JENumber AS 전표번호			
                                        , JournalEntries.JELineNumber AS 전표라인번호			
                                        , JournalEntries.Year AS 회계연도			
                                        , JournalEntries.Period AS 회계기간			
                                        , JournalEntries.EffectiveDate AS 전기일			
                                        , JournalEntries.EntryDate AS 입력일			
                                        , JournalEntries.Amount AS 금액			
                                        , JournalEntries.FunctionalCurrencyCode AS 통화			
                                        , JournalEntries.GLAccountNumber AS 계정코드			
                                        , #TMPCOA.GLAccountName AS 계정명			
                                        , JournalEntries.Source AS 전표유형			
                                        , JournalEntries.PreparerID AS 입력자			
                                        , JournalEntries.ApproverID AS 승인자			
                                        , JournalEntries.JEDescription AS 전표헤더적요			
                                        , JournalEntries.JELineDescription AS 전표라인적요			
                                       {NewSelect}			
                               FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                               [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                               WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                               AND JournalEntries.JELINEID = Details.JENumberID 				        	
                               AND ABS(JournalEntries.Amount) >= {TE} 
                               {Account}
                               {NewSQL}
                               {AutoManual}
                               {DebitCredit}
                               ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                               DROP TABLE #TMPCOA			
                            '''.format(field=self.selected_project_id, TE=self.tempTE,
                                       Account=self.checked_account10,
                                       NewSQL=self.NewSQL, AutoManual=self.ManualAuto, NewSelect=self.NewSelect,
                                       DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE 추출
        elif self.rbtn2.isChecked():

            sql = '''
                                   SET NOCOUNT ON
                                   SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA
                                   FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                   GROUP BY CoA.GLAccountNumber
                                   SELECT 			
                                                JournalEntries.BusinessUnit AS 회사코드			
                                                , JournalEntries.JENumber AS 전표번호			
                                                , JournalEntries.JELineNumber AS 전표라인번호			
                                                , JournalEntries.Year AS 회계연도			
                                                , JournalEntries.Period AS 회계기간			
                                                , JournalEntries.EffectiveDate AS 전기일			
                                                , JournalEntries.EntryDate AS 입력일			
                                                , JournalEntries.Amount AS 금액			
                                                , JournalEntries.FunctionalCurrencyCode AS 통화			
                                                , JournalEntries.GLAccountNumber AS 계정코드			
                                                , #TMPCOA.GLAccountName AS 계정명			
                                                , JournalEntries.Source AS 전표유형			
                                                , JournalEntries.PreparerID AS 입력자			
                                                , JournalEntries.ApproverID AS 승인자			
                                                , JournalEntries.JEDescription AS 전표헤더적요			
                                                , JournalEntries.JELineDescription AS 전표라인적요			
                                                {NewSelect}	
                                   FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries, #TMPCOA,
                                    [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details
                                   WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 
                                   AND JournalEntries.JELINEID = Details.JENumberID 
                                   AND Details.JEIdentifierID IN 		
                                                    (	
                                                    SELECT DISTINCT Details.JEIdentifierID	
                                                    FROM  [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,
                                                    [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                                    WHERE JournalEntries.JELINEID = Details.JENumberID 
                                                    AND ABS(JournalEntries.Amount) >= {TE}	
                                                    {Account}
                                                    {NewSQL}
                                                    {AutoManual}
                                                    {DebitCredit}
                                                    ) 
                                   ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                                   DROP TABLE #TMPCOA			
                            '''.format(field=self.selected_project_id, TE=self.tempTE,
                                       Account=self.checked_account10,
                                       NewSQL=self.NewSQL, AutoManual=self.ManualAuto, NewSelect=self.NewSelect,
                                       DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario07",
                                                             "---Filtered Result  Scenario07---\n" + sql]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario07",
                                                               "---Filtered JE  Scenario07---\n" + sql]

        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate10.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.communicate10.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate10.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 8-1번)
    def extButtonClicked12(self):
        cursor = self.cnxn.cursor()

        ### 기능영역 선택 시 Reference 추출
        if self.checkF.isChecked():
            sql = '''
                        SET NOCOUNT ON;																	
                        SELECT 																	
                            Details.JEIdentifierID AS JENumber,															
                            JournalEntries.GLAccountNumber, 																
                            JournalEntries.Debit, 																
                            JournalEntries.Credit, 																
                            JournalEntries.Amount,																
                            JournalEntries.Segment01																
                            INTO #tmp																
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,																	
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details																
                        WHERE JournalEntries.JELINEID = Details.JENumberID 																																	
                        AND ABS(JournalEntries.Amount) >= {TE}																	
                        {AutoManual}																	
                                SELECT															
                                        LVL4.GL_Functional_Area AS 기능영역,													
                                        LVL4.GL_Account_Number AS 계정코드,													
                                        LVL4.GL_Account_Name AS 계정명,													
                                        LVL4.Account_Type AS 계정대분류,													
                                        LVL4.Account_Class AS 계정중분류,													
                                        LVL4.GL_Account_Position AS '차/대',													
                                        LVL4.Posting_Type AS 상대계정유형,													
                                        LVL4.Analysis_GL_Functional_Area AS 상대계정기능영역,													
                                        LVL4.Analysis_GL_Account_Number AS 상대계정코드,													
                                        LVL4.Analysis_GL_Account_Name AS 상대계정명,													
                                        LVL4.Analysis_Account_Type AS 상대계정대분류,													
                                        LVL4.Analysis_Account_Class AS 상대계정중분류,													
                                        LVL4.Analysis_Position AS '상대계정차/대',													
                                        LVL4.Sum_Of_Debit_Amount AS 차변합계금액,													
                                        LVL4.Sum_Of_Credit_Amount AS 대변합계금액,													
                                        LVL4.JE_Line_Count AS 전표라인수													
                                FROM                                                                                                    															
                                (                                                                                                       															
                                       SELECT  LVL3.FunctionalArea1 AS GL_Functional_Area,                                                                                                                                          														
                                               LVL3.GLAccountNumber1 AS GL_Account_Number,                                                                                                                                												
                                               MAX(LVL3.GLAccountName1) AS GL_Account_Name,                                                                                    												
                                               MAX(LVL3.AccountType1) AS Account_Type,												
                                               LVL3.AccountClass1 AS Account_Class,												
                                               LVL3.DivideDC1 AS GL_Account_Position,                                                                                 												
                                               CASE                                                                                                                      												
                                               WHEN LVL3.GLAccountNumber1 = LVL3.GLAccountNumber2 and  LVL3.DivideDC1  = LVL3.DivideDC2 THEN '1.Analysis Account'                                                                                                                            												
                                               WHEN LVL3.GLAccountNumber1 <> LVL3.GLAccountNumber2 and LVL3.DivideDC1 = LVL3.DivideDC2 THEN '3.Reference Account'                                                                                                                           												
                                               ELSE '2.Correspondent Account'                                                                                                                   												
                                               END AS Posting_Type, 												
                                               LVL3.FunctionalArea2 AS Analysis_GL_Functional_Area,												
                                               LVL3.GLAccountNumber2 AS Analysis_GL_Account_Number,                                                                                                                        												
                                               MAX(LVL3.GLAccountName2) AS Analysis_GL_ACcount_Name,                                                                                  												
                                               MAX(LVL3.AccountType2) AS Analysis_Account_Type,												
                                               LVL3.AccountClass2 AS Analysis_Account_Class,												
                                               LVL3.DivideDC2 AS Analysis_Position,                                                                                                            												
                                               SUM(LVL3.SumOfDebit2) AS Sum_Of_Debit_Amount,                                                                                                                                 												
                                               SUM(LVL3.SumOfCredit2) AS Sum_Of_Credit_Amount,                                                                                                                               												
                                               SUM(LVL3.Cnt2) AS JE_Line_Count												
                                       FROM                                                                                             														
                                       (                                                                                                														
                                               SELECT *                                                                                         												
                                               FROM                                                                                     												
                                                      (                                                                                										
                                                                     SELECT                                                             						
                                                                            LVL1_1.JENumber1,                                                         				
                                                                            LVL1_1.GLAccountNumber1,                                                          				
                                                                            MAX(LVL1_1.CoA_GLAccountName1) AS GLAccountName1,                                                            				
                                                                            MAX(LVL1_1.AccountType1) AS AccountType1,				
                                                                            LVL1_1.AccountClass1 AS AccountClass1,				
                                                                            SUM(LVL1_1.Debit1) AS SumOfDebit1,                                                       				
                                                                            SUM(LVL1_1.Credit1) AS SumOfCredit1,                                                      				
                                                                            DivideDC1,                                                         				
                                                                            COUNT(*) AS Cnt1,				
                                                                            LVL1_1.FunctionalArea1 AS FunctionalArea1				
                                                                     FROM                                                               						
                                                                     (                                                                  						
                                                                                    SELECT                                               		
                                                                                           #tmp.JENumber AS JENumber1,                                          	
                                                                                           #tmp.GLAccountNumber AS GLAccountNumber1,                                          	
                                                                                           CoA.GLAccountNumber AS CoA_GLAccountNumber1,                                       	
                                                                                           CoA.GLAccountName AS CoA_GLAccountName1,                                      	
                                                                                           CoA.AccountType AS AccountType1,	
                                                                                           CoA.AccountClass AS AccountClass1,	
                                                                                           #tmp.Debit AS Debit1,                                             	
                                                                                           #tmp.Credit AS Credit1,                                            	
                                                                                           #tmp.Amount AS Amount1,	
                                                                                           #tmp.Segment01 AS FunctionalArea1,	
                                                                                           CASE                                         	
                                                                                           WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       	
                                                                                           END AS 'DivideDC1'                                            	
                                                                                    FROM #tmp, [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] CoA                                                		
                                                                                    WHERE CONCAT(#tmp.GLAccountNumber, #tmp.Segment01) = CONCAT(CoA.GLAccountNumber, CoA.Segment01)                                               		
                                                                     ) LVL1_1                                                                  						
                                                                     GROUP BY LVL1_1.JENumber1, LVL1_1.GLAccountNumber1, LVL1_1.DivideDC1, LVL1_1.FunctionalArea1, LVL1_1.AccountClass1                                                               						
                                                      ) LVL2_1,                                                                                										
                                                      (                                                                                 										
                                                                     SELECT                                                             						
                                                                            LVL1_2.JENumber2,                                                         				
                                                                            LVL1_2.GLAccountNumber2,                                                          				
                                                                            MAX(LVL1_2.CoA_GLAccountName2) AS GLAccountName2,                                                            				
                                                                            MAX(LVL1_2.AccountType2) AS AccountType2, 				
                                                                            LVL1_2.AccountClass2 AS AccountClass2,				
                                                                            SUM(LVL1_2.Debit2) AS SumOfDebit2,                                                       				
                                                                            SUM(LVL1_2.Credit2) AS SumOfCredit2,                                                      				
                                                                            DivideDC2,                                                         				
                                                                            COUNT(*) AS Cnt2,				
                                                                            LVL1_2.FunctionalArea2 AS FunctionalArea2				
                                                                     FROM                                                               						
                                                                     (                                                                  						
                                                                                    SELECT #tmp.JENumber AS JENumber2,                                                  		
                                                                                           #tmp.GLAccountNumber AS GLAccountNumber2,                                          	
                                                                                           CoA.GLAccountNumber AS CoA_GLAccountNumber2,                                       	
                                                                                           CoA.GLAccountName AS CoA_GLAccountName2,                                      	
                                                                                           CoA.AccountType AS AccountType2, 	
                                                                                           CoA.AccountClass AS AccountClass2,	
                                                                                           #tmp.Debit AS Debit2,                                             	
                                                                                           #tmp.Credit AS Credit2,                                            	
                                                                                           #tmp.Amount AS Amount2,	
                                                                                           #tmp.Segment01 AS FunctionalArea2,	
                                                                                           CASE                                         	
                                                                                           WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       	
                                                                                           END AS 'DivideDC2'                                            	
                                                                                    FROM #tmp, [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] CoA                                                		
                                                                                    WHERE CONCAT(#tmp.GLAccountNumber, #tmp.Segment01) = CONCAT(CoA.GLAccountNumber, CoA.Segment01)                                            		
                                                                     ) LVL1_2                                                                  						
                                                                     GROUP BY LVL1_2.JENumber2, LVL1_2.GLAccountNumber2, LVL1_2.DivideDC2, LVL1_2.FunctionalArea2, LVL1_2.AccountClass2                                                              						
                                                      ) LVL2_2                                                                                 										
                                               WHERE LVL2_1.JENumber1 = LVL2_2.JENumber2                                                                                      												
                                       ) LVL3                                                                                                  														
                                       GROUP BY LVL3.GLAccountNumber1, LVL3.DivideDC1, LVL3.GLAccountNumber2, LVL3.DivideDC2, LVL3.FunctionalArea1, LVL3.FunctionalArea2, LVL3.AccountClass1, LVL3.AccountClass2                                                                                          														
                                ) LVL4                                                                                                                                                                                                  															
                                where LVL4.Posting_Type = '2.Correspondent Account'															
                                {AccountA}															
                                {DebitCreditA}															
                                {AccountB}															
                                {DebitCreditB}															
                                ORDER BY LVL4.GL_Account_Number, LVL4.GL_Account_Position, LVL4.Posting_Type, LVL4.Analysis_GL_Account_Number    															
                        DROP TABLE #TMP																	
                                        '''.format(field=self.selected_project_id, DebitCreditA=self.tempStateA,
                                                   AccountA=self.checked_accountA, AccountB=self.checked_accountB,
                                                   DebitCreditB=self.tempStateB, TE=self.temp_TE,
                                                   AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

            ### 비경상적계정 선택여부 컬럽 추가
            self.dataframe['비경상적계정 선택여부'] = ''

            ### 기능영역 미선택 후, Reference 추출 시, 쿼리 저장
            self.my_query.loc[self.temp_Sheet + "_Reference"] = [self.temp_Sheet + "_Reference", "Scenario08",
                                                                 "---Filtered Result  Scenario08---\n" + sql]

        ### 기능영역 미선택 시 Reference 추출
        else:
            sql = '''
                        SET NOCOUNT ON;																	
                        SELECT COA.GLAccountNumber,																	
                               MAX(COA.GLAccountName) AS GLAccountName, 																
                               MAX(COA.AccountType) AS AccountType,																
                               MAX(COA.AccountClass) AS AccountClass																
                               INTO #TMPCOA																
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA																	
                        GROUP BY COA.GLAccountNumber																	
                        SELECT 																	
                            Details.JEIdentifierID AS JENumber,															
                            JournalEntries.GLAccountNumber, 																
                            JournalEntries.Debit, 																
                            JournalEntries.Credit, 																
                            JournalEntries.Amount 																
                            INTO #tmp																
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,																	
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details																
                        WHERE JournalEntries.JELINEID = Details.JENumberID 															
                        AND ABS(JournalEntries.Amount) >= {TE}																	
                        {AutoManual}																
                                SELECT 															
                                        LVL4.GL_Account_Number AS 계정코드,													
                                        LVL4.GL_Account_Name AS 계정명,													
                                        LVL4.Account_Type AS 계정대분류,													
                                        LVL4.Account_Class AS 계정중분류,													
                                        LVL4.GL_Account_Position AS '차/대',													
                                        LVL4.Posting_Type AS 상대계정유형,													
                                        LVL4.Analysis_GL_Account_Number AS 상대계정코드,													
                                        LVL4.Analysis_GL_Account_Name AS 상대계정명,													
                                        LVL4.Analysis_Account_Type AS 상대계정대분류,													
                                        LVL4.Analysis_Account_Class AS 상대계정중분류,													
                                        LVL4.Analysis_Position AS '상대계정차/대',													
                                        LVL4.Sum_Of_Debit_Amount AS 차변합계금액,													
                                        LVL4.Sum_Of_Credit_Amount AS 대변합계금액,													
                                        LVL4.JE_Line_Count AS 전표라인수													
                                FROM                                                                                                    															
                                (                                                                                                       															
                                       SELECT                                                                                                                                            														
                                               LVL3.GLAccountNumber1 AS GL_Account_Number,                                                                                                                                												
                                               MAX(LVL3.GLAccountName1) AS GL_ACcount_Name,                                                                                    												
                                               MAX(LVL3.AccountType1) AS Account_Type, 												
                                               LVL3.AccountClass1 AS Account_Class,												
                                               LVL3.DivideDC1 AS GL_Account_Position,                                                                                 												
                                               CASE                                                                                                                      												
                                               WHEN LVL3.GLAccountNumber1 = LVL3.GLAccountNumber2 and  LVL3.DivideDC1  = LVL3.DivideDC2 THEN '1.Analysis Account'                                                                                                                            												
                                               WHEN LVL3.GLAccountNumber1 <> LVL3.GLAccountNumber2 and LVL3.DivideDC1 = LVL3.DivideDC2 THEN '3.Reference Account'                                                                                                                           												
                                               ELSE '2.Correspondent Account'                                                                                                                   												
                                               END AS Posting_Type,                                                                                                                      												
                                               LVL3.GLAccountNumber2 AS Analysis_GL_Account_Number,                                                                                                                        												
                                               MAX(LVL3.GLAccountName2) AS Analysis_GL_ACcount_Name,                                                                                  												
                                               MAX(LVL3.AccountType2) AS Analysis_Account_Type, 												
                                               LVL3.AccountClass2 AS Analysis_Account_Class,												
                                               LVL3.DivideDC2 AS Analysis_Position,                                                                                                            												
                                               SUM(LVL3.SumOfDebit2) AS Sum_Of_Debit_Amount,                                                                                                                                 												
                                               SUM(LVL3.SumOfCredit2) AS Sum_Of_Credit_Amount,                                                                                                                               												
                                               SUM(LVL3.Cnt2) AS JE_Line_Count                                                                                                                                     												
                                       FROM                                                                                             														
                                       (                                                                                                														
                                               SELECT *                                                                                         												
                                               FROM                                                                                     												
                                                      (                                                                                										
                                                                     SELECT                                                             						
                                                                            LVL1_1.JENumber1,                                                         				
                                                                            LVL1_1.GLAccountNumber1,                                                          				
                                                                            MAX(LVL1_1.CoA_GLAccountName1) AS GLAccountName1,                                                            				
                                                                            MAX(LVL1_1.AccountType1) AS AccountType1, 				
                                                                            LVL1_1.AccountClass1 AS AccountClass1,				
                                                                            SUM(LVL1_1.Debit1) AS SumOfDebit1,                                                       				
                                                                            SUM(LVL1_1.Credit1) AS SumOfCredit1,                                                      				
                                                                            DivideDC1,                                                         				
                                                                            COUNT(*) AS Cnt1                                                          				
                                                                     FROM                                                               						
                                                                     (                                                                  						
                                                                                    SELECT                                               		
                                                                                           #tmp.JENumber AS JENumber1,                                          	
                                                                                           #tmp.GLAccountNumber AS GLAccountNumber1,                                          	
                                                                                           #TMPCOA.GLAccountNumber AS CoA_GLAccountNumber1,                                       	
                                                                                           #TMPCOA.GLAccountName AS CoA_GLAccountName1,                                      	
                                                                                           #TMPCOA.AccountType AS AccountType1,  	
                                                                                           #TMPCOA.AccountClass AS AccountClass1,	
                                                                                           #tmp.Debit AS Debit1,                                             	
                                                                                           #tmp.Credit AS Credit1,                                            	
                                                                                           #tmp.Amount AS Amount1,                                            	
                                                                                           CASE                                         	
                                                                                           WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       	
                                                                                           END AS 'DivideDC1'                                            	
                                                                                    FROM #tmp, #TMPCOA                                                		
                                                                                    WHERE #tmp.GLAccountNumber = #TMPCOA.GLAccountNumber                                                		
                                                                     ) LVL1_1                                                                  						
                                                                     GROUP BY LVL1_1.JENumber1, LVL1_1.GLAccountNumber1, LVL1_1.DivideDC1, LVL1_1.AccountClass1                                                                						
                                                      ) LVL2_1,                                                                                										
                                                      (                                                                                 										
                                                                     SELECT                                                             						
                                                                            LVL1_2.JENumber2,                                                         				
                                                                            LVL1_2.GLAccountNumber2,                                                          				
                                                                            MAX(LVL1_2.CoA_GLAccountName2) AS GLAccountName2,                                                            				
                                                                            MAX(LVL1_2.AccountType2) AS AccountType2,				
                                                                            LVL1_2.AccountClass2 AS AccountClass2,				
                                                                            SUM(LVL1_2.Debit2) AS SumOfDebit2,                                                       				
                                                                            SUM(LVL1_2.Credit2) AS SumOfCredit2,                                                      				
                                                                            DivideDC2,                                                         				
                                                                            COUNT(*) AS Cnt2                                                          				
                                                                     FROM                                                               						
                                                                     (                                                                  						
                                                                                    SELECT #tmp.JENumber AS JENumber2,                                                  		
                                                                                           #tmp.GLAccountNumber AS GLAccountNumber2,                                          	
                                                                                           #TMPCOA.GLAccountNumber AS CoA_GLAccountNumber2,                                       	
                                                                                           #TMPCOA.GLAccountName AS CoA_GLAccountName2,                                      	
                                                                                           #TMPCOA.AccountType AS AccountType2, 	
                                                                                           #TMPCOA.AccountClass AS AccountClass2,	
                                                                                           #tmp.Debit AS Debit2,                                             	
                                                                                           #tmp.Credit AS Credit2,                                            	
                                                                                           #tmp.Amount AS Amount2,                                            	
                                                                                           CASE                                         	
                                                                                           WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       	
                                                                                           END AS 'DivideDC2'                                            	
                                                                                    FROM #tmp, #TMPCOA                                                		
                                                                                    WHERE #tmp.GLAccountNumber = #TMPCOA.GLAccountNumber                                                		
                                                                     ) LVL1_2                                                                  						
                                                                     GROUP BY LVL1_2.JENumber2, LVL1_2.GLAccountNumber2, LVL1_2.DivideDC2, LVL1_2.AccountClass2                                                               						
                                                      ) LVL2_2                                                                                 										
                                               WHERE LVL2_1.JENumber1 = LVL2_2.JENumber2                                                                                      												
                                       ) LVL3                                                                                                  														
                                       GROUP BY LVL3.GLAccountNumber1, LVL3.DivideDC1, LVL3.GLAccountNumber2, LVL3.DivideDC2, LVL3.AccountClass1, LVL3.AccountClass2                                                                                          														
                                ) LVL4                                                                                                                                                                                                  															
                                where LVL4.Posting_Type = '2.Correspondent Account'															
                                {AccountA}															
                                {DebitCreditA}															
                                {AccountB}															
                                {DebitCreditB}															
                                ORDER BY LVL4.GL_Account_Number, LVL4.GL_Account_Position, LVL4.Posting_Type, LVL4.Analysis_GL_Account_Number 															
                        DROP TABLE #TMP, #TMPCOA																	
            '''.format(field=self.selected_project_id, DebitCreditA=self.tempStateA,
                       AccountA=self.checked_accountA, AccountB=self.checked_accountB,
                       DebitCreditB=self.tempStateB, TE=self.temp_TE, AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)
            self.dataframe['비경상적계정 선택여부'] = ''

            ### 기능영역 선택 후, Reference 추출 시, 쿼리 저장
            self.my_query.loc[self.temp_Sheet + "_Reference"] = [self.temp_Sheet + "_Reference", "Scenario08",
                                                                 "---Filtered Result  Scenario08---\n" + sql]

        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            self.scenario_dic[self.temp_Sheet + '_Reference'] = self.dataframe.head(1000)
            self.combo_sheet.addItem(self.temp_Sheet + '_Reference')
            self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)
            self.communicate12.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[중요성금액: " + str(
                self.temp_TE) + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            self.scenario_dic[self.temp_Sheet + '_Reference'] = self.dataframe
            self.combo_sheet.addItem(self.temp_Sheet + '_Reference')
            self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate12.closeApp.emit()

        else:
            self.scenario_dic[self.temp_Sheet + '_Reference'] = self.dataframe
            self.combo_sheet.addItem(self.temp_Sheet + '_Reference')
            self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            self.communicate12.closeApp.emit()

    def CursorChange(self, row):
        """Cursor 입력값들에 적절하게 따옴표를 넣어주는 함수"""
        if row == 'NULL':
            return row
        else:
            return "'{}'".format(row)

    ### 쿼리문 관련 함수 (시나리오 8-2번)
    def extButtonClickedC(self):
        dflist = []
        cursorindex = []

        ### 기능영역 선택시 (커서문 형식)
        if self.checkF2.isChecked():
            index = self.wbC[self.wbC.iloc[:, 16].notnull()].iloc[:, [0, 1, 5, 7, 8, 12]]
            index.iloc[:, 0] = index.iloc[:, 0].fillna('NULL')
            index.iloc[:, 3] = index.iloc[:, 3].fillna('NULL')
            for i in range(len(index.columns)):
                index.iloc[:, i] = index.iloc[:, i].apply(lambda row: self.CursorChange(row))

            for i in range(len(index)):
                cursorindex.append("(" + str(index.iloc[i, 0]) + ","
                                   + str(index.iloc[i, 1]) + ","
                                   + str(index.iloc[i, 2]) + ","
                                   + str(index.iloc[i, 3]) + ","
                                   + str(index.iloc[i, 4]) + ","
                                   + str(index.iloc[i, 5]) + ")")
            cursortext = ',\n'.join(cursorindex)

        ### 기능영역 선택하지 않을 시 (커서문 형식)
        else:
            index = self.wbC[self.wbC.iloc[:, 14].notnull()].iloc[:, [0, 4, 6, 10]]
            for i in range(len(index.columns)):
                index.iloc[:, i] = index.iloc[:, i].apply(lambda row: self.CursorChange(row))

            for i in range(len(index)):
                cursorindex.append("(" + str(index.iloc[i, 0]) + ','
                                   + str(index.iloc[i, 1]) + ','
                                   + str(index.iloc[i, 2]) + ','
                                   + str(index.iloc[i, 3]) + ')')
            cursortext = ',\n'.join(cursorindex)

        if not self.checkF2.isChecked():  ### 기능영영 무
            if self.rbtn1.isChecked():  ### JE Line
                sql = '''
                    SET NOCOUNT ON		
                    --****************************************************Filter Table***************************************************																
                    CREATE TABLE #filter																
                    (GLAccountNumber VARCHAR(100), Debit_Credit VARCHAR(100), AL_GLAccountNumber VARCHAR(100), AL_Debit_Credit VARCHAR(100))																
                    INSERT INTO #filter																
                    VALUES																
                    {cursor}																
                    --****************************************************Insert ProjectID***************************************************																
                    SELECT 																
                        JournalEntries.BusinessUnit,															
                        JournalEntries.JENumber AS TMPJENumber, 															
                        Details.JEIdentifierID AS JENumber, 															
                        JournalEntries.JELineNumber,															
                        JournalEntries.Year,															
                        JournalEntries.Period,															
                        JournalEntries.EffectiveDate,															
                        JournalEntries.EntryDate,															
                        JournalEntries.Debit,															
                        JournalEntries.Credit,															
                        JournalEntries.Amount,															
                        JournalEntries.FunctionalCurrencyCode,															
                        JournalEntries.GLAccountNumber,															
                        JournalEntries.Source,															
                        JournalEntries.PreparerID,															
                        JournalEntries.ApproverID,															
                        JournalEntries.JEDescription,															
                        JournalEntries.JELineDescription															
                        INTO #JEData															
                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,																
                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details															
                    WHERE JournalEntries.JELINEID = Details.JENumberID 																															
                    AND ABS(JournalEntries.Amount) >= {TE}																
                    {AutoManual}															
                    SELECT  JournalEntries.BusinessUnit,																
                            JournalEntries.JENumber,														
                            Details.JEIdentifierID,														
                            JournalEntries.JELineNumber,																	
                            JournalEntries.EffectiveDate,																	
                            JournalEntries.EntryDate,																	
                            JournalEntries.Year,														
                            JournalEntries.Period,																	
                            JournalEntries.GLAccountNumber,																	
                            JournalEntries.Source,														
                            JournalEntries.Debit,																	
                            JournalEntries.Credit,																	
                            JournalEntries.Amount,																	
                            JournalEntries.FunctionalCurrencyCode,																	
                            JournalEntries.JEDescription,																	
                            JournalEntries.JELineDescription,																	
                            JournalEntries.PreparerID,																	
                            JournalEntries.ApproverID  INTO #JEData2																	
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,																	
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details														
                        WHERE JournalEntries.JELINEID = Details.JENumberID																	
                        AND Details.JEIdentifierID IN (															
                                                        SELECT DISTINCT Details.JEIdentifierID							
                                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,							
                                                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details							
                                                        WHERE JournalEntries.JELINEID = Details.JENumberID													
                                                        AND ABS(JournalEntries.Amount) >= {TE}							
                                                        {AutoManual}							
                                                      ) 													                                      
                    SELECT COA.GLAccountNumber,																
                           MAX(COA.GLAccountName) AS GLAccountName, 															
                           MAX(COA.AccountType) AS AccountType															
                           INTO #COAData															
                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA																
                    GROUP BY COA.GLAccountNumber																                          
                    --****************************************************Result Table***************************************************																
                    CREATE TABLE #result																
                    (																
                    BusinessUnit NVARCHAR(100),																
                    TMPJENumber NVARCHAR(100), 																
                    JENumber NVARCHAR(100),	 															
                    JELineNumber BIGINT,																
                    YEAR NVARCHAR(25),																
                    Period NVARCHAR(25),																
                    EffectiveDate DATE,																
                    EntryDate DATE,																
                    Debit NUMERIC(21,6),																
                    Credit NUMERIC(21,6),																
                    Amount NUMERIC(21,6),																
                    FunctionalCurrencyCode NVARCHAR(50),																
                    GLAccountNumber NVARCHAR(100),																
                    Source NVARCHAR(100),																
                    PreparerID NVARCHAR(100),																
                    ApproverID NVARCHAR(100),																
                    JEDescription NVARCHAR(200),																
                    JELineDescription NVARCHAR(200)																
                    )																
                    --****************************************************Cursor Start***************************************************																
                    DECLARE cur CURSOR FOR 																
                    SELECT GLAccountNumber, Debit_Credit, AL_GLAccountNumber, AL_Debit_Credit FROM #filter																
                    DECLARE @GLAccountNumber VARCHAR(100)																
                    DECLARE @Debit_Credit VARCHAR(100)																
                    DECLARE @AL_GLAccountNumber VARCHAR(100)																
                    DECLARE @AL_Debit_Credit VARCHAR(100)																
                    OPEN cur																
                    Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit																
                    WHILE(@@FETCH_STATUS <> -1)																
                    BEGIN;																
                    IF (@Debit_Credit = 'Debit')																
                        IF (@AL_Debit_Credit='Debit') /* Debit/Debit */															
                            INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                            GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                            SELECT JE1.BusinessUnit, JE1.TMPJENumber, JE1.JENumber, JE1.JELineNumber, JE1.Year, JE1.Period, JE1.EffectiveDate, JE1.EntryDate, JE1.Debit, JE1.Credit, JE1.Amount, JE1.FunctionalCurrencyCode, 														
                            JE1.GLAccountNumber, JE1.Source, JE1.PreparerID, JE1.ApproverID, JE1.JEDescription, JE1.JELineDescription FROM #JEData JE1														
                            WHERE JE1.JENumber IN (														
                                SELECT DISTINCT(JE1_1.JENumber)													
                                FROM #JEData JE1_1													
                                WHERE JE1_1.GLAccountNumber = @GLAccountNumber AND JE1_1.Debit<>0													
                                ) AND JE1.GLAccountNumber = @AL_GLAccountNumber AND JE1.Debit<>0													
                        ELSE /* Debit/Credit */															
                            INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                            GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                            SELECT JE2.BusinessUnit, JE2.TMPJENumber, JE2.JENumber, JE2.JELineNumber, JE2.Year, JE2.Period, JE2.EffectiveDate, JE2.EntryDate, JE2.Debit, JE2.Credit, JE2.Amount, JE2.FunctionalCurrencyCode, 														
                            JE2.GLAccountNumber, JE2.Source, JE2.PreparerID, JE2.ApproverID, JE2.JEDescription, JE2.JELineDescription FROM #JEData JE2														
                            WHERE JE2.JENumber IN (														
                                SELECT DISTINCT(JE2_1.JENumber)													
                                FROM #JEData JE2_1													
                                WHERE JE2_1.GLAccountNumber = @GLAccountNumber AND JE2_1.Debit<>0													
                                ) AND JE2.GLAccountNumber = @AL_GLAccountNumber AND JE2.Credit<>0													
                    ELSE																
                        IF (@AL_Debit_Credit='Debit') /* Credit/Debit */															
                            INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                            GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                            SELECT JE3.BusinessUnit, JE3.TMPJENumber, JE3.JENumber, JE3.JELineNumber, JE3.Year, JE3.Period, JE3.EffectiveDate, JE3.EntryDate, JE3.Debit, JE3.Credit, JE3.Amount, JE3.FunctionalCurrencyCode, 														
                            JE3.GLAccountNumber, JE3.Source, JE3.PreparerID, JE3.ApproverID, JE3.JEDescription, JE3.JELineDescription FROM #JEData JE3														
                            WHERE JE3.JENumber IN (														
                                SELECT DISTINCT(JE3_1.JENumber)													
                                FROM #JEData JE3_1													
                                WHERE JE3_1.GLAccountNumber = @GLAccountNumber AND JE3_1.Credit<>0													
                                ) AND JE3.GLAccountNumber = @AL_GLAccountNumber AND JE3.Debit<>0													
                        ELSE /* Credit/Credit */															
                            INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                            GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                            SELECT JE4.BusinessUnit, JE4.TMPJENumber, JE4.JENumber, JE4.JELineNumber, JE4.Year, JE4.Period, JE4.EffectiveDate, JE4.EntryDate, JE4.Debit, JE4.Credit, JE4.Amount, JE4.FunctionalCurrencyCode, 														
                            JE4.GLAccountNumber, JE4.Source, JE4.PreparerID, JE4.ApproverID, JE4.JEDescription, JE4.JELineDescription FROM #JEData JE4														
                            WHERE JE4.JENumber IN (														
                                SELECT DISTINCT(JE4_1.JENumber)													
                                FROM #JEData JE4_1													
                                WHERE JE4_1.GLAccountNumber = @GLAccountNumber AND JE4_1.Credit<>0													
                                ) AND JE4.GLAccountNumber = @AL_GLAccountNumber AND JE4.Credit<>0													
                    Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit																
                    END;																
                    Close cur;																
                    Deallocate cur																
                    --****************************************************Filtered Result_1***************************************************																
                    SELECT distinct 																
                        BusinessUnit AS 회사코드,															
                        TMPJENumber AS 전표번호, 															
                        JELineNumber AS 전표라인번호,															
                        Year AS 회계연도,															
                        Period AS 회계기간,															
                        EffectiveDate AS 전기일,															
                        EntryDate AS 입력일,															
                        Amount AS 금액,															
                        FunctionalCurrencyCode AS 통화,															
                        #result.GLAccountNumber AS 계정코드,															
                        COA.GLAccountName AS 계정명,															
                        Source AS 전표유형,															
                        PreparerID AS 입력자,															
                        ApproverID AS 승인자,															
                        JEDescription AS 전표헤더적요,															
                        JELineDescription AS 전표라인적요															
                    FROM #result, #COAData COA																
                    WHERE #result.GLAccountNumber = COA.GLAccountNumber																
                    ORDER BY 전표번호, 전표라인번호	
                    DROP TABLE #filter, #JEData, #result, #COAData, #JEData2																											
                '''.format(field=self.selected_project_id, cursor=cursortext, TE=self.temp_TE,
                           AutoManual=self.ManualAuto)

            elif self.rbtn2.isChecked():  ### JE
                sql = '''
                            SET NOCOUNT ON													
                            --****************************************************Filter Table***************************************************																
                            CREATE TABLE #filter																
                            (GLAccountNumber VARCHAR(100), Debit_Credit VARCHAR(100), AL_GLAccountNumber VARCHAR(100), AL_Debit_Credit VARCHAR(100))																
                            INSERT INTO #filter																
                            VALUES																
                            {cursor}																
                            --****************************************************Insert ProjectID***************************************************																
                            SELECT 																
                                JournalEntries.BusinessUnit,															
                                JournalEntries.JENumber AS TMPJENumber, 															
                                Details.JEIdentifierID AS JENumber, 															
                                JournalEntries.JELineNumber,															
                                JournalEntries.Year,															
                                JournalEntries.Period,															
                                JournalEntries.EffectiveDate,															
                                JournalEntries.EntryDate,															
                                JournalEntries.Debit,															
                                JournalEntries.Credit,															
                                JournalEntries.Amount,															
                                JournalEntries.FunctionalCurrencyCode,															
                                JournalEntries.GLAccountNumber,															
                                JournalEntries.Source,															
                                JournalEntries.PreparerID,															
                                JournalEntries.ApproverID,															
                                JournalEntries.JEDescription,															
                                JournalEntries.JELineDescription															
                                INTO #JEData															
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,																
                                [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details															
                            WHERE JournalEntries.JELINEID = Details.JENumberID 																															
                            AND ABS(JournalEntries.Amount) >= {TE}																
                            {AutoManual}															
                            SELECT  JournalEntries.BusinessUnit,																
                                    JournalEntries.JENumber,														
                                    Details.JEIdentifierID,														
                                    JournalEntries.JELineNumber,																	
                                    JournalEntries.EffectiveDate,																	
                                    JournalEntries.EntryDate,																	
                                    JournalEntries.Year,														
                                    JournalEntries.Period,																	
                                    JournalEntries.GLAccountNumber,																	
                                    JournalEntries.Source,														
                                    JournalEntries.Debit,																	
                                    JournalEntries.Credit,																	
                                    JournalEntries.Amount,																	
                                    JournalEntries.FunctionalCurrencyCode,																	
                                    JournalEntries.JEDescription,																	
                                    JournalEntries.JELineDescription,																	
                                    JournalEntries.PreparerID,																	
                                    JournalEntries.ApproverID  INTO #JEData2																	
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,																	
                                    [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details														
                                WHERE JournalEntries.JELINEID = Details.JENumberID																	
                                AND Details.JEIdentifierID IN (															
                                                                SELECT DISTINCT Details.JEIdentifierID							
                                                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,							
                                                                [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details							
                                                                WHERE JournalEntries.JELINEID = Details.JENumberID													
                                                                AND ABS(JournalEntries.Amount) >= {TE}							
                                                                {AutoManual}							
                                                              ) 																							
                            SELECT COA.GLAccountNumber,																
                                   MAX(COA.GLAccountName) AS GLAccountName, 															
                                   MAX(COA.AccountType) AS AccountType															
                                   INTO #COAData															
                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA																
                            GROUP BY COA.GLAccountNumber																                          
                            --****************************************************Result Table***************************************************																
                            CREATE TABLE #result																
                            (																
                            BusinessUnit NVARCHAR(100),																
                            TMPJENumber NVARCHAR(100), 																
                            JENumber NVARCHAR(100),	 															
                            JELineNumber BIGINT,																
                            YEAR NVARCHAR(25),																
                            Period NVARCHAR(25),																
                            EffectiveDate DATE,																
                            EntryDate DATE,																
                            Debit NUMERIC(21,6),																
                            Credit NUMERIC(21,6),																
                            Amount NUMERIC(21,6),																
                            FunctionalCurrencyCode NVARCHAR(50),																
                            GLAccountNumber NVARCHAR(100),																
                            Source NVARCHAR(100),																
                            PreparerID NVARCHAR(100),																
                            ApproverID NVARCHAR(100),																
                            JEDescription NVARCHAR(200),																
                            JELineDescription NVARCHAR(200)																
                            )																
                            --****************************************************Cursor Start***************************************************																
                            DECLARE cur CURSOR FOR 																
                            SELECT GLAccountNumber, Debit_Credit, AL_GLAccountNumber, AL_Debit_Credit FROM #filter																
                            DECLARE @GLAccountNumber VARCHAR(100)																
                            DECLARE @Debit_Credit VARCHAR(100)																
                            DECLARE @AL_GLAccountNumber VARCHAR(100)																
                            DECLARE @AL_Debit_Credit VARCHAR(100)																
                            OPEN cur																
                            Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit																
                            WHILE(@@FETCH_STATUS <> -1)																
                            BEGIN;																
                            IF (@Debit_Credit = 'Debit')																
                                IF (@AL_Debit_Credit='Debit') /* Debit/Debit */															
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                                    GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                                    SELECT JE1.BusinessUnit, JE1.TMPJENumber, JE1.JENumber, JE1.JELineNumber, JE1.Year, JE1.Period, JE1.EffectiveDate, JE1.EntryDate, JE1.Debit, JE1.Credit, JE1.Amount, JE1.FunctionalCurrencyCode, 														
                                    JE1.GLAccountNumber, JE1.Source, JE1.PreparerID, JE1.ApproverID, JE1.JEDescription, JE1.JELineDescription FROM #JEData JE1														
                                    WHERE JE1.JENumber IN (														
                                        SELECT DISTINCT(JE1_1.JENumber)													
                                        FROM #JEData JE1_1													
                                        WHERE JE1_1.GLAccountNumber = @GLAccountNumber AND JE1_1.Debit<>0													
                                        ) AND JE1.GLAccountNumber = @AL_GLAccountNumber AND JE1.Debit<>0													
                                ELSE /* Debit/Credit */															
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                                    GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                                    SELECT JE2.BusinessUnit, JE2.TMPJENumber, JE2.JENumber, JE2.JELineNumber, JE2.Year, JE2.Period, JE2.EffectiveDate, JE2.EntryDate, JE2.Debit, JE2.Credit, JE2.Amount, JE2.FunctionalCurrencyCode, 														
                                    JE2.GLAccountNumber, JE2.Source, JE2.PreparerID, JE2.ApproverID, JE2.JEDescription, JE2.JELineDescription FROM #JEData JE2														
                                    WHERE JE2.JENumber IN (														
                                        SELECT DISTINCT(JE2_1.JENumber)													
                                        FROM #JEData JE2_1													
                                        WHERE JE2_1.GLAccountNumber = @GLAccountNumber AND JE2_1.Debit<>0													
                                        ) AND JE2.GLAccountNumber = @AL_GLAccountNumber AND JE2.Credit<>0													
                            ELSE																
                                IF (@AL_Debit_Credit='Debit') /* Credit/Debit */															
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                                    GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                                    SELECT JE3.BusinessUnit, JE3.TMPJENumber, JE3.JENumber, JE3.JELineNumber, JE3.Year, JE3.Period, JE3.EffectiveDate, JE3.EntryDate, JE3.Debit, JE3.Credit, JE3.Amount, JE3.FunctionalCurrencyCode, 														
                                    JE3.GLAccountNumber, JE3.Source, JE3.PreparerID, JE3.ApproverID, JE3.JEDescription, JE3.JELineDescription FROM #JEData JE3														
                                    WHERE JE3.JENumber IN (														
                                        SELECT DISTINCT(JE3_1.JENumber)													
                                        FROM #JEData JE3_1													
                                        WHERE JE3_1.GLAccountNumber = @GLAccountNumber AND JE3_1.Credit<>0													
                                        ) AND JE3.GLAccountNumber = @AL_GLAccountNumber AND JE3.Debit<>0													
                                ELSE /* Credit/Credit */															
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, Debit, Credit, Amount, FunctionalCurrencyCode, 														
                                    GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)														
                                    SELECT JE4.BusinessUnit, JE4.TMPJENumber, JE4.JENumber, JE4.JELineNumber, JE4.Year, JE4.Period, JE4.EffectiveDate, JE4.EntryDate, JE4.Debit, JE4.Credit, JE4.Amount, JE4.FunctionalCurrencyCode, 														
                                    JE4.GLAccountNumber, JE4.Source, JE4.PreparerID, JE4.ApproverID, JE4.JEDescription, JE4.JELineDescription FROM #JEData JE4														
                                    WHERE JE4.JENumber IN (														
                                        SELECT DISTINCT(JE4_1.JENumber)													
                                        FROM #JEData JE4_1													
                                        WHERE JE4_1.GLAccountNumber = @GLAccountNumber AND JE4_1.Credit<>0													
                                        ) AND JE4.GLAccountNumber = @AL_GLAccountNumber AND JE4.Credit<>0													
                            Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit																
                            END;																
                            Close cur;																
                            Deallocate cur																
                            	--****************************************************Filtered 전표추출***************************************************																		
                            SELECT 												
                                #JEData2.BusinessUnit AS 회사코드											
                                , #JEData2.JENumber AS 전표번호											
                                , #JEData2.JELineNumber AS 전표라인번호											
                                , #JEData2.Year AS 회계연도											
                                , #JEData2.Period AS 회계기간											
                                , #JEData2.EffectiveDate AS 전기일											
                                , #JEData2.EntryDate AS 입력일											
                                , #JEData2.Amount AS 금액											
                                , #JEData2.FunctionalCurrencyCode AS 통화											
                                , #JEData2.GLAccountNumber AS 계정코드											
                                , #COAData.GLAccountName AS 계정명											
                                , #JEData2.Source AS 전표유형											
                                , #JEData2.PreparerID AS 입력자											
                                , #JEData2.ApproverID AS 승인자											
                                , #JEData2.JEDescription AS 전표헤더적요											
                                , #JEData2.JELineDescription AS 전표라인적요											
                            FROM #JEData2, #COAData												
                            WHERE #JEData2.GLAccountNumber = #COAData.GLAccountNumber																						
                            AND #JEData2.JEIdentifierID IN --JEIdentifierID												
                                    (										
                                     select distinct JENumber -- JEIdentifierID										
                                     from #result																
                                    )										
                            ORDER BY 전표번호, 전표라인번호												                                                                   								
                            DROP TABLE #filter, #JEData, #result, #COAData, #JEData2													
                                    '''.format(field=self.selected_project_id, cursor=cursortext,
                                               TE=self.temp_TE, AutoManual=self.ManualAuto)

        else:  ### 기능영역 유
            if self.rbtn1.isChecked():  ### JE Line
                sql = '''
                        SET NOCOUNT ON									
                        	--****************************************************Filter Table***************************************************																
                        CREATE TABLE #filter																
                        (GL_Functional_Area VARCHAR(100), GLAccountNumber VARCHAR(100), Debit_Credit VARCHAR(100), 																
                         AL_Functional_Area VARCHAR(100), AL_GLAccountNumber VARCHAR(100), AL_Debit_Credit VARCHAR(100))																
                        INSERT INTO #filter																
                        VALUES																
                        {cursor}
                        --****************************************************Insert ProjectID***************************************************																
                        SELECT 																
                            JournalEntries.BusinessUnit,															
                            JournalEntries.JENumber AS TMPJENumber, 															
                            Details.JEIdentifierID AS JENumber, 															
                            JournalEntries.JELineNumber,															
                            JournalEntries.Year,															
                            JournalEntries.Period,															
                            JournalEntries.EffectiveDate,															
                            JournalEntries.EntryDate,															
                            JournalEntries.Debit,															
                            JournalEntries.Credit,															
                            JournalEntries.Amount,															
                            JournalEntries.FunctionalCurrencyCode,															
                            JournalEntries.Segment01,															
                            JournalEntries.GLAccountNumber,															
                            JournalEntries.Source,															
                            JournalEntries.PreparerID,															
                            JournalEntries.ApproverID,															
                            JournalEntries.JEDescription,															
                            JournalEntries.JELineDescription															
                            INTO #JEData															
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,																
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details															
                        WHERE JournalEntries.JELINEID = Details.JENumberID 																															
                        AND ABS(JournalEntries.Amount) >= {TE}																
                        {AutoManual}																
                        SELECT  JournalEntries.BusinessUnit,																
                                JournalEntries.JENumber,														
                                Details.JEIdentifierID,														
                                JournalEntries.JELineNumber,																	
                                JournalEntries.EffectiveDate,																	
                                JournalEntries.EntryDate,																	
                                JournalEntries.Year,														
                                JournalEntries.Period,																	
                                JournalEntries.Segment01,														
                                JournalEntries.GLAccountNumber,																	
                                JournalEntries.Source,														
                                JournalEntries.Debit,																	
                                JournalEntries.Credit,																	
                                JournalEntries.Amount,																	
                                JournalEntries.FunctionalCurrencyCode,																	
                                JournalEntries.JEDescription,																	
                                JournalEntries.JELineDescription,																	
                                JournalEntries.PreparerID,																	
                                JournalEntries.ApproverID  INTO #JEData2																	
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,																	
                                [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details														
                            WHERE JournalEntries.JELINEID = Details.JENumberID																	
                            AND Details.JEIdentifierID IN (															
                                                            SELECT DISTINCT Details.JEIdentifierID							
                                                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,							
                                                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details							
                                                            WHERE JournalEntries.JELINEID = Details.JENumberID												
                                                            AND ABS(JournalEntries.Amount) >= {TE}							
                                                            {AutoManual}							
                                                           )																							
                        SELECT COA.GLAccountNumber,																
                               COA.GLAccountName AS GLAccountName, 															
                               COA.AccountType AS AccountType,															
                               COA.Segment01 AS Segment01															
                               INTO #COAData															
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA																                                                       
                        --****************************************************Result Table***************************************************																
                        CREATE TABLE #result																
                        (																
                        BusinessUnit NVARCHAR(100),																
                        TMPJENumber NVARCHAR(100), 																
                        JENumber NVARCHAR(100),	 															
                        JELineNumber BIGINT,																
                        YEAR NVARCHAR(25),																
                        Period NVARCHAR(25),																
                        EffectiveDate DATE,																
                        EntryDate DATE,																
                        Debit NUMERIC(21,6),																
                        Credit NUMERIC(21,6),																
                        Amount NUMERIC(21,6),																
                        FunctionalCurrencyCode NVARCHAR(50),																
                        Segment01 NVARCHAR(100),																
                        GLAccountNumber NVARCHAR(100),																
                        Source NVARCHAR(100),																
                        PreparerID NVARCHAR(100),																
                        ApproverID NVARCHAR(100),																
                        JEDescription NVARCHAR(200),																
                        JELineDescription NVARCHAR(200)																
                        )																                                            
                        --****************************************************Cursor Start***************************************************																
                        DECLARE cur CURSOR FOR 																
                        SELECT GL_Functional_Area, GLAccountNumber, Debit_Credit, AL_Functional_Area, AL_GLAccountNumber, AL_Debit_Credit FROM #filter																
                        DECLARE @GL_Functional_Area VARCHAR(100)																
                        DECLARE @GLAccountNumber VARCHAR(100)																
                        DECLARE @Debit_Credit VARCHAR(100)																
                        DECLARE @AL_Functional_Area VARCHAR(100)																
                        DECLARE @AL_GLAccountNumber VARCHAR(100)																
                        DECLARE @AL_Debit_Credit VARCHAR(100)																
                        OPEN cur																
                        Fetch Next From cur INTO @GL_Functional_Area, @GLAccountNumber, @Debit_Credit, @AL_Functional_Area, @AL_GLAccountNumber, @AL_Debit_Credit																
                        WHILE(@@FETCH_STATUS <> -1)																
                        BEGIN;																                                                      
                            IF (@Debit_Credit = 'Debit')															
                                IF (@AL_Debit_Credit='Debit') /* Debit/Debit */														
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                    Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                    SELECT JE1.BusinessUnit, JE1.TMPJENumber, JE1.JENumber, JE1.JELineNumber, JE1.Year, JE1.Period, JE1.EffectiveDate, JE1.EntryDate,													
                                    JE1.Debit, JE1.Credit, JE1.Amount, JE1.FunctionalCurrencyCode, JE1.Segment01, JE1.GLAccountNumber, JE1.Source, JE1.PreparerID, 													
                                    JE1.ApproverID, JE1.JEDescription, JE1.JELineDescription FROM #JEData JE1													
                                    WHERE JE1.JENumber IN 													
                                        (												
                                        SELECT DISTINCT(JE1_1.JENumber)												
                                        FROM #JEData JE1_1												
                                        WHERE JE1_1.GLAccountNumber = @GLAccountNumber AND ((JE1_1.Segment01=@GL_Functional_Area) OR (JE1_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                        AND JE1_1.Debit<>0												
                                        ) 												
                                    AND JE1.GLAccountNumber = @AL_GLAccountNumber AND ((JE1.Segment01=@AL_Functional_Area) OR (JE1.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                    AND JE1.Debit<>0													
                                ELSE /* Debit/Credit */														
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                    Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                    SELECT JE2.BusinessUnit, JE2.TMPJENumber, JE2.JENumber, JE2.JELineNumber, JE2.Year, JE2.Period, JE2.EffectiveDate, JE2.EntryDate, 													
                                    JE2.Debit, JE2.Credit, JE2.Amount, JE2.FunctionalCurrencyCode, JE2.Segment01, JE2.GLAccountNumber, JE2.Source, JE2.PreparerID, 													
                                    JE2.ApproverID, JE2.JEDescription, JE2.JELineDescription FROM #JEData JE2													
                                    WHERE JE2.JENumber IN 													
                                        (												
                                        SELECT DISTINCT(JE2_1.JENumber)												
                                        FROM #JEData JE2_1												
                                        WHERE JE2_1.GLAccountNumber = @GLAccountNumber AND ((JE2_1.Segment01=@GL_Functional_Area) OR (JE2_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                        AND JE2_1.Debit<>0												
                                        ) 												
                                    AND JE2.GLAccountNumber = @AL_GLAccountNumber AND ((JE2.Segment01=@AL_Functional_Area) OR (JE2.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                    AND JE2.Credit<>0													
                            ELSE															
                                IF (@AL_Debit_Credit='Debit') /* Credit/Debit */														
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                    Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                    SELECT JE3.BusinessUnit, JE3.TMPJENumber, JE3.JENumber, JE3.JELineNumber, JE3.Year, JE3.Period, JE3.EffectiveDate, JE3.EntryDate, 													
                                    JE3.Debit, JE3.Credit, JE3.Amount, JE3.FunctionalCurrencyCode, JE3.Segment01, JE3.GLAccountNumber, JE3.Source, JE3.PreparerID, 													
                                    JE3.ApproverID, JE3.JEDescription, JE3.JELineDescription FROM #JEData JE3													
                                    WHERE JE3.JENumber IN 													
                                        (												
                                        SELECT DISTINCT(JE3_1.JENumber)												
                                        FROM #JEData JE3_1												
                                        WHERE JE3_1.GLAccountNumber = @GLAccountNumber AND ((JE3_1.Segment01=@GL_Functional_Area) OR (JE3_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                        AND JE3_1.Credit<>0												
                                        ) 												
                                    AND JE3.GLAccountNumber = @AL_GLAccountNumber AND ((JE3.Segment01=@AL_Functional_Area) OR (JE3.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                    AND JE3.Debit<>0													
                                ELSE /* Credit/Credit */														
                                    INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                    Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                    SELECT JE4.BusinessUnit, JE4.TMPJENumber, JE4.JENumber, JE4.JELineNumber, JE4.Year, JE4.Period, JE4.EffectiveDate, JE4.EntryDate, 													
                                    JE4.Debit, JE4.Credit, JE4.Amount, JE4.FunctionalCurrencyCode, JE4.Segment01, JE4.GLAccountNumber, JE4.Source, JE4.PreparerID, 													
                                    JE4.ApproverID, JE4.JEDescription, JE4.JELineDescription FROM #JEData JE4													
                                    WHERE JE4.JENumber IN 													
                                        (												
                                        SELECT DISTINCT(JE4_1.JENumber)												
                                        FROM #JEData JE4_1												
                                        WHERE JE4_1.GLAccountNumber = @GLAccountNumber AND ((JE4_1.Segment01=@GL_Functional_Area) OR (JE4_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                        AND JE4_1.Credit<>0												
                                        ) 												
                                    AND JE4.GLAccountNumber = @AL_GLAccountNumber AND ((JE4.Segment01=@AL_Functional_Area) OR (JE4.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                    AND JE4.Credit<>0													
                        Fetch Next From cur INTO @GL_Functional_Area, @GLAccountNumber, @Debit_Credit, @AL_Functional_Area, @AL_GLAccountNumber, @AL_Debit_Credit																
                        END;																
                        Close cur;																
                        Deallocate cur																                                                                
                        --****************************************************Filtered Result_1***************************************************																
                        SELECT DISTINCT																
                            BusinessUnit AS 회사코드,															
                            TMPJENumber AS 전표번호, 															
                            JELineNumber AS 전표라인번호,															
                            Year AS 회계연도,															
                            Period AS 회계기간,															
                            EffectiveDate AS 전기일,															
                            EntryDate AS 입력일,															
                            Amount AS 금액,															
                            FunctionalCurrencyCode AS 통화,															
                            #result.Segment01 AS Segment01,															
                            #result.GLAccountNumber AS 계정코드,															
                            COA.GLAccountName AS 계정명,															
                            Source AS 전표유형,															
                            PreparerID AS 입력자,															
                            ApproverID AS 승인자,															
                            JEDescription AS 전표헤더적요,															
                            JELineDescription AS 전표라인적요															
                        FROM #result, #COAData COA																
                        WHERE CONCAT(#result.GLAccountNumber,#result.Segment01) = CONCAT(COA.GLAccountNumber,COA.Segment01)																
                        ORDER BY 전표번호, 전표라인번호																
                        DROP TABLE #filter, #JEData, #result, #COAData, #JEData2																
                                    '''.format(field=self.selected_project_id, cursor=cursortext,
                                               TE=self.temp_TE, AutoManual=self.ManualAuto)

            elif self.rbtn2.isChecked():  ### JE
                sql = '''
                SET NOCOUNT ON	
                	--****************************************************Filter Table***************************************************																
                    CREATE TABLE #filter																
                    (GL_Functional_Area VARCHAR(100), GLAccountNumber VARCHAR(100), Debit_Credit VARCHAR(100), 																
                     AL_Functional_Area VARCHAR(100), AL_GLAccountNumber VARCHAR(100), AL_Debit_Credit VARCHAR(100))																
                    INSERT INTO #filter																
                    VALUES																
                    {cursor}                                             
                    --****************************************************Insert ProjectID***************************************************																
                    SELECT 																
                        JournalEntries.BusinessUnit,															
                        JournalEntries.JENumber AS TMPJENumber, -- 전표번호															
                        Details.JEIdentifierID AS JENumber, -- JEIdentifierID를 전표 세트 식별자로 사용함.															
                        JournalEntries.JELineNumber,															
                        JournalEntries.Year,															
                        JournalEntries.Period,															
                        JournalEntries.EffectiveDate,															
                        JournalEntries.EntryDate,															
                        JournalEntries.Debit,															
                        JournalEntries.Credit,															
                        JournalEntries.Amount,															
                        JournalEntries.FunctionalCurrencyCode,															
                        JournalEntries.Segment01,															
                        JournalEntries.GLAccountNumber,															
                        JournalEntries.Source,															
                        JournalEntries.PreparerID,															
                        JournalEntries.ApproverID,															
                        JournalEntries.JEDescription,															
                        JournalEntries.JELineDescription															
                        INTO #JEData															
                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,																
                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details															
                    WHERE JournalEntries.JELINEID = Details.JENumberID 																															
                    AND ABS(JournalEntries.Amount) >= {TE}
                    {AutoManual}																                                                                
                    SELECT  JournalEntries.BusinessUnit,																
                            JournalEntries.JENumber,														
                            Details.JEIdentifierID,														
                            JournalEntries.JELineNumber,																	
                            JournalEntries.EffectiveDate,																	
                            JournalEntries.EntryDate,																	
                            JournalEntries.Year,														
                            JournalEntries.Period,																	
                            JournalEntries.Segment01,														
                            JournalEntries.GLAccountNumber,																	
                            JournalEntries.Source,														
                            JournalEntries.Debit,																	
                            JournalEntries.Credit,																	
                            JournalEntries.Amount,																	
                            JournalEntries.FunctionalCurrencyCode,																	
                            JournalEntries.JEDescription,																	
                            JournalEntries.JELineDescription,																	
                            JournalEntries.PreparerID,																	
                            JournalEntries.ApproverID  INTO #JEData2																	
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,																	
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details														
                        WHERE JournalEntries.JELINEID = Details.JENumberID																	
                        AND Details.JEIdentifierID IN (															
                                                        SELECT DISTINCT Details.JEIdentifierID							
                                                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries,							
                                                        [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details							
                                                        WHERE JournalEntries.JELINEID = Details.JENumberID													
                                                        AND ABS(JournalEntries.Amount) >= {TE}							
                                                        {AutoManual}						
                                                       )																						
                    SELECT COA.GLAccountNumber,																
                           COA.GLAccountName AS GLAccountName, 															
                           COA.AccountType AS AccountType,															
                           COA.Segment01 AS Segment01															
                           INTO #COAData															
                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA																
                    --****************************************************Result Table***************************************************																
                    CREATE TABLE #result																
                    (																
                    BusinessUnit NVARCHAR(100),																
                    TMPJENumber NVARCHAR(100), 																
                    JENumber NVARCHAR(100),	 															
                    JELineNumber BIGINT,																
                    YEAR NVARCHAR(25),																
                    Period NVARCHAR(25),																
                    EffectiveDate DATE,																
                    EntryDate DATE,																
                    Debit NUMERIC(21,6),																
                    Credit NUMERIC(21,6),																
                    Amount NUMERIC(21,6),																
                    FunctionalCurrencyCode NVARCHAR(50),																
                    Segment01 NVARCHAR(100),																
                    GLAccountNumber NVARCHAR(100),																
                    Source NVARCHAR(100),																
                    PreparerID NVARCHAR(100),																
                    ApproverID NVARCHAR(100),																
                    JEDescription NVARCHAR(200),																
                    JELineDescription NVARCHAR(200)																
                    )																
                    --****************************************************Cursor Start***************************************************																
                    DECLARE cur CURSOR FOR 																
                    SELECT GL_Functional_Area, GLAccountNumber, Debit_Credit, AL_Functional_Area, AL_GLAccountNumber, AL_Debit_Credit FROM #filter																
                    DECLARE @GL_Functional_Area VARCHAR(100)																
                    DECLARE @GLAccountNumber VARCHAR(100)																
                    DECLARE @Debit_Credit VARCHAR(100)																
                    DECLARE @AL_Functional_Area VARCHAR(100)																
                    DECLARE @AL_GLAccountNumber VARCHAR(100)																
                    DECLARE @AL_Debit_Credit VARCHAR(100)																
                    OPEN cur																
                    Fetch Next From cur INTO @GL_Functional_Area, @GLAccountNumber, @Debit_Credit, @AL_Functional_Area, @AL_GLAccountNumber, @AL_Debit_Credit																
                    WHILE(@@FETCH_STATUS <> -1)																
                    BEGIN;																                                               
                        IF (@Debit_Credit = 'Debit')															
                            IF (@AL_Debit_Credit='Debit') /* Debit/Debit */														
                                INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                SELECT JE1.BusinessUnit, JE1.TMPJENumber, JE1.JENumber, JE1.JELineNumber, JE1.Year, JE1.Period, JE1.EffectiveDate, JE1.EntryDate,													
                                JE1.Debit, JE1.Credit, JE1.Amount, JE1.FunctionalCurrencyCode, JE1.Segment01, JE1.GLAccountNumber, JE1.Source, JE1.PreparerID, 													
                                JE1.ApproverID, JE1.JEDescription, JE1.JELineDescription FROM #JEData JE1													
                                WHERE JE1.JENumber IN 													
                                    (												
                                    SELECT DISTINCT(JE1_1.JENumber)												
                                    FROM #JEData JE1_1												
                                    WHERE JE1_1.GLAccountNumber = @GLAccountNumber AND ((JE1_1.Segment01=@GL_Functional_Area) OR (JE1_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                    AND JE1_1.Debit<>0												
                                    ) 												
                                AND JE1.GLAccountNumber = @AL_GLAccountNumber AND ((JE1.Segment01=@AL_Functional_Area) OR (JE1.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                AND JE1.Debit<>0													
                            ELSE /* Debit/Credit */														
                                INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                SELECT JE2.BusinessUnit, JE2.TMPJENumber, JE2.JENumber, JE2.JELineNumber, JE2.Year, JE2.Period, JE2.EffectiveDate, JE2.EntryDate, 													
                                JE2.Debit, JE2.Credit, JE2.Amount, JE2.FunctionalCurrencyCode, JE2.Segment01, JE2.GLAccountNumber, JE2.Source, JE2.PreparerID, 													
                                JE2.ApproverID, JE2.JEDescription, JE2.JELineDescription FROM #JEData JE2													
                                WHERE JE2.JENumber IN 													
                                    (												
                                    SELECT DISTINCT(JE2_1.JENumber)												
                                    FROM #JEData JE2_1												
                                    WHERE JE2_1.GLAccountNumber = @GLAccountNumber AND ((JE2_1.Segment01=@GL_Functional_Area) OR (JE2_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                    AND JE2_1.Debit<>0												
                                    ) 												
                                AND JE2.GLAccountNumber = @AL_GLAccountNumber AND ((JE2.Segment01=@AL_Functional_Area) OR (JE2.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                AND JE2.Credit<>0													
                        ELSE															
                            IF (@AL_Debit_Credit='Debit') /* Credit/Debit */														
                                INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                SELECT JE3.BusinessUnit, JE3.TMPJENumber, JE3.JENumber, JE3.JELineNumber, JE3.Year, JE3.Period, JE3.EffectiveDate, JE3.EntryDate, 													
                                JE3.Debit, JE3.Credit, JE3.Amount, JE3.FunctionalCurrencyCode, JE3.Segment01, JE3.GLAccountNumber, JE3.Source, JE3.PreparerID, 													
                                JE3.ApproverID, JE3.JEDescription, JE3.JELineDescription FROM #JEData JE3													
                                WHERE JE3.JENumber IN 													
                                    (												
                                    SELECT DISTINCT(JE3_1.JENumber)												
                                    FROM #JEData JE3_1												
                                    WHERE JE3_1.GLAccountNumber = @GLAccountNumber AND ((JE3_1.Segment01=@GL_Functional_Area) OR (JE3_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                    AND JE3_1.Credit<>0												
                                    ) 												
                                AND JE3.GLAccountNumber = @AL_GLAccountNumber AND ((JE3.Segment01=@AL_Functional_Area) OR (JE3.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                AND JE3.Debit<>0													
                            ELSE /* Credit/Credit */														
                                INSERT INTO #result (BusinessUnit, TMPJENumber, JENumber, JELineNumber, Year, Period, EffectiveDate, EntryDate, 													
                                Debit, Credit, Amount, FunctionalCurrencyCode, Segment01, GLAccountNumber, Source, PreparerID, ApproverID, JEDescription, JELineDescription)													
                                SELECT JE4.BusinessUnit, JE4.TMPJENumber, JE4.JENumber, JE4.JELineNumber, JE4.Year, JE4.Period, JE4.EffectiveDate, JE4.EntryDate, 													
                                JE4.Debit, JE4.Credit, JE4.Amount, JE4.FunctionalCurrencyCode, JE4.Segment01, JE4.GLAccountNumber, JE4.Source, JE4.PreparerID, 													
                                JE4.ApproverID, JE4.JEDescription, JE4.JELineDescription FROM #JEData JE4													
                                WHERE JE4.JENumber IN 													
                                    (												
                                    SELECT DISTINCT(JE4_1.JENumber)												
                                    FROM #JEData JE4_1												
                                    WHERE JE4_1.GLAccountNumber = @GLAccountNumber AND ((JE4_1.Segment01=@GL_Functional_Area) OR (JE4_1.Segment01 IS NULL AND @GL_Functional_Area IS NULL)) 												
                                    AND JE4_1.Credit<>0												
                                    ) 												
                                AND JE4.GLAccountNumber = @AL_GLAccountNumber AND ((JE4.Segment01=@AL_Functional_Area) OR (JE4.Segment01 IS NULL AND @AL_Functional_Area IS NULL)) 													
                                AND JE4.Credit<>0													
                    Fetch Next From cur INTO @GL_Functional_Area, @GLAccountNumber, @Debit_Credit, @AL_Functional_Area, @AL_GLAccountNumber, @AL_Debit_Credit																
                    END;																
                    Close cur;																
                    Deallocate cur
                    	--****************************************************Filtered 전표추출***************************************************																						
                    SELECT 													
                        #JEData2.BusinessUnit AS 회사코드												
                        , #JEData2.JENumber AS 전표번호												
                        , #JEData2.JELineNumber AS 전표라인번호												
                        , #JEData2.Year AS 회계연도												
                        , #JEData2.Period AS 회계기간												
                        , #JEData2.EffectiveDate AS 전기일												
                        , #JEData2.EntryDate AS 입력일												
                        , #JEData2.Amount AS 금액												
                        , #JEData2.FunctionalCurrencyCode AS 통화												
                        , #JEData2.Segment01 AS Segment01												
                        , #JEData2.GLAccountNumber AS 계정코드												
                        , #COAData.GLAccountName AS 계정명												
                        , #JEData2.Source AS 전표유형												
                        , #JEData2.PreparerID AS 입력자												
                        , #JEData2.ApproverID AS 승인자												
                        , #JEData2.JEDescription AS 전표헤더적요												
                        , #JEData2.JELineDescription AS 전표라인적요												
                    FROM #JEData2, #COAData													
                    WHERE CONCAT(#JEData2.GLAccountNumber,#JEData2.Segment01) = CONCAT(#COAData.GLAccountNumber,#COAData.Segment01)																								
                    AND #JEData2.JEIdentifierID IN --JEIdentifierID													
                            (											
                             select distinct JENumber -- JEIdentifierID											
                             from #result																					
                            )											
                    ORDER BY 전표번호, 전표라인번호																																					
                    DROP TABLE #filter, #JEData, #result, #COAData, #JEData2	        								              
                                    '''.format(field=self.selected_project_id, cursor=cursortext,
                                               TE=self.temp_TE, AutoManual=self.ManualAuto)
        readlist = pd.read_sql(sql, self.cnxn)
        dflist.append(readlist)

        self.dataframe = pd.concat(dflist, ignore_index=True)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario08",
                                                             "---Filtered Result  Scenario08---\n" + sql]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario08",
                                                               "---Filtered JE  Scenario08---\n" + sql]

        if len(self.dataframe) == 0:
            self.communicateC.closeApp2.emit(cursortext)

        elif len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicateC.closeApp2.emit(cursortext)

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            self.communicateC.closeApp2.emit(cursortext)

    ### 쿼리문 관련 함수 (시나리오 9번)
    def extButtonClicked13(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        ### JE Line
        if self.rbtn1.isChecked():
            sql_query = '''
                                SET NOCOUNT ON				
                                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                GROUP BY CoA.GLAccountNumber				
                                SELECT				
                                    JournalEntries.BusinessUnit AS 회사코드			
                                    , JournalEntries.JENumber AS 전표번호			
                                    , JournalEntries.JELineNumber AS 전표라인번호			
                                    , JournalEntries.Year AS 회계연도			
                                    , JournalEntries.Period AS 회계기간			
                                    , JournalEntries.EffectiveDate AS 전기일			
                                    , JournalEntries.EntryDate AS 입력일			
                                    , JournalEntries.Amount AS 금액			
                                    , JournalEntries.FunctionalCurrencyCode AS 통화			
                                    , JournalEntries.GLAccountNumber AS 계정코드			
                                    , #TMPCOA.GLAccountName AS 계정명			
                                    , JournalEntries.Source AS 전표유형			
                                    , JournalEntries.PreparerID AS 입력자			
                                    , JournalEntries.ApproverID AS 승인자			
                                    , JournalEntries.JEDescription AS 전표헤더적요			
                                    , JournalEntries.JELineDescription AS 전표라인적요			
        		                    {NewSelect}  
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                    #TMPCOA,			
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                AND JournalEntries.JELINEID = Details.JENumberID 							
                                {Continuous} 		
                                AND ABS(JournalEntries.Amount) >= {TE}			
                                {Account}			
                                {NewSQL}
                                {DebitCredit}
                                {AutoManual}				
                                ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                                DROP TABLE #TMPCOA				
                                        '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                   Account=self.checked_account13,
                                                   DebitCredit=self.debitcredit,
                                                   NewSQL=self.NewSQL,
                                                   AutoManual=self.ManualAuto, NewSelect=self.NewSelect,
                                                   Continuous=self.filter_Continuous)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

            ### JE - Journals
        elif self.rbtn2.isChecked():
            sql_query = '''
                                    SET NOCOUNT ON				
                                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                                    GROUP BY CoA.GLAccountNumber				
                                    SELECT				
                                        JournalEntries.BusinessUnit AS 회사코드			
                                        , JournalEntries.JENumber AS 전표번호			
                                        , JournalEntries.JELineNumber AS 전표라인번호			
                                        , JournalEntries.Year AS 회계연도			
                                        , JournalEntries.Period AS 회계기간			
                                        , JournalEntries.EffectiveDate AS 전기일			
                                        , JournalEntries.EntryDate AS 입력일			
                                        , JournalEntries.Amount AS 금액			
                                        , JournalEntries.FunctionalCurrencyCode AS 통화			
                                        , JournalEntries.GLAccountNumber AS 계정코드			
                                        , #TMPCOA.GLAccountName AS 계정명			
                                        , JournalEntries.Source AS 전표유형			
                                        , JournalEntries.PreparerID AS 입력자			
                                        , JournalEntries.ApproverID AS 승인자			
                                        , JournalEntries.JEDescription AS 전표헤더적요			
                                        , JournalEntries.JELineDescription AS 전표라인적요			
                                        {NewSelect}		
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                        #TMPCOA,			
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                                    AND JournalEntries.JELINEID = Details.JENumberID 							
                                    AND Details.JEIdentifierID IN				
                                            (		
                                             SELECT DISTINCT Details.JEIdentifierID		
                                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                                             {Continuous}		
                                             AND ABS(JournalEntries.Amount) >= {TE}		
                                             {Account}	
                                             {NewSQL}		
                                             {DebitCredit}
                                             {AutoManual}	
                                            )		
                                    ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                                    DROP TABLE #TMPCOA				
                                            '''.format(field=self.selected_project_id, TE=self.temp_TE,
                                                       Account=self.checked_account13,
                                                       DebitCredit=self.debitcredit,
                                                       NewSQL=self.NewSQL,
                                                       AutoManual=self.ManualAuto, NewSelect=self.NewSelect,
                                                       Continuous=self.filter_Continuous)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario09",
                                                             "---Filtered Result  Scenario09---\n" + sql_query]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario09",
                                                               "---Filtered JE  Scenario09---\n" + sql_query]
        ### 예외처리 3 - 최대 추출 라인수
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate13.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[연속된 숫자: " + str(self.temp_Continuous) + ','
                                                       + '중요성금액: ' + str(self.temp_TE)
                                                       + '], 라인수 ' + str(len(self.dataframe)) + '개입니다']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate13.closeApp.emit()

        else:
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate13.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 10번)
    def extButtonClicked14(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        if self.rbtn1.isChecked():  # JE Line- Result

            sql = '''
                SET NOCOUNT ON				
                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                GROUP BY CoA.GLAccountNumber				
                SELECT				
                    JournalEntries.BusinessUnit AS 회사코드			
                    , JournalEntries.JENumber AS 전표번호			
                    , JournalEntries.JELineNumber AS 전표라인번호			
                    , JournalEntries.Year AS 회계연도			
                    , JournalEntries.Period AS 회계기간			
                    , JournalEntries.EffectiveDate AS 전기일			
                    , JournalEntries.EntryDate AS 입력일			
                    , JournalEntries.Amount AS 금액			
                    , JournalEntries.FunctionalCurrencyCode AS 통화			
                    , JournalEntries.GLAccountNumber AS 계정코드			
                    , #TMPCOA.GLAccountName AS 계정명			
                    , JournalEntries.Source AS 전표유형			
                    , JournalEntries.PreparerID AS 입력자			
                    , JournalEntries.ApproverID AS 승인자			
                    , JournalEntries.JEDescription AS 전표헤더적요			
                    , JournalEntries.JELineDescription AS 전표라인적요			
                    {NewSelect}			
                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                    #TMPCOA,			
                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                AND JournalEntries.JELINEID = Details.JENumberID 
                AND ABS(JournalEntries.Amount) >= {TE}
                {KEY}
                {Account} 		
                {NewSQL}				
                {DebitCredit}			
                {AutoManual}				
                ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                DROP TABLE #TMPCOA				
                '''.format(field=self.selected_project_id, KEY=self.tempKey, TE=self.tempTE,
                           DebitCredit=self.debitcredit,
                           Account=self.checked_account14, NewSQL=self.NewSQL,
                           NewSelect=self.NewSelect, AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        elif self.rbtn2.isChecked():  # JE- Journals

            sql = '''
                SET NOCOUNT ON				
                SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                GROUP BY CoA.GLAccountNumber				
                SELECT				
                    JournalEntries.BusinessUnit AS 회사코드			
                    , JournalEntries.JENumber AS 전표번호			
                    , JournalEntries.JELineNumber AS 전표라인번호			
                    , JournalEntries.Year AS 회계연도			
                    , JournalEntries.Period AS 회계기간			
                    , JournalEntries.EffectiveDate AS 전기일			
                    , JournalEntries.EntryDate AS 입력일			
                    , JournalEntries.Amount AS 금액			
                    , JournalEntries.FunctionalCurrencyCode AS 통화			
                    , JournalEntries.GLAccountNumber AS 계정코드			
                    , #TMPCOA.GLAccountName AS 계정명			
                    , JournalEntries.Source AS 전표유형			
                    , JournalEntries.PreparerID AS 입력자			
                    , JournalEntries.ApproverID AS 승인자			
                    , JournalEntries.JEDescription AS 전표헤더적요			
                    , JournalEntries.JELineDescription AS 전표라인적요			
                    {NewSelect}			
                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                    #TMPCOA,			
                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                AND JournalEntries.JELINEID = Details.JENumberID 				
                AND Details.JEIdentifierID IN				
                        (		
                         SELECT DISTINCT Details.JEIdentifierID		
                         FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                         WHERE JournalEntries.JELINEID = Details.JENumberID 		
                         AND ABS(JournalEntries.Amount) >= {TE} 
                         {KEY}
                         {Account} 		
                         {NewSQL}				
                         {DebitCredit}			
                         {AutoManual}		
                         )		
                ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                DROP TABLE #TMPCOA				
                '''.format(field=self.selected_project_id, KEY=self.tempKey, TE=self.tempTE,
                           DebitCredit=self.debitcredit,
                           Account=self.checked_account14, NewSQL=self.NewSQL,
                           NewSelect=self.NewSelect, AutoManual=self.ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE Line- Result 추출 시, 쿼리 저장
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario10",
                                                             "---Filtered Result  Scenario10---\n" + sql]
        ### JE- Journals 추출 시, 쿼리 저장
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario10",
                                                               "---Filtered JE  Scenario10---\n" + sql]
        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate14.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.communicate14.closeApp.emit()

        ### 0건이 아닐 경우, 시트&데이터 추출
        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate14.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 11번)
    def extButtonClicked15(self):
        cursor = self.cnxn.cursor()

        ### JE Line
        if self.rbtn1.isChecked():

            sql = '''
                SET NOCOUNT ON				
                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                    GROUP BY CoA.GLAccountNumber				
                    SELECT				
                        JournalEntries.BusinessUnit AS 회사코드			
                        , JournalEntries.JENumber AS 전표번호			
                        , JournalEntries.JELineNumber AS 전표라인번호			
                        , JournalEntries.Year AS 회계연도			
                        , JournalEntries.Period AS 회계기간			
                        , JournalEntries.EffectiveDate AS 전기일			
                        , CONVERT(CHAR(10), CONVERT(DATE, JournalEntries.UserDefined1), 23)  AS 증빙일		
                        , JournalEntries.EntryDate AS 입력일			
                        , JournalEntries.Amount AS 금액			
                        , JournalEntries.FunctionalCurrencyCode AS 통화			
                        , JournalEntries.GLAccountNumber AS 계정코드			
                        , #TMPCOA.GLAccountName AS 계정명			
                        , JournalEntries.Source AS 전표유형			
                        , JournalEntries.PreparerID AS 입력자			
                        , JournalEntries.ApproverID AS 승인자			
                        , JournalEntries.JEDescription AS 전표헤더적요			
                        , JournalEntries.JELineDescription AS 전표라인적요			
                        {NewSelect}		       
                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                        #TMPCOA,			
                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                    AND JournalEntries.JELINEID = Details.JENumberID 						
                    AND Month(JournalEntries.UserDefined1) <> Month(JournalEntries.EffectiveDate) 				
                    AND ABS(JournalEntries.Amount) >= {TE} 				
                    {Account}					
                    {NewSQL} 			
                    {AutoManual}
                    {DebitCredit}	  						
                    ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                    DROP TABLE #TMPCOA						
                '''.format(field=self.selected_project_id, TE=self.tempTE,
                           Account=self.checked_account15, NewSQL=self.NewSQL,
                           NewSelect=self.NewSelect, AutoManual=self.ManualAuto, DebitCredit=self.debitcredit)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE
        elif self.rbtn2.isChecked():

            sql = '''
                SET NOCOUNT ON				
                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                    GROUP BY CoA.GLAccountNumber				
                    SELECT				
                        JournalEntries.BusinessUnit AS 회사코드			
                        , JournalEntries.JENumber AS 전표번호			
                        , JournalEntries.JELineNumber AS 전표라인번호			
                        , JournalEntries.Year AS 회계연도			
                        , JournalEntries.Period AS 회계기간			
                        , JournalEntries.EffectiveDate AS 전기일			
                        , CONVERT(CHAR(10), CONVERT(DATE, JournalEntries.UserDefined1), 23)  AS 증빙일		
                        , JournalEntries.EntryDate AS 입력일			
                        , JournalEntries.Amount AS 금액			
                        , JournalEntries.FunctionalCurrencyCode AS 통화			
                        , JournalEntries.GLAccountNumber AS 계정코드			
                        , #TMPCOA.GLAccountName AS 계정명			
                        , JournalEntries.Source AS 전표유형			
                        , JournalEntries.PreparerID AS 입력자			
                        , JournalEntries.ApproverID AS 승인자			
                        , JournalEntries.JEDescription AS 전표헤더적요			
                        , JournalEntries.JELineDescription AS 전표라인적요			
                        {NewSelect}			 
                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                        #TMPCOA,			
                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                    AND JournalEntries.JELINEID = Details.JENumberID 					
                    AND Details.JEIdentifierID IN				
                            (		
                             SELECT DISTINCT Details.JEIdentifierID		
                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                             WHERE JournalEntries.JELINEID = Details.JENumberID 		
                             AND Month(JournalEntries.UserDefined1) <> Month(JournalEntries.EffectiveDate) 		
                             AND ABS(JournalEntries.Amount) >= {TE} 	
                             {Account} 	
                             {NewSQL}
                             {AutoManual}
                             {DebitCredit}	
                            )		
                    ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                    DROP TABLE #TMPCOA						
                '''.format(field=self.selected_project_id, TE=self.tempTE,
                           Account=self.checked_account15, NewSQL=self.NewSQL,
                           NewSelect=self.NewSelect, AutoManual=self.ManualAuto, DebitCredit=self.debitcredit)
            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario11",
                                                             "---Filtered Result  Scenario11---\n" + sql]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario11",
                                                               "---Filtered JE  Scenario11---\n" + sql]

        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate15.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.communicate15.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate15.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 12번)
    def extButtonClicked16(self):
        cursor = self.cnxn.cursor()

        ### JE Line 추출
        if self.rbtn1.isChecked():

            if self.debitcredit != '':

                sql = '''
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber;			
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}			
                                        
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 				
                        
                        {Account}			
                        {Date}
                        {NewSQL}					
                        {DebitCredit}								
                        {AutoManual}	
                        AND (				
                             SELECT SUM(ABS(JournalEntries1.Amount))			
                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                  [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                             WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                             AND Details1.JEIdentifierID = Details.JEIdentifierID	
                             {SubAccount}
                             {SubDate}
                             {SubNewSQL}
                             {SubDebitCredit}
                             {SubAutoManual}			 
                             GROUP BY Details1.JEIdentifierID			
                            ) >= {TE}	-- 중요성 금액(이상으로)			
                                        
                        ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA										
                        '''.format(field=self.selected_project_id, Account=self.checked_account16, TE=self.temp_TE, Date=self.EntryDate,
                                   NewSQL=self.NewSQL, DebitCredit=self.debitcredit, NewSelect=self.NewSelect, AutoManual=self.ManualAuto,
                                   SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate, SubNewSQL=self.sub_NewSQL,
                                   SubDebitCredit=self.sub_debitcredit, SubAutoManual=self.sub_ManualAuto)

            else:

                sql = """
                		SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber;
                        			
                        (
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}	
                                        
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 				
                        
                        {Account}			
                        {Date}
                        {NewSQL}												
                        {AutoManual}
                        AND JournalEntries.Credit = 0
                        AND (				
                             SELECT SUM(ABS(JournalEntries1.Amount))			
                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                  [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                             WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                             AND Details1.JEIdentifierID = Details.JEIdentifierID	
                             {SubAccount}
                             {SubDate}
                             {SubNewSQL}
                             {SubAutoManual}
                             AND JournalEntries1.Credit = 0					 
                             GROUP BY Details1.JEIdentifierID			
                            ) >= {TE}	-- 중요성 금액(이상으로)			
                                        
                        )			
                        Union
                        (
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}	
                                        
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                        
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 				
                        
                        {Account}			
                        {Date}
                        {NewSQL}												
                        {AutoManual}
                        AND JournalEntries.Debit = 0				
                        
                        AND (				
                             SELECT SUM(ABS(JournalEntries1.Amount))			
                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                  [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                             WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                             AND Details1.JEIdentifierID = Details.JEIdentifierID	
                             {SubAccount}
                             {SubDate}
                             {SubNewSQL}
                             {SubAutoManual}
                             AND JournalEntries1.Debit = 0						 
                             GROUP BY Details1.JEIdentifierID			
                            ) >= {TE}	-- 중요성 금액(이상으로)			
                                        
                        )
                        ORDER BY 전표번호, 전표라인번호
                        DROP TABLE #TMPCOA		
                        """.format(field=self.selected_project_id, Account=self.checked_account16, TE=self.temp_TE, Date=self.EntryDate,
                                   NewSQL=self.NewSQL, NewSelect=self.NewSelect, AutoManual=self.ManualAuto,
                                   SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate, SubNewSQL=self.sub_NewSQL,
                                   SubAutoManual=self.sub_ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE 추출
        elif self.rbtn2.isChecked():

            if self.debitcredit != '':
                sql = '''
                        SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber;			
                        SELECT				
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}			
    
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                             [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
    
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 	
                        AND Details.JEIdentifierID IN
                            (			
                             SELECT DISTINCT Details.JEIdentifierID		
                             FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                             WHERE JournalEntries.JELINEID = Details.JENumberID 	
                             {Account}			
                             {Date}
                             {NewSQL}					
                             {DebitCredit}								
                             {AutoManual}	
                             AND (				
                                 SELECT SUM(ABS(JournalEntries1.Amount))			
                                 FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                      [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                 WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                 AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                 {SubAccount}
                                 {SubDate}
                                 {SubNewSQL}
                                 {SubDebitCredit}
                                 {SubAutoManual}			 
                                 GROUP BY Details1.JEIdentifierID			
                                 ) >= {TE}	-- 중요성 금액(이상으로)			
                            )
                        ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                        DROP TABLE #TMPCOA										
                        '''.format(field=self.selected_project_id, Account=self.checked_account16, TE=self.temp_TE, Date=self.EntryDate,
                                   NewSQL=self.NewSQL, DebitCredit=self.debitcredit, NewSelect=self.NewSelect, AutoManual=self.ManualAuto,
                                   SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate, SubNewSQL=self.sub_NewSQL,
                                   SubDebitCredit=self.sub_debitcredit, SubAutoManual=self.sub_ManualAuto)
            else:
                sql = """
                		SET NOCOUNT ON				
                        SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                        GROUP BY CoA.GLAccountNumber;			
                        
                        SELECT
                            JournalEntries.BusinessUnit AS 회사코드			
                            , JournalEntries.JENumber AS 전표번호			
                            , JournalEntries.JELineNumber AS 전표라인번호			
                            , JournalEntries.Year AS 회계연도			
                            , JournalEntries.Period AS 회계기간			
                            , JournalEntries.EffectiveDate AS 전기일			
                            , JournalEntries.EntryDate AS 입력일			
                            , JournalEntries.Amount AS 금액			
                            , JournalEntries.FunctionalCurrencyCode AS 통화			
                            , JournalEntries.GLAccountNumber AS 계정코드			
                            , #TMPCOA.GLAccountName AS 계정명			
                            , JournalEntries.Source AS 전표유형			
                            , JournalEntries.PreparerID AS 입력자			
                            , JournalEntries.ApproverID AS 승인자			
                            , JournalEntries.JEDescription AS 전표헤더적요			
                            , JournalEntries.JELineDescription AS 전표라인적요			
                            {NewSelect}	
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                            #TMPCOA,			
                            [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                    
                        WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                        AND JournalEntries.JELINEID = Details.JENumberID 				
                                    
                        AND Details.JEIdentifierID IN
                            (
                                (
                                SELECT DISTINCT Details.JEIdentifierID				
                                                
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                
                                WHERE JournalEntries.JELINEID = Details.JENumberID 				
                                
                                {Account}			
                                {Date}
                                {NewSQL}												
                                {AutoManual}
                                AND JournalEntries.Credit = 0
                                AND (				
                                     SELECT SUM(ABS(JournalEntries1.Amount))			
                                     FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                          [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                     WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                     AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                     {SubAccount}
                                     {SubDate}
                                     {SubNewSQL}
                                     {SubAutoManual}
                                     AND JournalEntries1.Credit = 0					 
                                     GROUP BY Details1.JEIdentifierID			
                                    ) >= {TE}	-- 중요성 금액(이상으로)			 
                                )			
                            Union
                                (
                                SELECT DISTINCT Details.JEIdentifierID		
                                                
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,			
                                     [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                                                
                                WHERE JournalEntries.JELINEID = Details.JENumberID 				
                                
                                {Account}			
                                {Date}
                                {NewSQL}												
                                {AutoManual}
                                AND JournalEntries.Debit = 0				
                                
                                AND (				
                                     SELECT SUM(ABS(JournalEntries1.Amount))			
                                     FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries1,			
                                          [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details1		
                                     WHERE JournalEntries1.JELINEID = Details1.JENumberID 			
                                     AND Details1.JEIdentifierID = Details.JEIdentifierID	
                                     {SubAccount}
                                     {SubDate}
                                     {SubNewSQL}
                                     {SubAutoManual}
                                     AND JournalEntries1.Debit = 0						 
                                     GROUP BY Details1.JEIdentifierID			
                                    ) >= {TE}	-- 중요성 금액(이상으로)			        
                                )
                            )
                        ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
		                DROP TABLE #TMPCOA			
                        """.format(field=self.selected_project_id, Account=self.checked_account16, TE=self.temp_TE, Date=self.EntryDate,
                                   NewSQL=self.NewSQL, NewSelect=self.NewSelect, AutoManual=self.ManualAuto,
                                   SubAccount=self.sub_checked_account16, SubDate=self.subEntryDate, SubNewSQL=self.sub_NewSQL,
                                   SubAutoManual=self.sub_ManualAuto)

            self.dataframe = pd.read_sql(sql, self.cnxn)

        ### JE Line 추출 시, 쿼리 저장
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario12",
                                                             "---Filtered Result  Scenario12---\n" + sql]

        ### JE 추출 시, 쿼리 저장
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario12",
                                                               "---Filtered JE  Scenario12---\n" + sql]

        ### 50만건 초과 추출 시, 상위 1000건 반환
        if len(self.dataframe) > 500000:
            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate16.closeApp.emit()

        ### 0건 추출 시, 문구 반환
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': ["중요성금액: " + str(self.temp_TE)
                             + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate16.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate16.closeApp.emit()

    ### 쿼리문 관련 함수 (시나리오 13번)
    def extButtonClicked17(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        ### JE Line
        if self.rbtn1.isChecked():
            sql_query = """
                    SET NOCOUNT ON				
                    SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                    FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                    GROUP BY CoA.GLAccountNumber				
                    SELECT				
                        JournalEntries.BusinessUnit AS 회사코드			
                        , JournalEntries.JENumber AS 전표번호			
                        , JournalEntries.JELineNumber AS 전표라인번호			
                        , JournalEntries.Year AS 회계연도			
                        , JournalEntries.Period AS 회계기간			
                        , JournalEntries.EffectiveDate AS 전기일			
                        , JournalEntries.EntryDate AS 입력일			
                        , JournalEntries.Amount AS 금액			
                        , JournalEntries.FunctionalCurrencyCode AS 통화			
                        , JournalEntries.GLAccountNumber AS 계정코드			
                        , #TMPCOA.GLAccountName AS 계정명			
                        , JournalEntries.Source AS 전표유형			
                        , JournalEntries.PreparerID AS 입력자			
                        , JournalEntries.ApproverID AS 승인자			
                        , JournalEntries.JEDescription AS 전표헤더적요			
                        , JournalEntries.JELineDescription AS 전표라인적요			
                        {NewSelect}
                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                        #TMPCOA,			
                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                    WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                    AND JournalEntries.JELINEID = Details.JENumberID 							
                    AND JournalEntries.PreparerID = JournalEntries.ApproverID			
                    AND ABS(JournalEntries.Amount) >= {TE} 		
                    {Account}				
                    {NewSQL}				
                    {DebitCredit}
                    {AutoManual}				
                    ORDER BY JournalEntries.JENumber,JournalEntries.JELineNumber				
                    DROP TABLE #TMPCOA				            
            """.format(field=self.selected_project_id, TE=self.temp_TE, Account=self.checked_account17,
                       DebitCredit=self.debitcredit, NewSQL=self.NewSQL,
                       AutoManual=self.ManualAuto, NewSelect=self.NewSelect)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### JE
        elif self.rbtn2.isChecked():
            sql_query = """
                            SET NOCOUNT ON				
                            SELECT CoA.GLAccountNumber, MAX(CoA.GLAccountName) AS GLAccountName INTO #TMPCOA				
                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA				
                            GROUP BY CoA.GLAccountNumber				
                            SELECT				
                                JournalEntries.BusinessUnit AS 회사코드			
                                , JournalEntries.JENumber AS 전표번호			
                                , JournalEntries.JELineNumber AS 전표라인번호			
                                , JournalEntries.Year AS 회계연도			
                                , JournalEntries.Period AS 회계기간			
                                , JournalEntries.EffectiveDate AS 전기일			
                                , JournalEntries.EntryDate AS 입력일			
                                , JournalEntries.Amount AS 금액			
                                , JournalEntries.FunctionalCurrencyCode AS 통화			
                                , JournalEntries.GLAccountNumber AS 계정코드			
                                , #TMPCOA.GLAccountName AS 계정명			
                                , JournalEntries.Source AS 전표유형			
                                , JournalEntries.PreparerID AS 입력자			
                                , JournalEntries.ApproverID AS 승인자			
                                , JournalEntries.JEDescription AS 전표헤더적요			
                                , JournalEntries.JELineDescription AS 전표라인적요			
                                {NewSelect}		
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,				
                                #TMPCOA,			
                                 [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details			
                            WHERE JournalEntries.GLAccountNumber = #TMPCOA.GLAccountNumber 				
                            AND JournalEntries.JELINEID = Details.JENumberID 							
                            AND Details.JEIdentifierID IN				
                                    (		
                                     SELECT DISTINCT Details.JEIdentifierID		
                                     FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,		
                                         [{field}_Reporting_Details_CY_01].[dbo].[JournalEntries] AS Details	
                                     WHERE JournalEntries.JELINEID = Details.JENumberID 
                                     AND JournalEntries.PreparerID = JournalEntries.ApproverID 
                                     AND ABS(JournalEntries.Amount) >= {TE}	
                                     {Account}	
                                     {NewSQL}			
                                     {DebitCredit}
                                     {AutoManual}			
                                        )	
                            ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                            DROP TABLE #TMPCOA				           
                        """.format(field=self.selected_project_id, TE=self.temp_TE, Account=self.checked_account17,
                                   DebitCredit=self.debitcredit, NewSQL=self.NewSQL,
                                   AutoManual=self.ManualAuto, NewSelect=self.NewSelect)

            self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario13",
                                                             "---Filtered Result  Scenario13---\n" + sql_query]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario13",
                                                               "---Filtered JE  Scenario13---\n" + sql_query]

        ### 예외처리 3 - 최대 추출 라인수
        if len(self.dataframe) > 500000:

            model = DataFrameModel(self.dataframe.head(1000))
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe.head(1000)
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate17.closeApp.emit()


        elif len(self.dataframe) == 0:

            self.dataframe = pd.DataFrame({'No Data': ["[중요성금액: " + str(self.temp_TE)
                                                       + '], 라인수 ' + str(len(self.dataframe)) + ' 개입니다']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate17.closeApp.emit()


        else:

            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate17.closeApp.emit()

    ### DataFrameModel 클래스와 연결하는 함수
    @pyqtSlot(QModelIndex)
    def slot_clicked_item(self, QModelIndex):
        self.stk_w.setCurrentIndex(QModelIndex.row())

    ### 프로그램을 종료하고자 할 때, 뜨는 팝업
    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Quit', '정말 종료하시겠습니까?\n현재 ' + str(len(self.combo_sheet)) + '개의 시트가 있습니다',
                                     QMessageBox.No | QMessageBox.Yes)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    ### 사용자가 파일을 저장하고자 할 때 실행되는 함수
    def saveFile(self):
        self.scenario_dic['JEA_Query'] = self.my_query
        ### 예외처리 1 - 데이터 프레임 빈 경우
        if self.dataframe is None:
            self.MessageBox_Open("저장할 데이터가 없습니다.")
            return
        ### 예외처리 2 - 딕셔너리 빈 경우
        if self.scenario_dic == {}:
            self.MessageBox_Open("저장할 Sheet가 없습니다.")
            return

        else:
            fileName = QFileDialog.getSaveFileName(self, self.tr("Save Data files"), "./",
                                                   self.tr("xlsx(*.xlsx);; All Files(*.*)"))
            path = fileName[0]
            ### 신규 저장
            if path == '':
                pass
            ### 덮어쓰기
            else:
                if os.path.isfile(path):
                    changecount = 0
                    addcount = 0
                    wb = openpyxl.load_workbook(path)
                    wb.create_sheet('Scenario Updated>>>')
                    ws_names = wb.get_sheet_names()
                    query_sheet = wb.get_sheet_by_name('JEA_Query')

                    for temp in list(self.scenario_dic.keys()):
                        if temp in ws_names:
                            changecount += 1
                            wb.remove(wb['' + temp + ''])
                        else:
                            addcount += 1

                    real_query = pd.DataFrame(query_sheet.values)
                    real_query.columns = ["Sheet name", "Scenario number", "Query"]

                    if changecount == 1:
                        self.scenario_dic['JEA_Query'] = pd.concat([real_query, self.scenario_dic['JEA_Query']])
                        self.scenario_dic['JEA_Query'] = self.scenario_dic['JEA_Query'].drop(
                            self.scenario_dic['JEA_Query'].index[0])

                    wb.save(path)

                    with pd.ExcelWriter('' + path + '', mode='a', engine='openpyxl') as writer:
                        for temp in self.scenario_dic:
                            self.scenario_dic['' + temp + ''].to_excel(writer, sheet_name='' + temp + '', index=False,
                                                                       freeze_panes=(1, 0))

                    query_wb = openpyxl.load_workbook(path)
                    sht = query_wb.get_sheet_by_name('JEA_Query')
                    query_wb.move_sheet(sht, addcount)
                    sht.sheet_properties.tabColor = '00FFFF'
                    query_wb.save(path)
                    self.MessageBox_Open("총 " + str(changecount) + "개 시트가 교체\n" + str(addcount) + "개 시트가 추가되었습니다.")

                else:
                    with pd.ExcelWriter('' + path + '', mode='w', engine='openpyxl') as writer:
                        for temp in self.scenario_dic:
                            self.scenario_dic['' + temp + ''].to_excel(writer, sheet_name='' + temp + '', index=False,
                                                                       freeze_panes=(1, 0))

                    query_wb_origin = openpyxl.load_workbook(path)
                    sht_origin = query_wb_origin.get_sheet_by_name('JEA_Query')
                    sht_origin.sheet_properties.tabColor = '00FFFF'
                    query_wb_origin.move_sheet(sht_origin, len(self.scenario_dic) - 1)

                    query_wb_origin.save(path)

                    self.MessageBox_Open("저장을 완료했습니다.")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
